import io
import unittest

from openpyxl import load_workbook

import report_builder
from routers import hosts as hosts_router
from routers.hosts import InspectExcelRequest, export_inspect_excel


EXPECTED_INSPECT_HEADERS = [
    "分组", "状态", "服务器IP", "服务器", "主机名", "负责人", "机房",
    "CPU使用率(%)", "CPU核心", "内存使用率(%)", "内存总量(GB)",
    "负载1m", "负载5m", "负载15m", "运行时长",
    "磁盘具体占用",
    "网络入(MB/s)", "网络出(MB/s)", "磁盘读(MB/s)", "磁盘写(MB/s)",
    "TCP连接", "TIME_WAIT", "进程占用Top10", "异常项",
]


def _fallback_result() -> dict:
    return {
        "instance": "10.0.0.8:ssh",
        "ip": "10.0.0.8",
        "hostname": "app-1",
        "os": "Linux 6.1",
        "cpu_cores": 4,
        "memory_gb": 8,
        "state": "fallback",
        "overall": "warning",
        "metrics_source": "ssh_python",
        "metrics": {
            "cpu_usage": 12.5,
            "mem_usage": 64.2,
            "mem_total_gb": 8,
            "load1": 0.4,
            "load5": 0.5,
            "load15": 0.6,
            "uptime_seconds": 3661,
            "net_recv_mbps": 1.0,
            "net_send_mbps": 2.0,
            "disk_read_mbps": 3.0,
            "disk_write_mbps": 4.0,
            "tcp_estab": 12,
            "tcp_tw": 3,
        },
        "partitions": [
            {"mountpoint": "/", "usage_pct": 80, "used_gb": 80, "total_gb": 100},
        ],
        "checks": [
            {"item": "CPU", "value": "12.5%", "status": "normal", "threshold": "<80%"},
            {"item": "磁盘 /", "value": "80%", "status": "warning", "threshold": "<80%"},
        ],
    }


class HostInspectionReportTests(unittest.IsolatedAsyncioTestCase):
    async def test_collect_inspect_data_uses_ssh_python_fallback_when_prometheus_missing(self):
        old_load_hosts = report_builder.load_hosts_list
        old_inspect_hosts = report_builder.prom.inspect_hosts
        old_score = report_builder.analyzer.calculate_host_health_score
        old_fallback = hosts_router._collect_inspection_fallbacks
        old_processes = report_builder._collect_process_top10
        old_groups = report_builder.load_groups

        async def fake_inspect_hosts(instances=None):
            return []

        async def fake_score(summary):
            return 88

        async def fake_fallback(hosts):
            return {"10.0.0.8": {"ok": True, "result": _fallback_result()}}

        async def fake_processes(hosts):
            return {"10.0.0.8": {"data": [{"pid": 1, "comm": "java", "service": "Java", "cpu": 12, "mem": 8}], "error": ""}}

        report_builder.load_hosts_list = lambda: [{
            "ip": "10.0.0.8", "hostname": "app-1", "group": "g1",
            "role": "应用服务器", "owner": "张三", "datacenter": "上海一号机房",
        }]
        report_builder.load_groups = lambda: [{"id": "g1", "name": "核心应用组"}]
        report_builder.prom.inspect_hosts = fake_inspect_hosts
        report_builder.analyzer.calculate_host_health_score = fake_score
        hosts_router._collect_inspection_fallbacks = fake_fallback
        report_builder._collect_process_top10 = fake_processes
        try:
            data = await report_builder.collect_inspect_data()
        finally:
            report_builder.load_hosts_list = old_load_hosts
            report_builder.load_groups = old_groups
            report_builder.prom.inspect_hosts = old_inspect_hosts
            report_builder.analyzer.calculate_host_health_score = old_score
            hosts_router._collect_inspection_fallbacks = old_fallback
            report_builder._collect_process_top10 = old_processes

        self.assertEqual(data["summary"]["metrics_fallback_count"], 1)
        self.assertEqual(data["summary"]["metrics_missing_count"], 0)
        self.assertEqual(data["results"][0]["metrics_source"], "ssh_python")
        self.assertEqual(data["all_hosts"][0]["cpu_cores"], 4)
        self.assertEqual(data["all_hosts"][0]["load1"], 0.4)
        self.assertEqual(data["all_hosts"][0]["tcp_tw"], 3)
        self.assertEqual(data["all_hosts"][0]["disk_write"], 4.0)
        self.assertEqual(data["all_hosts"][0]["group_name"], "核心应用组")
        self.assertEqual(data["all_hosts"][0]["owner"], "张三")
        self.assertEqual(data["all_hosts"][0]["process_top10"][0]["service"], "Java")
        self.assertEqual(data["group_sections"][0]["group_name"], "核心应用组")

    async def test_inspect_excel_uses_runtime_status_columns(self):
        response = await export_inspect_excel(
            InspectExcelRequest(
                results=[_fallback_result()],
                summary={
                    "total": 1,
                    "normal": 0,
                    "warning": 1,
                    "critical": 0,
                    "metrics_updated_count": 1,
                    "metrics_fallback_count": 1,
                    "metrics_missing_count": 0,
                },
            )
        )

        wb = load_workbook(io.BytesIO(response.body))
        ws = wb["全部主机明细"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, EXPECTED_INSPECT_HEADERS)
        values = [cell.value for cell in ws[2]]
        self.assertEqual(values[0], "未分组")
        self.assertEqual(values[1], "警告")
        self.assertEqual(values[2], "10.0.0.8")
        self.assertIn("/ 80%", values[15])
        self.assertIn("磁盘 /: 80%", values[23])


if __name__ == "__main__":
    unittest.main()
