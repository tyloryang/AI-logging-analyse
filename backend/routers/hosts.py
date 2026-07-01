"""CMDB 主机管理（手动录入）+ 主机巡检路由。

路由前缀：/api/hosts/*
"""
import json
import logging
import os
import shlex
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote

import asyncio
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Depends
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from auth.deps import current_user
from auth.models import User

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

from state import (
    prom, analyzer,
    REPORTS_DIR,
    load_hosts_list, save_hosts_list,
    load_groups, save_groups,
    encrypt_password, decrypt_password,
    load_credentials, save_credentials,
    get_user_allowed_groups,
    load_user_groups, save_user_groups,
)
from ssh_utils import ssh_connect_options

logger = logging.getLogger(__name__)
router = APIRouter()

_NOW = lambda: datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def _filter_hosts_by_user(hosts: list[dict], user: User) -> list[dict]:
    """超管可见全部；普通用户只能看自己分组内的主机。"""
    if user.is_superuser:
        return hosts
    allowed = get_user_allowed_groups(user.id)
    if allowed is None:
        return hosts  # 兜底：超管标记缺失时放行
    # allowed 为空列表 → 只能看无分组主机；有值 → 只能看对应分组
    return [h for h in hosts if h.get("group", "") in allowed]


def _env_float(name: str, default: float) -> float:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        value = float(raw)
        return value if value > 0 else default
    except Exception:
        logger.warning("[hosts] 环境变量 %s=%r 不是有效数字，使用默认值 %s", name, raw, default)
        return default


_HOST_SYNC_SSH_CONNECT_TIMEOUT = _env_float("HOST_SYNC_SSH_CONNECT_TIMEOUT", 12.0)
_HOST_SYNC_SSH_COMMAND_TIMEOUT = _env_float("HOST_SYNC_SSH_COMMAND_TIMEOUT", 20.0)
_HOST_INSPECT_FALLBACK_CONCURRENCY = max(1, int(_env_float("HOST_INSPECT_FALLBACK_CONCURRENCY", 5.0)))

# ── 主机 CRUD ──────────────────────────────────────────────────────────────────

class HostCreateRequest(BaseModel):
    hostname:      str = ""   # 可选，留空时自动用 IP 填充
    ip:            str
    platform:      str = "Linux"        # Linux / Windows / Network / Other
    os_version:    str = ""
    cpu_cores:     Optional[int] = None
    memory_gb:     Optional[float] = None
    disk_gb:       Optional[float] = None
    status:        str = "active"       # active / offline / maintenance
    env:           str = "production"   # production / staging / development / testing / dr
    role:          str = ""
    owner:         str = ""
    datacenter:    str = ""
    group:         str = ""
    ssh_port:      int = 22
    ssh_user:      str = ""
    ssh_password:  str = ""
    credential_id: str = ""
    notes:         str = ""
    labels:        dict = {}


class HostUpdateRequest(BaseModel):
    hostname:      Optional[str] = None
    ip:            Optional[str] = None
    platform:      Optional[str] = None
    os_version:    Optional[str] = None
    cpu_cores:     Optional[int] = None
    memory_gb:     Optional[float] = None
    disk_gb:       Optional[float] = None
    status:        Optional[str] = None
    env:           Optional[str] = None
    role:          Optional[str] = None
    owner:         Optional[str] = None
    datacenter:    Optional[str] = None
    group:         Optional[str] = None
    ssh_port:      Optional[int] = None
    ssh_user:      Optional[str] = None
    ssh_password:  Optional[str] = None
    credential_id: Optional[str] = None
    notes:         Optional[str] = None
    labels:        Optional[dict] = None


class BatchCredentialRequest(BaseModel):
    host_ids:      Optional[list] = None   # None = 全部可见主机
    group:         Optional[str] = None    # 按分组过滤（与 host_ids 互斥）
    credential_id: Optional[str] = None
    ssh_user:      Optional[str] = None
    ssh_port:      Optional[int] = None
    ssh_password:  Optional[str] = None


@router.post("/api/hosts/batch-credential")
async def batch_apply_credential(body: BatchCredentialRequest, user: User = Depends(current_user)):
    """批量为主机应用 SSH 凭证或密码。"""
    if not body.credential_id and not body.ssh_password:
        raise HTTPException(status_code=400, detail="必须提供凭证 ID 或 SSH 密码")

    all_hosts = load_hosts_list()
    visible = _filter_hosts_by_user(all_hosts, user)

    if body.host_ids:
        visible = [h for h in visible if h.get("id") in set(body.host_ids)]
    elif body.group:
        visible = [h for h in visible if h.get("group") == body.group]

    if not visible:
        return {"ok": True, "updated": 0}

    ids_to_update = {h["id"] for h in visible}
    now = _NOW()
    updated = 0
    for h in all_hosts:
        if h.get("id") not in ids_to_update:
            continue
        if body.credential_id:
            h["credential_id"] = body.credential_id
            h["ssh_password"] = ""
            creds = {c["id"]: c for c in load_credentials()}
            cred = creds.get(body.credential_id, {})
            if not body.ssh_user:
                h["ssh_user"] = cred.get("username", h.get("ssh_user") or "root")
            if not body.ssh_port:
                h["ssh_port"] = cred.get("port", h.get("ssh_port") or 22)
        else:
            h["ssh_password"] = encrypt_password(body.ssh_password)
            h["credential_id"] = ""
        if body.ssh_user:
            h["ssh_user"] = body.ssh_user
        if body.ssh_port:
            h["ssh_port"] = body.ssh_port
        h["updated_at"] = now
        updated += 1

    save_hosts_list(all_hosts)
    return {"ok": True, "updated": updated}


@router.get("/api/hosts")
async def list_hosts(user: User = Depends(current_user)):
    """获取主机列表（非管理员只能看自己分组内的主机）。"""
    hosts = _filter_hosts_by_user(load_hosts_list(), user)
    credentials = {c.get("id"): c for c in load_credentials()}
    # 脱敏：不返回加密密码
    safe = []
    for h in hosts:
        entry = dict(h)
        credential_id = entry.get("credential_id", "")
        entry["ssh_saved"] = bool(entry.pop("ssh_password", "")) or bool(credential_id)
        if credential_id and credential_id in credentials:
            cred = credentials[credential_id]
            entry["credential_name"] = cred.get("name", credential_id)
            entry["credential_username"] = cred.get("username", "")
            entry["credential_port"] = cred.get("port", 22)
        safe.append(entry)
    return {"data": safe, "total": len(safe)}


@router.post("/api/hosts")
async def create_host(body: HostCreateRequest):
    """新增主机。"""
    hosts = load_hosts_list()
    # 校验 IP 唯一性
    if any(h.get("ip") == body.ip for h in hosts):
        raise HTTPException(status_code=409, detail=f"IP {body.ip} 已存在")
    now = _NOW()
    entry: dict = {
        "id":           str(uuid.uuid4()),
        "hostname":     body.hostname.strip() or body.ip,  # 留空时用 IP
        "ip":           body.ip,
        "platform":     body.platform,
        "os_version":   body.os_version,
        "cpu_cores":    body.cpu_cores,
        "memory_gb":    body.memory_gb,
        "disk_gb":      body.disk_gb,
        "status":       body.status,
        "env":          body.env,
        "role":         body.role,
        "owner":        body.owner,
        "datacenter":   body.datacenter,
        "group":        body.group,
        "ssh_port":     body.ssh_port,
        "ssh_user":     body.ssh_user,
        "ssh_password": "" if body.credential_id else (encrypt_password(body.ssh_password) if body.ssh_password else ""),
        "credential_id": body.credential_id,
        "notes":        body.notes,
        "labels":       body.labels,
        "created_at":   now,
        "updated_at":   now,
    }
    hosts.append(entry)
    save_hosts_list(hosts)
    result = dict(entry)
    result["ssh_saved"] = bool(result.pop("ssh_password", "")) or bool(result.get("credential_id"))
    return result


@router.put("/api/hosts/{host_id}")
async def update_host(host_id: str, body: HostUpdateRequest):
    """更新主机信息。"""
    hosts = load_hosts_list()
    idx = next((i for i, h in enumerate(hosts) if h.get("id") == host_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="主机不存在")
    host = hosts[idx]
    fields = ["hostname", "ip", "platform", "os_version", "cpu_cores", "memory_gb",
              "disk_gb", "status", "env", "role", "owner", "datacenter", "group",
              "ssh_port", "ssh_user", "credential_id", "notes", "labels"]
    for f in fields:
        val = getattr(body, f)
        if val is not None:
            host[f] = val
    if host.get("credential_id"):
        host["ssh_password"] = ""
    elif body.ssh_password:
        host["ssh_password"] = encrypt_password(body.ssh_password) if body.ssh_password else ""
    host["updated_at"] = _NOW()
    hosts[idx] = host
    save_hosts_list(hosts)
    result = dict(host)
    result["ssh_saved"] = bool(result.pop("ssh_password", "")) or bool(result.get("credential_id"))
    return result


@router.delete("/api/hosts/{host_id}")
async def delete_host(host_id: str):
    """删除主机。"""
    hosts = load_hosts_list()
    new_hosts = [h for h in hosts if h.get("id") != host_id]
    if len(new_hosts) == len(hosts):
        raise HTTPException(status_code=404, detail="主机不存在")
    save_hosts_list(new_hosts)
    return {"ok": True}


# ── 导出 / 导入 ──────────────────────────────────────────────────────────────

_EXPORT_COLS = [
    ("id",            "主机ID"),
    ("hostname",      "主机名"),
    ("ip",            "IP 地址"),
    ("platform",      "平台"),
    ("os_version",    "操作系统版本"),
    ("cpu_cores",     "CPU 核心数"),
    ("memory_gb",     "内存(GB)"),
    ("disk_gb",       "磁盘(GB)"),
    ("status",        "状态"),
    ("env",           "环境"),
    ("role",          "用途/角色"),
    ("owner",         "负责人"),
    ("datacenter",    "机房/区域"),
    ("group",         "分组ID"),
    ("group_name",    "分组名称"),
    ("ssh_port",      "SSH 端口"),
    ("ssh_user",      "SSH 用户名"),
    ("credential_id", "SSH 凭证ID"),
    ("ssh_password",  "SSH 密码密文"),
    ("notes",         "备注"),
    ("labels",        "标签(JSON)"),
    ("created_at",    "创建时间"),
    ("updated_at",    "更新时间"),
]
_EXPORT_COL_WIDTHS = {
    "id": 38,
    "hostname": 18,
    "ip": 16,
    "platform": 12,
    "os_version": 22,
    "cpu_cores": 12,
    "memory_gb": 12,
    "disk_gb": 12,
    "status": 10,
    "env": 10,
    "role": 18,
    "owner": 14,
    "datacenter": 18,
    "group": 38,
    "group_name": 18,
    "ssh_port": 10,
    "ssh_user": 14,
    "credential_id": 38,
    "ssh_password": 40,
    "notes": 28,
    "labels": 28,
    "created_at": 22,
    "updated_at": 22,
}
_ENV_LABEL = {"production":"生产","staging":"预发","development":"开发","testing":"测试","dr":"容灾"}
_STATUS_LABEL = {"active":"在线","offline":"离线","maintenance":"维护中"}
_HEADER_CLEANUP_TABLE = str.maketrans("", "", " _-()（）[]{}:/：\t\r\n*＊,，.。")


def _normalize_import_header(value: str) -> str:
    return str(value or "").strip().translate(_HEADER_CLEANUP_TABLE).lower()


_IMPORT_COL_MAP = {}
for _field, _label in _EXPORT_COLS:
    _IMPORT_COL_MAP[_normalize_import_header(_label)] = _field
    _IMPORT_COL_MAP[_normalize_import_header(_field)] = _field
_IMPORT_COL_MAP.update({
    _normalize_import_header("分组"): "group_name",
    _normalize_import_header("group"): "group_name",
    _normalize_import_header("group_id"): "group",
    _normalize_import_header("分组名"): "group_name",
    _normalize_import_header("分组名称"): "group_name",
    _normalize_import_header("分组不存在自动创建"): "group_name",  # 模板列名
    _normalize_import_header("分组不存在则自动创建"): "group_name",
    _normalize_import_header("ssh密码明文自动加密并建凭据"): "ssh_password",  # 模板列名
    _normalize_import_header("标签"): "labels",
    # IP / 主机名 常见别名
    _normalize_import_header("ip"): "ip",
    _normalize_import_header("ip地址"): "ip",
    _normalize_import_header("地址"): "ip",
    _normalize_import_header("address"): "ip",
    _normalize_import_header("hostname"): "hostname",
    _normalize_import_header("名称"): "hostname",
    _normalize_import_header("主机"): "hostname",
    # SSH 用户名 / 密码 常见别名（模板用的中文描述也能识别）
    _normalize_import_header("用户名"): "ssh_user",
    _normalize_import_header("ssh用户"): "ssh_user",
    _normalize_import_header("user"): "ssh_user",
    _normalize_import_header("username"): "ssh_user",
    _normalize_import_header("ssh密码"): "ssh_password",
    _normalize_import_header("ssh密码明文"): "ssh_password",
    _normalize_import_header("ssh密码明文自动加密并建凭据"): "ssh_password",
    _normalize_import_header("密码"): "ssh_password",
    _normalize_import_header("password"): "ssh_password",
    _normalize_import_header("port"): "ssh_port",
    _normalize_import_header("端口"): "ssh_port",
    _normalize_import_header("ssh端口"): "ssh_port",
    # 业务字段别名
    _normalize_import_header("角色"): "role",
    _normalize_import_header("用途"): "role",
    _normalize_import_header("应用"): "role",
    _normalize_import_header("application"): "role",
    _normalize_import_header("owner"): "owner",
    _normalize_import_header("负责人"): "owner",
    _normalize_import_header("机房"): "datacenter",
    _normalize_import_header("region"): "datacenter",
    _normalize_import_header("dc"): "datacenter",
    _normalize_import_header("env"): "env",
    _normalize_import_header("环境"): "env",
    _normalize_import_header("status"): "status",
    _normalize_import_header("状态"): "status",
})


def _normalize_text(value) -> str:
    return str(value or "").strip()


def _parse_optional_int(value) -> Optional[int]:
    text = _normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _parse_optional_float(value) -> Optional[float]:
    text = _normalize_text(value)
    if not text:
        return None
    try:
        return round(float(text), 1)
    except (TypeError, ValueError):
        return None


def _parse_labels_cell(value) -> dict:
    text = _normalize_text(value)
    if not text:
        return {}
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return {
                str(k).strip(): "" if v is None else str(v)
                for k, v in parsed.items()
                if str(k).strip()
            }
    except Exception:
        pass

    labels: dict[str, str] = {}
    for line in text.splitlines():
        item = line.strip().rstrip(",;")
        if not item:
            continue
        if "=" in item:
            key, raw_val = item.split("=", 1)
        elif ":" in item:
            key, raw_val = item.split(":", 1)
        else:
            key, raw_val = item, ""
        key = key.strip()
        if key:
            labels[key] = raw_val.strip()
    if labels:
        return labels
    raise ValueError("标签需为 JSON 对象，或按每行 key=value 填写")


def _normalize_imported_ssh_password(value) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    if text.startswith("gAAAA"):
        return text
    try:
        decrypt_password(text)
        return text
    except Exception:
        return encrypt_password(text)


def _resolve_import_group(
    entry: dict,
    group_ids: set[str],
    group_name_map: dict[str, str],
    auto_create: bool = True,
    groups: list[dict] | None = None,
) -> tuple[Optional[str], Optional[str], Optional[dict]]:
    """解析导入行的分组归属。返回 (group_id, error, new_group_dict)。
    auto_create=True 时，分组名不存在则自动创建一个新分组（不报错）。
    """
    has_group_id = "group" in entry
    has_group_name = "group_name" in entry
    if not has_group_id and not has_group_name:
        return None, None, None

    raw_group_id = _normalize_text(entry.get("group", "")) if has_group_id else ""
    raw_group_name = _normalize_text(entry.get("group_name", "")) if has_group_name else ""

    if raw_group_id:
        if raw_group_id in group_ids:
            return raw_group_id, None, None
        resolved = group_name_map.get(raw_group_id.casefold())
        if resolved:
            return resolved, None, None
        return raw_group_id, None, None

    if raw_group_name:
        resolved = group_name_map.get(raw_group_name.casefold())
        if resolved:
            return resolved, None, None
        if auto_create:
            new_id = "grp_" + uuid.uuid4().hex[:8]
            new_group = {"id": new_id, "name": raw_group_name, "description": "Excel 导入自动创建"}
            if groups is not None:
                groups.append(new_group)
            group_ids.add(new_id)
            group_name_map[raw_group_name.casefold()] = new_id
            return new_id, None, new_group
        return None, f"分组「{raw_group_name}」不存在", None

    return "", None, None


def _resolve_import_credential(
    entry: dict,
    credentials: list[dict],
    cred_ids: set[str],
    cred_signature_map: dict[str, str],
    auto_create: bool = True,
) -> tuple[Optional[str], Optional[dict]]:
    """根据 ssh_user + ssh_password 自动匹配或创建凭据。
    返回 (credential_id, new_credential_dict)。已显式提供 credential_id 时直接返回。
    """
    cred_id = _normalize_text(entry.get("credential_id", ""))
    if cred_id and cred_id in cred_ids:
        return cred_id, None

    user = _normalize_text(entry.get("ssh_user", ""))
    raw_pw = _normalize_text(entry.get("ssh_password", ""))
    port = _parse_optional_int(entry.get("ssh_port")) or 22
    if not user or not raw_pw:
        return cred_id or None, None

    # 通过 user@port + 密码指纹去重（密文/明文统一指纹）
    plain_pw = raw_pw
    if raw_pw.startswith("gAAAA"):
        try:
            plain_pw = decrypt_password(raw_pw)
        except Exception:
            plain_pw = raw_pw
    import hashlib
    sig = f"{user}@{port}#" + hashlib.sha1(plain_pw.encode("utf-8", "ignore")).hexdigest()[:12]

    if sig in cred_signature_map:
        return cred_signature_map[sig], None

    if not auto_create:
        return cred_id or None, None

    new_id = "cred_" + uuid.uuid4().hex[:8]
    new_cred = {
        "id": new_id,
        "name": f"{user}@:{port} (导入自建)",
        "username": user,
        "port": port,
        "password": encrypt_password(plain_pw),
        "description": "Excel 导入自动创建",
        "created_at": _NOW(),
    }
    credentials.append(new_cred)
    cred_ids.add(new_id)
    cred_signature_map[sig] = new_id
    return new_id, new_cred


@router.get("/api/hosts/export")
async def export_hosts(user: User = Depends(current_user)):
    """导出主机 Excel（非管理员只能导出自己分组的主机）。"""
    hosts = _filter_hosts_by_user(load_hosts_list(), user)
    groups = load_groups()
    group_map = {g["id"]: g["name"] for g in groups}

    wb = Workbook()
    ws = wb.active
    ws.title = "主机列表"

    header_fill = PatternFill("solid", fgColor="FF1F3A5F")
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    border_side = Side(style="thin", color="FFCCCCCC")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = [label for _, label in _EXPORT_COLS]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border

    # 列宽
    for i, (field, _) in enumerate(_EXPORT_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = _EXPORT_COL_WIDTHS.get(field, 16)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, h in enumerate(hosts, 2):
        for col_idx, (field, _) in enumerate(_EXPORT_COLS, 1):
            if field == "group_name":
                val = group_map.get(h.get("group", ""), "")
            elif field == "labels":
                val = json.dumps(h.get("labels") or {}, ensure_ascii=False, sort_keys=True)
            else:
                val = h.get(field, "")
            if field == "status":
                val = _STATUS_LABEL.get(val, val)
            elif field == "env":
                val = _ENV_LABEL.get(val, val)
            c = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
            c.alignment = Alignment(vertical="center", wrap_text=field in {"notes", "labels", "ssh_password"})
            c.border = cell_border
            c.font = Font(size=9)
        ws.row_dimensions[row_idx].height = 16

    # 说明 sheet
    ws2 = wb.create_sheet("填写说明")
    tips = [
        ("字段", "可选值 / 说明"),
        ("主机ID", "备份回滚时建议保留；留空会自动生成"),
        ("主机名", "可选，留空自动用 IP"),
        ("IP 地址", "必填，不能重复"),
        ("平台", "Linux / Windows / Network / Other"),
        ("状态", "在线 / 离线 / 维护中"),
        ("环境", "生产 / 预发 / 开发 / 测试 / 容灾"),
        ("分组ID", "优先使用，适合完整备份恢复；即使当前系统暂无该分组也会保留"),
        ("分组名称", "便于人工编辑；导入时会自动匹配到现有分组 ID"),
        ("SSH 端口", "默认 22"),
        ("SSH 凭证ID", "填写凭证库中的凭证 ID；有值时会覆盖主机密码"),
        ("SSH 密码密文", "导出的是已加密密文，适合备份；导入明文时也会自动加密"),
        ("标签(JSON)", "支持 JSON 对象，或每行 key=value"),
        ("创建时间/更新时间", "可保留导出值用于完整恢复；留空则自动补当前时间"),
    ]
    for r, (k, v) in enumerate(tips, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws2.cell(row=r, column=2, value=v).font = Font(size=10)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"主机CMDB_{now_str}.xlsx"
    encoded = quote(filename, encoding="utf-8")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def _parse_import_rows(content: bytes, filename: str) -> list[dict]:
    """统一解析 xlsx/csv 上传文件 → 字典行列表，原始列名保留。"""
    from openpyxl import load_workbook
    import csv

    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return list(csv.DictReader(text.splitlines()))
    if lower.endswith((".xlsx", ".xls")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        col_names = [str(h).strip() if h else "" for h in headers_row]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for i, v in enumerate(row):
                if i < len(col_names) and col_names[i]:
                    d[col_names[i]] = (str(v).strip() if v is not None else "")
            rows.append(d)
        wb.close()
        return rows
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


def _process_import(
    rows: list[dict],
    groups: list[dict],
    credentials: list[dict],
    hosts: list[dict],
    *,
    conflict: str = "skip",
    auto_create_group: bool = True,
    auto_create_credential: bool = True,
    dry_run: bool = False,
) -> dict:
    """核心导入处理：分组/凭据自动创建、未知列入 labels、IP 去重。
    dry_run=True 时不修改原列表，仅返回预览统计。
    """
    now = _NOW()
    group_ids = {str(g.get("id", "")).strip() for g in groups if str(g.get("id", "")).strip()}
    group_name_map = {
        str(g.get("name", "")).strip().casefold(): str(g.get("id", "")).strip()
        for g in groups if str(g.get("name", "")).strip() and str(g.get("id", "")).strip()
    }
    cred_ids = {str(c.get("id", "")).strip() for c in credentials if str(c.get("id", "")).strip()}

    # 凭据指纹：username@port#sha1(pw)
    import hashlib
    cred_signature_map: dict[str, str] = {}
    for c in credentials:
        u = str(c.get("username", "")).strip()
        p = c.get("port") or 22
        raw_pw = str(c.get("password", "")).strip()
        if not u or not raw_pw:
            continue
        try:
            plain = decrypt_password(raw_pw) if raw_pw.startswith("gAAAA") else raw_pw
        except Exception:
            plain = raw_pw
        sig = f"{u}@{p}#" + hashlib.sha1(plain.encode("utf-8", "ignore")).hexdigest()[:12]
        cred_signature_map[sig] = c.get("id", "")

    existing_ips = {h.get("ip", ""): i for i, h in enumerate(hosts)}
    existing_ids = {h.get("id", ""): i for i, h in enumerate(hosts) if h.get("id")}
    rev_status = {v: k for k, v in _STATUS_LABEL.items()}
    rev_env = {v: k for k, v in _ENV_LABEL.items()}

    added = skipped = updated = errors = 0
    new_groups: list[dict] = []
    new_credentials: list[dict] = []
    error_details: list[str] = []
    preview_added: list[dict] = []
    preview_updated: list[dict] = []

    for row_num, raw in enumerate(rows, 2):
        # 列名映射：已知列 → 字段；未知列 → 标签
        entry: dict = {}
        extra_labels: dict = {}
        for raw_key, val in raw.items():
            field = _IMPORT_COL_MAP.get(_normalize_import_header(raw_key))
            if field:
                entry[field] = _normalize_text(val)
            elif raw_key and _normalize_text(val):
                extra_labels[str(raw_key).strip()] = _normalize_text(val)

        ip = _normalize_text(entry.get("ip", ""))
        if not ip:
            # 整行空跳过，不计入错误
            if not any(_normalize_text(v) for v in raw.values()):
                continue
            error_details.append(f"第 {row_num} 行：IP 地址为空，跳过")
            errors += 1
            continue

        if "id" in entry:
            entry["id"] = _normalize_text(entry.get("id", ""))
            if not entry["id"]:
                entry.pop("id", None)

        if "hostname" in entry:
            entry["hostname"] = _normalize_text(entry.get("hostname", "")) or ip
        if "platform" in entry:
            entry["platform"] = _normalize_text(entry.get("platform", "")) or "Linux"
        if "status" in entry:
            raw_status = _normalize_text(entry.get("status", ""))
            entry["status"] = rev_status.get(raw_status, raw_status or "active")
        if "env" in entry:
            raw_env = _normalize_text(entry.get("env", ""))
            entry["env"] = rev_env.get(raw_env, raw_env or "production")
        if "cpu_cores" in entry:
            entry["cpu_cores"] = _parse_optional_int(entry.get("cpu_cores"))
        if "memory_gb" in entry:
            entry["memory_gb"] = _parse_optional_float(entry.get("memory_gb"))
        if "disk_gb" in entry:
            entry["disk_gb"] = _parse_optional_float(entry.get("disk_gb"))
        if "ssh_port" in entry:
            parsed_port = _parse_optional_int(entry.get("ssh_port"))
            entry["ssh_port"] = parsed_port if parsed_port and parsed_port > 0 else 22
        if "ssh_user" in entry:
            entry["ssh_user"] = _normalize_text(entry.get("ssh_user", ""))
        if "credential_id" in entry:
            entry["credential_id"] = _normalize_text(entry.get("credential_id", ""))
        if "ssh_password" in entry:
            entry["ssh_password"] = _normalize_imported_ssh_password(entry.get("ssh_password"))
            if entry["ssh_password"] and "credential_id" not in entry:
                entry["credential_id"] = ""
        if "labels" in entry:
            try:
                entry["labels"] = _parse_labels_cell(entry.get("labels"))
            except ValueError as exc:
                error_details.append(f"第 {row_num} 行：{exc}")
                errors += 1
                continue
        # 未知列名合并入 labels（关键能力：Excel 任意附加列不丢失）
        if extra_labels:
            current_labels = entry.get("labels") if isinstance(entry.get("labels"), dict) else {}
            entry["labels"] = {**current_labels, **extra_labels}

        group_value, group_error, _new_group = _resolve_import_group(
            entry, group_ids, group_name_map,
            auto_create=auto_create_group, groups=groups,
        )
        entry.pop("group_name", None)
        if group_error:
            error_details.append(f"第 {row_num} 行：{group_error}")
            errors += 1
            continue
        if group_value is not None:
            entry["group"] = group_value
        if _new_group:
            new_groups.append(_new_group)

        # 凭据自动匹配/创建：根据 user+port+password 指纹
        cred_value, _new_cred = _resolve_import_credential(
            entry, credentials, cred_ids, cred_signature_map,
            auto_create=auto_create_credential,
        )
        if cred_value is not None:
            entry["credential_id"] = cred_value
            entry["ssh_password"] = ""   # 绑定了凭据就不再单独存密码
        if _new_cred:
            new_credentials.append(_new_cred)

        if ip in existing_ips:
            if conflict == "update":
                idx = existing_ips[ip]
                old_id = hosts[idx].get("id", "")
                incoming_id = entry.get("id", "")
                if incoming_id:
                    duplicated_idx = existing_ids.get(incoming_id)
                    if duplicated_idx is not None and duplicated_idx != idx:
                        error_details.append(f"第 {row_num} 行：主机 ID {incoming_id} 已被其他主机占用")
                        errors += 1
                        continue

                merged = dict(hosts[idx])
                merged.update(entry)
                merged["ip"] = ip
                merged["hostname"] = _normalize_text(merged.get("hostname", "")) or ip
                merged["platform"] = _normalize_text(merged.get("platform", "")) or "Linux"
                merged["status"] = _normalize_text(merged.get("status", "")) or "active"
                merged["env"] = _normalize_text(merged.get("env", "")) or "production"
                merged["ssh_port"] = _parse_optional_int(merged.get("ssh_port")) or 22
                if merged.get("credential_id"):
                    merged["ssh_password"] = ""
                else:
                    merged["ssh_password"] = _normalize_text(merged.get("ssh_password", ""))
                if not isinstance(merged.get("labels"), dict):
                    merged["labels"] = {}
                merged["created_at"] = _normalize_text(merged.get("created_at", "")) or hosts[idx].get("created_at") or now
                if "updated_at" not in entry:
                    merged["updated_at"] = now
                else:
                    merged["updated_at"] = _normalize_text(merged.get("updated_at", "")) or now
                merged["id"] = _normalize_text(merged.get("id", "")) or old_id or str(uuid.uuid4())
                if not dry_run:
                    hosts[idx] = merged
                    if old_id and old_id != merged["id"]:
                        existing_ids.pop(old_id, None)
                    existing_ids[merged["id"]] = idx
                preview_updated.append({
                    "ip": ip, "hostname": merged["hostname"],
                    "group": merged.get("group", ""), "env": merged.get("env", ""),
                })
                updated += 1
            else:
                skipped += 1
        else:
            entry.setdefault("id", str(uuid.uuid4()))
            if entry["id"] in existing_ids:
                error_details.append(f"第 {row_num} 行：主机 ID {entry['id']} 已存在")
                errors += 1
                continue
            new_entry = {
                "id": entry["id"],
                "hostname": ip,
                "ip": ip,
                "platform": "Linux",
                "os_version": "",
                "cpu_cores": None,
                "memory_gb": None,
                "disk_gb": None,
                "status": "active",
                "env": "production",
                "role": "",
                "owner": "",
                "datacenter": "",
                "group": "",
                "ssh_port": 22,
                "ssh_user": "",
                "ssh_password": "",
                "credential_id": "",
                "notes": "",
                "labels": {},
                "created_at": now,
                "updated_at": now,
            }
            new_entry.update(entry)
            new_entry["hostname"] = _normalize_text(new_entry.get("hostname", "")) or ip
            new_entry["platform"] = _normalize_text(new_entry.get("platform", "")) or "Linux"
            new_entry["status"] = _normalize_text(new_entry.get("status", "")) or "active"
            new_entry["env"] = _normalize_text(new_entry.get("env", "")) or "production"
            new_entry["ssh_port"] = _parse_optional_int(new_entry.get("ssh_port")) or 22
            if new_entry.get("credential_id"):
                new_entry["ssh_password"] = ""
            else:
                new_entry["ssh_password"] = _normalize_text(new_entry.get("ssh_password", ""))
            if not isinstance(new_entry.get("labels"), dict):
                new_entry["labels"] = {}
            new_entry["created_at"] = _normalize_text(new_entry.get("created_at", "")) or now
            new_entry["updated_at"] = _normalize_text(new_entry.get("updated_at", "")) or now
            if not dry_run:
                hosts.append(new_entry)
                existing_ips[ip] = len(hosts) - 1
                existing_ids[new_entry["id"]] = len(hosts) - 1
            preview_added.append({
                "ip": ip, "hostname": new_entry["hostname"],
                "group": new_entry.get("group", ""), "env": new_entry.get("env", ""),
            })
            added += 1

    if not dry_run:
        if new_groups:
            save_groups(groups)
        if new_credentials:
            save_credentials(credentials)
        save_hosts_list(hosts)

    return {
        "ok": True,
        "dry_run": dry_run,
        "total": len(rows),
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "new_groups": [{"id": g["id"], "name": g["name"]} for g in new_groups],
        "new_credentials": [{"id": c["id"], "name": c["name"]} for c in new_credentials],
        "preview_added": preview_added[:30],
        "preview_updated": preview_updated[:30],
        "error_details": error_details[:30],
        "message": (
            f"{'预览' if dry_run else '导入完成'}：新增 {added} 台，更新 {updated} 台，"
            f"跳过 {skipped} 台，错误 {errors} 条；"
            f"自动创建 {len(new_groups)} 个分组、{len(new_credentials)} 个凭据"
        ),
    }


@router.post("/api/hosts/import")
async def import_hosts(
    file: UploadFile = File(...),
    conflict: str = Query("skip", description="重复 IP 策略：skip=跳过 / update=覆盖"),
    auto_create_group: bool = Query(True, description="分组名不存在时自动创建"),
    auto_create_credential: bool = Query(True, description="ssh_user+password 自动建凭据"),
):
    """从 Excel/CSV 批量导入主机，分组/凭据自动创建，未知列入 labels。"""
    content = await file.read()
    rows = _parse_import_rows(content, file.filename or "")
    return _process_import(
        rows,
        groups=load_groups(),
        credentials=load_credentials(),
        hosts=load_hosts_list(),
        conflict=conflict,
        auto_create_group=auto_create_group,
        auto_create_credential=auto_create_credential,
        dry_run=False,
    )


@router.post("/api/hosts/import/preview")
async def import_hosts_preview(
    file: UploadFile = File(...),
    conflict: str = Query("skip"),
    auto_create_group: bool = Query(True),
    auto_create_credential: bool = Query(True),
):
    """预览导入结果（不入库），返回将新增/更新的主机及自动创建的分组/凭据列表。"""
    content = await file.read()
    rows = _parse_import_rows(content, file.filename or "")
    # 用副本避免污染当前状态
    return _process_import(
        rows,
        groups=[dict(g) for g in load_groups()],
        credentials=[dict(c) for c in load_credentials()],
        hosts=[dict(h) for h in load_hosts_list()],
        conflict=conflict,
        auto_create_group=auto_create_group,
        auto_create_credential=auto_create_credential,
        dry_run=True,
    )


@router.get("/api/hosts/template")
async def download_template():
    """下载预填好的 Excel 模板（含示例行 + 字段说明 + 下拉校验）。"""
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "主机导入模板"

    # 简化模板：只保留高频字段，IP 必填；其余可选
    template_cols = [
        ("ip",            "IP 地址 *",   "192.168.1.10",   16, ""),
        ("hostname",      "主机名",       "web-server-01",  20, ""),
        ("group_name",    "分组",         "生产-Web",        18, ""),
        ("env",           "环境",         "生产",            10, "生产,预发,开发,测试,容灾"),
        ("status",        "状态",         "在线",            10, "在线,离线,维护中"),
        ("platform",      "平台",         "Linux",          12, "Linux,Windows,Network,Other"),
        ("role",          "用途/角色",    "Web 服务",        18, ""),
        ("owner",         "负责人",       "ops",            12, ""),
        ("datacenter",    "机房/区域",    "cn-shanghai-1",  16, ""),
        ("ssh_user",      "SSH 用户名",   "root",           12, ""),
        ("ssh_port",      "SSH 端口",     "22",             10, ""),
        ("ssh_password",  "SSH 密码",     "your-password",  20, ""),
        ("notes",         "备注",         "示例：核心服务",   24, ""),
    ]

    header_fill = PatternFill("solid", fgColor="FFD97757")
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    border_side = Side(style="thin", color="FFCCCCCC")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col_idx, (field, label, sample, width, options) in enumerate(template_cols, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 第 2 行写示例
        ex = ws.cell(row=2, column=col_idx, value=sample)
        ex.font = Font(size=9, color="FF999999", italic=True)
        ex.border = cell_border

        # 下拉校验
        if options:
            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
            dv.add(f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}1000")
            ws.add_data_validation(dv)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # 说明 sheet
    ws2 = wb.create_sheet("使用说明")
    tips = [
        ("说明", "内容"),
        ("最小必填", "仅 IP 地址必填，其他字段可全部留空（系统自动补默认值）"),
        ("分组", "填分组名称即可。分组不存在会自动创建，无需先去分组管理建分组"),
        ("凭据", "填 SSH 用户名 + 密码，系统自动加密并创建凭据，相同 user@port+密码会复用现有凭据"),
        ("附加字段", "Excel 任意附加列（如「业务线」「应用名」）会自动放进主机标签，不会丢失"),
        ("批量更新", "重复 IP 默认跳过，需要更新可在导入页选「覆盖」策略"),
        ("示例行", "第 2 行的灰色斜体是示例，导入前请删除或覆盖"),
        ("中英文均可", "列名识别中文、英文、大小写、空格、下划线、括号都没问题"),
    ]
    for r, (k, v) in enumerate(tips, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=(r == 1), size=10)
        ws2.cell(row=r, column=2, value=v).font = Font(size=10)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "CMDB主机导入模板.xlsx"
    encoded = quote(filename, encoding="utf-8")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


class BatchUpdateRequest(BaseModel):
    host_ids: list[str]
    fields: dict   # 任意字段：group/env/status/role/owner/datacenter/labels...


@router.post("/api/hosts/batch-update")
async def batch_update_hosts(body: BatchUpdateRequest, user: User = Depends(current_user)):
    """批量更新选中主机的任意字段（list 视图浮动栏使用）。"""
    if not body.host_ids:
        raise HTTPException(400, "至少选中一个主机")
    if not body.fields:
        raise HTTPException(400, "未提供要更新的字段")

    allowed_fields = {
        "platform", "os_version", "status", "env", "role", "owner",
        "datacenter", "group", "labels", "notes",
    }
    fields = {k: v for k, v in body.fields.items() if k in allowed_fields}
    if not fields:
        raise HTTPException(400, "字段名不在允许范围")

    hosts = load_hosts_list()
    target_ids = set(body.host_ids)
    now = _NOW()
    updated = 0
    for h in hosts:
        if h.get("id") not in target_ids:
            continue
        if "labels" in fields and isinstance(fields["labels"], dict):
            cur = h.get("labels") if isinstance(h.get("labels"), dict) else {}
            h["labels"] = {**cur, **fields["labels"]}
            applied = {k: v for k, v in fields.items() if k != "labels"}
            h.update(applied)
        else:
            h.update(fields)
        h["updated_at"] = now
        updated += 1
    save_hosts_list(hosts)
    return {"ok": True, "updated": updated}


# ── SSH 同步系统信息 ─────────────────────────────────────────────────────────

_SYNC_SCRIPT_SHELL = r"""
OS=$(grep -m1 PRETTY_NAME /etc/os-release 2>/dev/null | cut -d'"' -f2)
[ -z "$OS" ] && OS=$(lsb_release -d 2>/dev/null | awk -F'\t' '{print $2}')
[ -z "$OS" ] && OS=$(uname -sr 2>/dev/null)
CPU=$(nproc 2>/dev/null)
[ -z "$CPU" ] && CPU=$(grep -c '^processor' /proc/cpuinfo 2>/dev/null)
[ -z "$CPU" ] && CPU=0
MEM_KB=$(awk '/MemTotal/{print $2}' /proc/meminfo 2>/dev/null)
[ -z "$MEM_KB" ] && MEM_KB=0
DISK_KB=$(df / 2>/dev/null | awk 'NR==2{print $2}')
[ -z "$DISK_KB" ] && DISK_KB=0
HN=$(hostname -s 2>/dev/null)
[ -z "$HN" ] && HN=$(hostname 2>/dev/null)

cpu_snapshot() {
  awk '/^cpu /{print $2" "$3" "$4" "$5" "$6" "$7" "$8" "$9}' /proc/stat 2>/dev/null
}
net_snapshot() {
  awk 'NR>2 {gsub(":","",$1); if ($1!="lo") {rx+=$2; tx+=$10}} END{print rx+0" "tx+0}' /proc/net/dev 2>/dev/null
}
diskio_snapshot() {
  awk '$3 ~ /^(sd[a-z]+|vd[a-z]+|xvd[a-z]+|hd[a-z]+|nvme[0-9]+n[0-9]+)$/ {r+=$6; w+=$10} END{print (r+0)*512" "(w+0)*512}' /proc/diskstats 2>/dev/null
}

CPU_A=$(cpu_snapshot)
NET_A=$(net_snapshot)
DIO_A=$(diskio_snapshot)
sleep 1
CPU_B=$(cpu_snapshot)
NET_B=$(net_snapshot)
DIO_B=$(diskio_snapshot)

CPU_USAGE=$(awk -v a="$CPU_A" -v b="$CPU_B" 'BEGIN{
  split(a,x); split(b,y);
  idle1=x[4]+x[5]; idle2=y[4]+y[5];
  for(i=1;i<=8;i++){total1+=x[i]; total2+=y[i]}
  dt=total2-total1; di=idle2-idle1;
  if(dt>0) printf "%.1f", (dt-di)*100/dt; else print "";
}')
MEM_USAGE=$(awk '/MemTotal/{t=$2}/MemAvailable/{a=$2}END{if(t>0) printf "%.1f", (t-a)*100/t; else print ""}' /proc/meminfo 2>/dev/null)
LOAD1=$(awk '{print $1}' /proc/loadavg 2>/dev/null)
LOAD5=$(awk '{print $2}' /proc/loadavg 2>/dev/null)
LOAD15=$(awk '{print $3}' /proc/loadavg 2>/dev/null)
UPTIME_SECONDS=$(awk '{printf "%.0f", $1}' /proc/uptime 2>/dev/null)
NET_RX_BPS=$(awk -v a="$NET_A" -v b="$NET_B" 'BEGIN{split(a,x); split(b,y); v=y[1]-x[1]; print v>0?v:0}')
NET_TX_BPS=$(awk -v a="$NET_A" -v b="$NET_B" 'BEGIN{split(a,x); split(b,y); v=y[2]-x[2]; print v>0?v:0}')
DIO_READ_BPS=$(awk -v a="$DIO_A" -v b="$DIO_B" 'BEGIN{split(a,x); split(b,y); v=y[1]-x[1]; print v>0?v:0}')
DIO_WRITE_BPS=$(awk -v a="$DIO_A" -v b="$DIO_B" 'BEGIN{split(a,x); split(b,y); v=y[2]-x[2]; print v>0?v:0}')

if command -v ss >/dev/null 2>&1; then
  TCP_CONN=$(ss -ant 2>/dev/null | awk 'NR>1{c++}END{print c+0}')
  TCP_TIME_WAIT=$(ss -ant state time-wait 2>/dev/null | awk 'NR>1{c++}END{print c+0}')
elif command -v netstat >/dev/null 2>&1; then
  TCP_CONN=$(netstat -ant 2>/dev/null | awk 'NR>2{c++}END{print c+0}')
  TCP_TIME_WAIT=$(netstat -ant 2>/dev/null | awk '$NF=="TIME_WAIT"{c++}END{print c+0}')
else
  TCP_CONN=$(awk 'FNR>1{c++}END{print c+0}' /proc/net/tcp /proc/net/tcp6 2>/dev/null)
  TCP_TIME_WAIT=$(awk 'FNR>1 && $4=="06"{c++}END{print c+0}' /proc/net/tcp /proc/net/tcp6 2>/dev/null)
fi

echo "OS_VER=${OS}"
echo "CPU_CORES=${CPU}"
echo "MEM_KB=${MEM_KB}"
echo "DISK_KB=${DISK_KB}"
echo "HOSTNAME=${HN}"
echo "CPU_USAGE_PCT=${CPU_USAGE}"
echo "MEM_USAGE_PCT=${MEM_USAGE}"
echo "LOAD1=${LOAD1}"
echo "LOAD5=${LOAD5}"
echo "LOAD15=${LOAD15}"
echo "UPTIME_SECONDS=${UPTIME_SECONDS}"
echo "DISK_IO_READ_BPS=${DIO_READ_BPS}"
echo "DISK_IO_WRITE_BPS=${DIO_WRITE_BPS}"
echo "NETWORK_RX_BPS=${NET_RX_BPS}"
echo "NETWORK_TX_BPS=${NET_TX_BPS}"
echo "TCP_CONNECTIONS=${TCP_CONN}"
echo "TCP_TIME_WAIT=${TCP_TIME_WAIT}"
df -P -k -x tmpfs -x devtmpfs -x squashfs 2>/dev/null | awk 'NR>1 {
  pct=$5; gsub("%","",pct);
  printf "DISK_USAGE=%s|%s|%s|%s|%s\n", $6, $2, $3, $4, pct
}'
"""


def _is_ssh_timeout(e: Exception) -> bool:
    name = type(e).__name__
    return isinstance(e, (TimeoutError, asyncio.TimeoutError)) or name == "TimeoutError"


def _ssh_error_msg(e: Exception, host: dict | None = None) -> str:
    """把 asyncssh / asyncio 异常转为对用户友好的中文提示。"""
    import asyncssh as _assh
    name = type(e).__name__
    msg  = str(e).strip()
    if _is_ssh_timeout(e):
        target = ""
        if host:
            target = f"（{host.get('ip') or '-'}:{host.get('ssh_port') or 22}）"
        return (
            f"SSH 连接超时{target}，请检查主机 IP 是否可达、SSH 端口是否开放；"
            f"当前连接超时 {_HOST_SYNC_SSH_CONNECT_TIMEOUT:g}s，可通过 HOST_SYNC_SSH_CONNECT_TIMEOUT 调整"
        )
    if isinstance(e, _assh.PermissionDenied):
        source = "SSH 凭证" if host and host.get("credential_id") else "SSH 配置"
        return f"SSH 认证失败，请检查{source}中的用户名、密码和端口是否正确，并确认目标主机允许密码登录"
    if isinstance(e, _assh.ConnectionLost):
        return "SSH 连接中断"
    if isinstance(e, _assh.DisconnectError):
        return f"SSH 断开连接：{msg or name}"
    if isinstance(e, OSError):
        if "refused" in msg.lower() or "111" in msg:
            return "SSH 连接被拒绝，请检查目标主机 SSH 服务是否运行"
        if "network" in msg.lower() or "unreachable" in msg.lower():
            return "网络不可达，请检查网络连接"
        return f"网络错误：{msg or name}"
    if isinstance(e, ValueError):
        return msg or name
    return f"{name}：{msg}" if msg else name


def _decode_credential_password(raw_password: str) -> str | None:
    if not raw_password:
        return None
    try:
        return decrypt_password(raw_password)
    except Exception as exc:
        if raw_password.startswith("gAAAAA"):
            raise ValueError("SSH 凭证解密失败，请检查 SSH_FERNET_KEY 是否与保存凭证时一致，或重新保存该凭证") from exc
        return raw_password


def _to_int(value: str) -> int | None:
    try:
        return int(float(value.strip()))
    except Exception:
        return None


def _to_float(value: str) -> float | None:
    try:
        return round(float(value.strip()), 1)
    except Exception:
        return None


def _format_uptime(seconds: int | None) -> str:
    if seconds is None or seconds < 0:
        return ""
    days, rem = divmod(seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, _ = divmod(rem, 60)
    parts: list[str] = []
    if days:
        parts.append(f"{days}天")
    if hours or days:
        parts.append(f"{hours}小时")
    parts.append(f"{minutes}分钟")
    return "".join(parts)


async def _ssh_sync(host: dict) -> dict:
    """通过 SSH 连接主机，读取系统基本信息，返回可更新字段 dict。"""
    import asyncssh

    ip       = host.get("ip", "")
    port     = int(host.get("ssh_port") or 22)
    username = host.get("ssh_user") or "root"

    # 凭证库优先
    password = None
    cred_id = host.get("credential_id", "")
    if cred_id:
        creds = load_credentials()
        cred = next((c for c in creds if c.get("id") == cred_id), None)
        if cred:
            username = cred.get("username", username)
            port = int(cred.get("port") or port)
            raw = cred.get("password", "")
            if raw:
                password = _decode_credential_password(raw)

    # 主机自身密码
    if not password:
        enc_pw = host.get("ssh_password", "")
        if enc_pw:
            try:
                password = decrypt_password(enc_pw)
            except Exception:
                pass

    if not ip:
        raise ValueError("主机没有配置 IP")
    if not password:
        raise ValueError("主机未配置 SSH 密码，请先在编辑页面填写 SSH 密码或关联凭证")

    connect_kwargs = ssh_connect_options(
        host=ip, port=port, username=username, password=password,
        connect_timeout=_HOST_SYNC_SSH_CONNECT_TIMEOUT,
    )

    async with asyncssh.connect(**connect_kwargs) as conn:
        result = await conn.run(_SYNC_SCRIPT_SHELL, timeout=_HOST_SYNC_SSH_COMMAND_TIMEOUT)
        output = result.stdout or ""

    info: dict = {}
    disk_usage: list[dict] = []
    for line in output.splitlines():
        line = line.strip()
        if line.startswith("OS_VER="):
            v = line[7:].strip()
            if v:
                info["os_version"] = v
        elif line.startswith("CPU_CORES="):
            value = _to_int(line[10:])
            if value is not None:
                info["cpu_cores"] = value
        elif line.startswith("MEM_KB="):
            kb = _to_int(line[7:])
            if kb is not None:
                info["memory_gb"] = round(kb / 1024 / 1024, 1) if kb > 0 else None
        elif line.startswith("DISK_KB="):
            kb = _to_int(line[8:])
            if kb is not None:
                info["disk_gb"] = round(kb / 1024 / 1024, 1) if kb > 0 else None
        elif line.startswith("HOSTNAME="):
            v = line[9:].strip()
            if v:
                info["hostname"] = v
        elif line.startswith("CPU_USAGE_PCT="):
            value = _to_float(line[14:])
            if value is not None:
                info["cpu_usage_pct"] = value
        elif line.startswith("MEM_USAGE_PCT="):
            value = _to_float(line[14:])
            if value is not None:
                info["memory_usage_pct"] = value
        elif line.startswith("LOAD1="):
            value = _to_float(line[6:])
            if value is not None:
                info["load1"] = value
        elif line.startswith("LOAD5="):
            value = _to_float(line[6:])
            if value is not None:
                info["load5"] = value
        elif line.startswith("LOAD15="):
            value = _to_float(line[7:])
            if value is not None:
                info["load15"] = value
        elif line.startswith("UPTIME_SECONDS="):
            value = _to_int(line[15:])
            if value is not None:
                info["uptime_seconds"] = value
                info["uptime_text"] = _format_uptime(value)
        elif line.startswith("DISK_IO_READ_BPS="):
            value = _to_int(line[17:])
            if value is not None:
                info["disk_io_read_bps"] = value
        elif line.startswith("DISK_IO_WRITE_BPS="):
            value = _to_int(line[18:])
            if value is not None:
                info["disk_io_write_bps"] = value
        elif line.startswith("NETWORK_RX_BPS="):
            value = _to_int(line[15:])
            if value is not None:
                info["network_rx_bps"] = value
        elif line.startswith("NETWORK_TX_BPS="):
            value = _to_int(line[15:])
            if value is not None:
                info["network_tx_bps"] = value
        elif line.startswith("TCP_CONNECTIONS="):
            value = _to_int(line[16:])
            if value is not None:
                info["tcp_connections"] = value
        elif line.startswith("TCP_TIME_WAIT="):
            value = _to_int(line[14:])
            if value is not None:
                info["tcp_time_wait"] = value
        elif line.startswith("DISK_USAGE="):
            parts = line[11:].split("|")
            if len(parts) == 5:
                mount, size_kb_raw, used_kb_raw, avail_kb_raw, pct_raw = parts
                size_kb = _to_int(size_kb_raw)
                used_kb = _to_int(used_kb_raw)
                avail_kb = _to_int(avail_kb_raw)
                pct = _to_float(pct_raw)
                if mount and size_kb is not None and used_kb is not None:
                    disk_usage.append({
                        "mount": mount,
                        "size_gb": round(size_kb / 1024 / 1024, 1),
                        "used_gb": round(used_kb / 1024 / 1024, 1),
                        "avail_gb": round((avail_kb or 0) / 1024 / 1024, 1),
                        "used_pct": pct,
                    })

    if disk_usage:
        info["disk_usage"] = disk_usage
    if any(k in info for k in (
        "cpu_usage_pct", "memory_usage_pct", "load5", "disk_usage",
        "disk_io_read_bps", "network_rx_bps", "tcp_connections", "uptime_seconds",
    )):
        info["metrics_updated_at"] = _NOW()

    return info


@router.post("/api/hosts/{host_id}/sync")
async def sync_host_info(host_id: str):
    """通过 SSH 同步主机系统信息（OS版本、CPU、内存、磁盘、主机名）。"""
    hosts = load_hosts_list()
    idx = next((i for i, h in enumerate(hosts) if h.get("id") == host_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="主机不存在")
    host = hosts[idx]
    try:
        info = await _ssh_sync(host)
    except Exception as e:
        detail = _ssh_error_msg(e, host)
        logger.warning(
            "[sync] 主机 %s (%s:%s) 同步失败: %s; raw=%r",
            host.get("hostname"),
            host.get("ip"),
            host.get("ssh_port") or 22,
            detail,
            e,
        )
        raise HTTPException(status_code=408 if _is_ssh_timeout(e) else 400, detail=detail)

    if not info:
        raise HTTPException(status_code=400, detail="SSH 连接成功但未获取到系统信息，请确认 shell 可正常执行")

    # 更新字段
    now = _NOW()
    for k, v in info.items():
        if v is not None:
            host[k] = v
    host["last_sync_at"] = now
    host["updated_at"] = now
    hosts[idx] = host
    save_hosts_list(hosts)

    result = dict(host)
    result["ssh_saved"] = bool(result.pop("ssh_password", "")) or bool(result.get("credential_id"))
    return {"ok": True, "updated": info, "host": result}


@router.post("/api/hosts/sync-all")
async def sync_all_hosts(user: User = Depends(current_user)):
    """一键同步（非管理员只同步自己分组的主机）。"""
    async def generate():
        all_hosts = load_hosts_list()
        visible   = _filter_hosts_by_user(all_hosts, user)
        hosts     = all_hosts  # 写回时用全量
        targets   = [h for h in visible if h.get("ssh_password") or h.get("credential_id")]
        total = len(targets)

        if total == 0:
            yield f"data: {json.dumps({'type':'done','success':0,'fail':0,'total':0,'message':'没有配置 SSH 凭证的主机'}, ensure_ascii=False)}\n\n"
            return

        yield f"data: {json.dumps({'type':'start','total':total}, ensure_ascii=False)}\n\n"

        sem = asyncio.Semaphore(5)

        async def _do_sync(host: dict) -> dict:
            async with sem:
                try:
                    info = await _ssh_sync(host)
                    return {
                        "ok": True,
                        "id": host["id"],
                        "hostname": host.get("hostname", ""),
                        "ip": host.get("ip", ""),
                        "info": info,
                    }
                except Exception as e:
                    return {
                        "ok": False,
                        "id": host["id"],
                        "hostname": host.get("hostname", ""),
                        "ip": host.get("ip", ""),
                        "error": _ssh_error_msg(e, host),
                    }

        # 全部并发执行，等待所有结果
        results = await asyncio.gather(*[_do_sync(h) for h in targets], return_exceptions=True)

        id_to_host = {h["id"]: h for h in hosts}
        success = fail = 0

        for res in results:
            if isinstance(res, Exception):
                fail += 1
                yield f"data: {json.dumps({'type':'progress','ok':False,'error':str(res),'success':success,'fail':fail,'total':total}, ensure_ascii=False)}\n\n"
                continue

            host_id = res["id"]
            h = id_to_host.get(host_id)

            if res["ok"] and h:
                now = _NOW()
                for k, v in res["info"].items():
                    if v is not None:
                        h[k] = v
                h["last_sync_at"] = now
                h["updated_at"] = now
                success += 1
                yield f"data: {json.dumps({'type':'progress','id':host_id,'hostname':res['hostname'],'ip':res['ip'],'ok':True,'updated':res['info'],'success':success,'fail':fail,'total':total}, ensure_ascii=False)}\n\n"
            else:
                fail += 1
                yield f"data: {json.dumps({'type':'progress','id':host_id,'hostname':res['hostname'],'ip':res['ip'],'ok':False,'error':res.get('error',''),'success':success,'fail':fail,'total':total}, ensure_ascii=False)}\n\n"

        # 所有主机处理完后统一保存
        save_hosts_list(hosts)

        msg = f"同步完成：成功 {success} 台，失败 {fail} 台"
        yield f"data: {json.dumps({'type':'done','success':success,'fail':fail,'total':total,'message':msg}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── 进程列表 ─────────────────────────────────────────────────────────────────

# 服务特征规则：(匹配关键字列表, 显示名, 颜色代号)
_SERVICE_RULES: list[tuple[list[str], str, str]] = [
    (["mysqld", "mysql"],              "MySQL",         "mysql"),
    (["java"],                         "Java",          "java"),
    (["python", "python3", "python2"], "Python",        "python"),
    (["node", "nodejs"],               "Node.js",       "node"),
    (["nginx"],                        "Nginx",         "nginx"),
    (["redis-server", "redis"],        "Redis",         "redis"),
    (["postgres", "postgresql"],       "PostgreSQL",    "postgres"),
    (["elasticsearch", "kibana"],      "Elastic",       "elastic"),
    (["docker", "containerd"],         "Docker",        "docker"),
    (["kubelet", "kube-proxy", "etcd"],"Kubernetes",    "k8s"),
    (["mongod", "mongo"],              "MongoDB",       "mongo"),
    (["rabbitmq"],                     "RabbitMQ",      "rabbit"),
    (["kafka", "zookeeper"],           "Kafka/ZK",      "kafka"),
    (["sshd"],                         "SSH",           "ssh"),
    (["php", "php-fpm"],               "PHP",           "php"),
    (["go"],                           "Go",            "go"),
]

_PROC_CMD = r"""
ps -eo pid,ppid,user,%cpu,%mem,rss,stat,comm,args --sort=-%cpu --no-headers 2>/dev/null | head -30 | awk '{
  pid=$1; ppid=$2; user=$3; cpu=$4; mem=$5; rss=$6; stat=$7; comm=$8;
  args="";
  for(i=9;i<=NF;i++) args=args" "$i;
  sub(/^ /,"",args);
  printf "%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\n", pid,ppid,user,cpu,mem,rss,stat,comm,args
}'
"""


def _classify_process(comm: str, args: str) -> tuple[str, str] | None:
    """返回 (service_name, color_key) 或 None。"""
    target = (comm + " " + args).lower()
    for keywords, name, color in _SERVICE_RULES:
        if any(kw in target for kw in keywords):
            return name, color
    return None


def _parse_processes(raw: str) -> list[dict]:
    """解析 ps 输出为进程列表，附加服务分类信息。"""
    procs: list[dict] = []
    for line in raw.strip().splitlines():
        parts = line.split("\t", 8)
        if len(parts) < 8:
            continue
        pid, ppid, user, cpu, mem, rss, stat, comm = parts[:8]
        args = parts[8] if len(parts) > 8 else comm
        try:
            cpu_f = float(cpu)
            mem_f = float(mem)
            rss_mb = round(int(rss) / 1024, 1)
        except (ValueError, TypeError):
            continue

        svc = _classify_process(comm, args)
        # 截断过长的命令行
        display_args = args[:120] + "…" if len(args) > 120 else args

        procs.append({
            "pid":      int(pid) if pid.isdigit() else 0,
            "user":     user,
            "cpu":      cpu_f,
            "mem":      mem_f,
            "rss_mb":   rss_mb,
            "stat":     stat,
            "comm":     comm,
            "args":     display_args,
            "service":  svc[0] if svc else None,
            "color":    svc[1] if svc else None,
        })
    return procs


def _prioritize(procs: list[dict]) -> list[dict]:
    """服务进程置顶，同层内按 CPU 排序，最多返回 15 条。"""
    service = [p for p in procs if p["service"]]
    others  = [p for p in procs if not p["service"]]
    # 服务进程按 CPU 降序，但去掉空闲的 kthreadd/idle 等
    service.sort(key=lambda p: -p["cpu"])
    others.sort(key=lambda p: -p["cpu"])
    combined = service + others
    return combined[:15]


_JAVA_DIAG_DIR = REPORTS_DIR / "java_diagnostics"
_JAVA_DIAG_DIR.mkdir(exist_ok=True)

_ARTHAS_PRESETS = {
    "dashboard": "dashboard -n 1",
    "jvm": "jvm",
    "thread_top": "thread -n 5",
    "thread_blocked": "thread -b",
    "memory": "memory",
}

_JAVA_FLAME_EVENTS = {"cpu", "alloc", "lock", "wall"}


class JavaArthasRequest(BaseModel):
    pid: int
    preset: str = "dashboard"
    command: str = ""


class JavaFlamegraphRequest(BaseModel):
    pid: int
    seconds: int = 30
    event: str = "cpu"


def _get_host_or_404(host_id: str) -> dict:
    host = next((h for h in load_hosts_list() if h.get("id") == host_id), None)
    if not host:
        raise HTTPException(status_code=404, detail="主机不存在")
    return host


def _resolve_host_ssh_auth(host: dict) -> tuple[str, str, int, str]:
    ip = str(host.get("ip") or "").strip()
    port = int(host.get("ssh_port") or 22)
    username = str(host.get("ssh_user") or "root").strip() or "root"
    password = None

    cred_id = host.get("credential_id", "")
    if cred_id:
        creds = load_credentials()
        cred = next((c for c in creds if c.get("id") == cred_id), None)
        if cred:
            username = str(cred.get("username") or username).strip() or username
            port = int(cred.get("port") or port)
            raw_pw = cred.get("password", "")
            if raw_pw:
                password = _decode_credential_password(raw_pw)

    if not password:
        enc_pw = host.get("ssh_password", "")
        if enc_pw:
            try:
                password = decrypt_password(enc_pw)
            except Exception:
                password = None

    if not ip:
        raise HTTPException(status_code=400, detail="主机没有配置 IP")
    if not password:
        raise HTTPException(status_code=400, detail="主机未配置 SSH 密码，无法执行诊断")
    return ip, username, port, password


def _java_artifact_url(filename: str) -> str:
    return f"/api/hosts/java-diagnostics/artifacts/{quote(filename)}"


def _save_java_artifact(filename: str, content: str, encoding: str = "utf-8") -> Path:
    path = _JAVA_DIAG_DIR / filename
    path.write_text(content, encoding=encoding)
    return path


async def _ssh_run(host: dict, command: str, timeout: float):
    import asyncssh

    ip, username, port, password = _resolve_host_ssh_auth(host)
    return await asyncssh.connect(
        **ssh_connect_options(
            host=ip,
            port=port,
            username=username,
            password=password,
            connect_timeout=_HOST_SYNC_SSH_CONNECT_TIMEOUT,
        )
    )


def _build_arthas_command(pid: int, arthas_command: str) -> str:
    return f"""set -e
# 清除代理，避免 curl/wget 走本地代理隧道失败
unset http_proxy https_proxy HTTP_PROXY HTTPS_PROXY ALL_PROXY all_proxy no_proxy NO_PROXY 2>/dev/null || true

WORKDIR=/tmp/aiops-java-tools
BOOT="$WORKDIR/arthas-boot.jar"
mkdir -p "$WORKDIR"

# 自动探测 java 路径，并为 Arthas 归一化 JAVA_HOME（JDK8 下不能落到 .../jre）
_find_java() {{
  # 1. PATH 里直接找，并解析为真实路径
  if command -v java >/dev/null 2>&1; then
    readlink -f "$(command -v java)" 2>/dev/null || command -v java
    return
  fi
  # 2. 从目标 JVM 进程的符号链接找（需要同用户或 root 才有权限）
  local exe
  exe=$(readlink -f /proc/{pid}/exe 2>/dev/null)
  if [ -n "$exe" ] && [ -x "$exe" ]; then echo "$exe"; return; fi
  # 3. /proc/{pid}/cmdline 首参数：很多启动脚本写绝对路径 (/data/jdk/bin/java -jar app.jar)
  local cmd0
  cmd0=$(tr '\\0' '\\n' < /proc/{pid}/cmdline 2>/dev/null | head -1)
  if [ -n "$cmd0" ] && [ -x "$cmd0" ] && [ "${{cmd0##*/}}" = "java" ]; then
    readlink -f "$cmd0" 2>/dev/null || echo "$cmd0"; return
  fi
  # 4. ps 命令行兜底（pid 已死也能从父进程拿到）
  local ps_cmd
  ps_cmd=$(ps -p {pid} -o args= 2>/dev/null | awk '{{print $1}}')
  if [ -n "$ps_cmd" ] && [ -x "$ps_cmd" ] && [ "${{ps_cmd##*/}}" = "java" ]; then
    readlink -f "$ps_cmd" 2>/dev/null || echo "$ps_cmd"; return
  fi
  # 5. 从进程 JAVA_HOME 环境变量找
  local jhome
  jhome=$(cat /proc/{pid}/environ 2>/dev/null | tr '\\0' '\\n' | grep '^JAVA_HOME=' | cut -d= -f2)
  if [ -n "$jhome" ] && [ "${{jhome%/jre}}" != "$jhome" ]; then
    jhome="${{jhome%/jre}}"
  fi
  if [ -n "$jhome" ] && [ -x "$jhome/bin/java" ]; then echo "$jhome/bin/java"; return; fi
  if [ -n "$jhome" ] && [ -x "$jhome/jre/bin/java" ]; then echo "$jhome/jre/bin/java"; return; fi
  if [ -n "$jhome" ] && [ "${{jhome#/}}" != "$jhome" ]; then
    if [ -x "/proc/{pid}/root$jhome/bin/java" ]; then echo "/proc/{pid}/root$jhome/bin/java"; return; fi
    if [ -x "/proc/{pid}/root$jhome/jre/bin/java" ]; then echo "/proc/{pid}/root$jhome/jre/bin/java"; return; fi
  fi
  local proc_path old_ifs dir candidate
  proc_path=$(cat /proc/{pid}/environ 2>/dev/null | tr '\\0' '\\n' | grep '^PATH=' | cut -d= -f2-)
  if [ -n "$proc_path" ]; then
    old_ifs="$IFS"; IFS=:
    for dir in $proc_path; do
      [ -z "$dir" ] && continue
      candidate="$dir/java"
      if [ -x "$candidate" ]; then readlink -f "$candidate" 2>/dev/null || echo "$candidate"; IFS="$old_ifs"; return; fi
      if [ "${{dir#/}}" != "$dir" ] && [ -x "/proc/{pid}/root$dir/java" ]; then echo "/proc/{pid}/root$dir/java"; IFS="$old_ifs"; return; fi
    done
    IFS="$old_ifs"
  fi
  # 6. 常见安装路径（覆盖容器/物理机/云主机常见目录）
  for p in /usr/local/java/bin/java /usr/lib/jvm/*/bin/java \\
           /opt/jdk*/bin/java /opt/java/*/bin/java /opt/jdk/*/bin/java \\
           /data/jdk*/bin/java /data/java*/bin/java /data/jdk/*/bin/java \\
           /app/jdk*/bin/java /app/java*/bin/java \\
           /srv/jdk*/bin/java /srv/java*/bin/java \\
           /home/*/jdk*/bin/java /home/*/java*/bin/java \\
           /root/jdk*/bin/java /root/java*/bin/java \\
           /usr/java/*/bin/java; do
    [ -x "$p" ] && {{ readlink -f "$p" 2>/dev/null || echo "$p"; return; }}
  done
  # 7. 最后兜底：限深 find，避免全盘扫描
  local found
  found=$(find /opt /data /app /srv /usr/local /usr/lib/jvm /root /home -maxdepth 5 \\
          -type f -executable -name java 2>/dev/null | head -1)
  if [ -n "$found" ]; then echo "$found"; return; fi
  # 失败前打印诊断信息到 stderr 帮助排查
  echo "[diag] PATH=$PATH" >&2
  echo "[diag] /proc/{pid}/exe -> $(readlink -f /proc/{pid}/exe 2>&1 || echo '权限不足')" >&2
  echo "[diag] /proc/{pid}/cmdline first arg -> ${{cmd0:-空}}" >&2
  echo "[diag] ps args first token -> ${{ps_cmd:-空}}" >&2
  echo "[diag] JAVA_HOME from env -> ${{jhome:-空}}" >&2
  echo "[diag] PATH from target env -> ${{proc_path:-空}}" >&2
  echo ""
}}

_find_jdk_home() {{
  local java_path="$1"
  local jhome=""
  if [ -n "$java_path" ]; then
    java_path=$(readlink -f "$java_path" 2>/dev/null || echo "$java_path")
    case "$java_path" in
      */jre/bin/java) jhome=$(dirname "$(dirname "$(dirname "$java_path")")") ;;
      */bin/java) jhome=$(dirname "$(dirname "$java_path")") ;;
    esac
  fi
  if [ -z "$jhome" ]; then
    jhome=$(cat /proc/{pid}/environ 2>/dev/null | tr '\\0' '\\n' | grep '^JAVA_HOME=' | cut -d= -f2)
  fi
  if [ -n "$jhome" ] && [ "${{jhome%/jre}}" != "$jhome" ]; then
    jhome="${{jhome%/jre}}"
  fi
  if [ -n "$jhome" ] && [ -f "$jhome/lib/tools.jar" ]; then
    echo "$jhome"
    return
  fi
  if [ -n "$jhome" ] && [ -x "$jhome/bin/javac" ]; then
    echo "$jhome"
    return
  fi
  echo "$jhome"
}}

JAVA=$(_find_java)
if [ -z "$JAVA" ]; then
  echo "" >&2
  echo "未找到 java 命令。请确认：" >&2
  echo "  1) 目标主机已安装可执行 java；JDK8 建议完整 JDK（tools.jar），Java 9+ 不需要 tools.jar" >&2
  echo "  2) PID {pid} 仍存活" >&2
  echo "  3) SSH 用户对 /proc/{pid}/exe 有读权限（同用户或 root）" >&2
  echo "  4) 或在系统 PATH、目标进程 PATH、JAVA_HOME 中配置 java" >&2
  echo "  上方 [diag] 输出包含诊断信息" >&2
  exit 2
fi
JAVA_HOME=$(_find_jdk_home "$JAVA")
if [ -n "$JAVA_HOME" ]; then
  export JAVA_HOME
fi
echo "使用 Java: $JAVA"
echo "使用 JAVA_HOME: $JAVA_HOME"

# 容器内 java 路径形如 /proc/<pid>/root/...，宿主机动态链接器找不到 libjli.so，
# 需要把容器内 JDK 的 lib 目录加入 LD_LIBRARY_PATH（兼容 JDK8/JDK11+ 多种布局）。
case "$JAVA" in
  /proc/*/root/*)
    java_proc_prefix=$(echo "$JAVA" | sed -E 's|(^/proc/[0-9]+/root)/.*|\\1|')
    if [ -n "$JAVA_HOME" ]; then
      jh_in_container="$JAVA_HOME"
    else
      jh_in_container=$(dirname "$(dirname "${{JAVA#$java_proc_prefix}}")")
    fi
    ld_extra=""
    for sub in lib lib/jli lib/server lib/amd64 lib/amd64/jli lib/amd64/server \
               jre/lib jre/lib/amd64 jre/lib/amd64/jli jre/lib/amd64/server; do
      candidate="$java_proc_prefix$jh_in_container/$sub"
      if [ -d "$candidate" ]; then
        ld_extra="$ld_extra:$candidate"
      fi
    done
    if [ -n "$ld_extra" ]; then
      export LD_LIBRARY_PATH="${{ld_extra#:}}${{LD_LIBRARY_PATH:+:$LD_LIBRARY_PATH}}"
      echo "使用 LD_LIBRARY_PATH: $LD_LIBRARY_PATH"
    fi
    ;;
esac

_download_file() {{
  local url="$1"
  local output="$2"

  if command -v curl >/dev/null 2>&1; then
    curl --noproxy '*' -fsSL -o "$output" "$url" && return 0
    echo "curl 下载失败: $url" >&2
    return 1
  fi

  if command -v wget >/dev/null 2>&1; then
    wget --no-proxy -qO "$output" "$url" && return 0
    echo "wget 下载失败: $url" >&2
    return 1
  fi

  echo "缺少 curl/wget，无法下载文件: $url" >&2
  return 1
}}

if [ ! -s "$BOOT" ]; then
  if ! _download_file https://arthas.aliyun.com/arthas-boot.jar "$BOOT"; then
    echo "下载 Arthas 失败，请检查目标机外网访问或代理配置" >&2
    exit 2
  fi
fi
"$JAVA" -jar "$BOOT" {pid} --action exec --command {shlex.quote(arthas_command)} 2>&1
"""


def _build_async_profiler_command(pid: int, seconds: int, event: str, remote_output: str) -> str:
    return f"""set -e
# 清除代理，避免 curl/wget 走本地代理隧道失败
unset http_proxy https_proxy HTTP_PROXY HTTPS_PROXY ALL_PROXY all_proxy no_proxy NO_PROXY 2>/dev/null || true

WORKDIR=/tmp/aiops-java-tools
PROFILE_HOME="$WORKDIR/async-profiler"
ARCH="$(uname -m)"
mkdir -p "$WORKDIR"

_find_java() {{
  if command -v java >/dev/null 2>&1; then
    readlink -f "$(command -v java)" 2>/dev/null || command -v java
    return
  fi
  local exe; exe=$(readlink -f /proc/{pid}/exe 2>/dev/null)
  if [ -n "$exe" ] && [ -x "$exe" ]; then echo "$exe"; return; fi
  local jhome; jhome=$(cat /proc/{pid}/environ 2>/dev/null | tr '\\0' '\\n' | grep '^JAVA_HOME=' | cut -d= -f2)
  if [ -n "$jhome" ] && [ "${{jhome%/jre}}" != "$jhome" ]; then
    jhome="${{jhome%/jre}}"
  fi
  if [ -n "$jhome" ] && [ -x "$jhome/bin/java" ]; then echo "$jhome/bin/java"; return; fi
  if [ -n "$jhome" ] && [ -x "$jhome/jre/bin/java" ]; then echo "$jhome/jre/bin/java"; return; fi
  if [ -n "$jhome" ] && [ "${{jhome#/}}" != "$jhome" ]; then
    if [ -x "/proc/{pid}/root$jhome/bin/java" ]; then echo "/proc/{pid}/root$jhome/bin/java"; return; fi
    if [ -x "/proc/{pid}/root$jhome/jre/bin/java" ]; then echo "/proc/{pid}/root$jhome/jre/bin/java"; return; fi
  fi
  local proc_path old_ifs dir candidate
  proc_path=$(cat /proc/{pid}/environ 2>/dev/null | tr '\\0' '\\n' | grep '^PATH=' | cut -d= -f2-)
  if [ -n "$proc_path" ]; then
    old_ifs="$IFS"; IFS=:
    for dir in $proc_path; do
      [ -z "$dir" ] && continue
      candidate="$dir/java"
      if [ -x "$candidate" ]; then readlink -f "$candidate" 2>/dev/null || echo "$candidate"; IFS="$old_ifs"; return; fi
      if [ "${{dir#/}}" != "$dir" ] && [ -x "/proc/{pid}/root$dir/java" ]; then echo "/proc/{pid}/root$dir/java"; IFS="$old_ifs"; return; fi
    done
    IFS="$old_ifs"
  fi
  for p in /usr/local/java/bin/java /usr/lib/jvm/*/bin/java /opt/jdk*/bin/java /opt/java/*/bin/java; do
    [ -x "$p" ] && {{ readlink -f "$p" 2>/dev/null || echo "$p"; return; }}
  done
  echo ""
}}
JAVA=$(_find_java)
if [ -z "$JAVA" ]; then
  echo "" >&2
  echo "未找到 java 命令。请确认目标主机已安装可执行 java 且 PID {pid} 存活。" >&2
  echo "上方 [diag] 输出包含 PATH / /proc/{pid}/exe / cmdline 等诊断信息" >&2
  exit 2
fi
_download_file() {{
  local url="$1"
  local output="$2"

  if command -v curl >/dev/null 2>&1; then
    curl --noproxy '*' -fsSL -o "$output" "$url" && return 0
    echo "curl 下载失败: $url" >&2
    return 1
  fi

  if command -v wget >/dev/null 2>&1; then
    wget --no-proxy -qO "$output" "$url" && return 0
    echo "wget 下载失败: $url" >&2
    return 1
  fi

  echo "缺少 curl/wget，无法下载文件: $url" >&2
  return 1
}}

case "$ARCH" in
  x86_64|amd64) PKG="linux-x64" ;;
  aarch64|arm64) PKG="linux-arm64" ;;
  *)
    echo "不支持的架构: $ARCH" >&2
    exit 2
    ;;
esac
if [ ! -x "$PROFILE_HOME/profiler.sh" ]; then
  cd "$WORKDIR"
  rm -rf "$PROFILE_HOME" async-profiler async-profiler-*.tar.gz async-profiler-*
  URL="https://github.com/async-profiler/async-profiler/releases/download/v3.0/async-profiler-3.0-$PKG.tar.gz"
  if ! _download_file "$URL" async-profiler.tar.gz; then
    echo "下载 async-profiler 失败，请检查目标机外网访问或代理配置" >&2
    exit 2
  fi
  tar -xzf async-profiler.tar.gz
  SRC_DIR="$(find "$WORKDIR" -maxdepth 1 -type d -name 'async-profiler-*' | head -1)"
  if [ -z "$SRC_DIR" ]; then
    echo "async-profiler 解压失败" >&2
    exit 2
  fi
  mv "$SRC_DIR" "$PROFILE_HOME"
fi
"$PROFILE_HOME/profiler.sh" -d {seconds} -e {event} -f {shlex.quote(remote_output)} {pid}
"""


@router.get("/api/hosts/{host_id}/processes")
async def get_host_processes(host_id: str):
    """通过 SSH 获取主机 Top 进程列表（服务进程置顶）。"""
    host = _get_host_or_404(host_id)

    try:
        async with await _ssh_run(host, _PROC_CMD, timeout=10) as conn:
            result = await conn.run(_PROC_CMD, timeout=10)
            raw = result.stdout or ""
    except Exception as e:
        raise HTTPException(status_code=400, detail=_ssh_error_msg(e, host))

    procs = _parse_processes(raw)
    procs = _prioritize(procs)

    # 统计检测到的服务
    detected = sorted({p["service"] for p in procs if p["service"]})
    return {"data": procs, "detected_services": detected, "total": len(procs)}


@router.post("/api/hosts/{host_id}/java-diagnostics/arthas")
async def run_java_arthas(host_id: str, body: JavaArthasRequest):
    host = _get_host_or_404(host_id)
    pid = int(body.pid or 0)
    if pid <= 0:
        raise HTTPException(status_code=400, detail="PID 必须大于 0")

    preset = (body.preset or "dashboard").strip()
    if preset == "custom":
        arthas_command = (body.command or "").strip()
        if not arthas_command:
            raise HTTPException(status_code=400, detail="自定义 Arthas 命令不能为空")
    else:
        arthas_command = _ARTHAS_PRESETS.get(preset)
        if not arthas_command:
            raise HTTPException(status_code=400, detail="不支持的 Arthas 预设")

    command = _build_arthas_command(pid, arthas_command)
    try:
        async with await _ssh_run(host, command, timeout=90) as conn:
            result = await conn.run(command, timeout=90, check=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=_ssh_error_msg(e, host))

    stdout = (result.stdout or "").strip()
    stderr = (result.stderr or "").strip()
    if result.exit_status != 0:
        detail = stderr or stdout or "Arthas 执行失败"
        raise HTTPException(status_code=400, detail=detail)

    artifact_name = f"arthas_{host_id}_{pid}_{uuid.uuid4().hex[:8]}.txt"
    preview = stdout or stderr or "(无输出)"
    _save_java_artifact(artifact_name, preview)
    return {
        "ok": True,
        "tool": "arthas",
        "pid": pid,
        "preset": preset,
        "command": arthas_command,
        "stdout": stdout,
        "stderr": stderr,
        "download_url": _java_artifact_url(artifact_name),
    }


@router.post("/api/hosts/{host_id}/java-diagnostics/flamegraph")
async def run_java_flamegraph(host_id: str, body: JavaFlamegraphRequest):
    host = _get_host_or_404(host_id)
    pid = int(body.pid or 0)
    seconds = int(body.seconds or 0)
    event = (body.event or "cpu").strip().lower()

    if pid <= 0:
        raise HTTPException(status_code=400, detail="PID 必须大于 0")
    if seconds < 10 or seconds > 300:
        raise HTTPException(status_code=400, detail="采样时长必须在 10 到 300 秒之间")
    if event not in _JAVA_FLAME_EVENTS:
        raise HTTPException(status_code=400, detail="不支持的火焰图事件类型")

    artifact_name = f"flamegraph_{host_id}_{pid}_{event}_{uuid.uuid4().hex[:8]}.html"
    remote_output = f"/tmp/aiops-java-tools/{artifact_name}"
    command = _build_async_profiler_command(pid, seconds, event, remote_output)

    try:
        async with await _ssh_run(host, command, timeout=seconds + 120) as conn:
            result = await conn.run(command, timeout=seconds + 120, check=False)
            stdout = (result.stdout or "").strip()
            stderr = (result.stderr or "").strip()
            if result.exit_status != 0:
                detail = stderr or stdout or "火焰图生成失败"
                raise HTTPException(status_code=400, detail=detail)

            async with conn.start_sftp_client() as sftp:
                local_path = _JAVA_DIAG_DIR / artifact_name
                await sftp.get(remote_output, str(local_path))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=_ssh_error_msg(e, host))

    size_bytes = (_JAVA_DIAG_DIR / artifact_name).stat().st_size
    return {
        "ok": True,
        "tool": "flamegraph",
        "pid": pid,
        "seconds": seconds,
        "event": event,
        "size_bytes": size_bytes,
        "download_url": _java_artifact_url(artifact_name),
        "stdout": stdout,
    }


@router.get("/api/hosts/java-diagnostics/artifacts/{filename}")
async def download_java_diag_artifact(filename: str):
    safe_name = Path(filename).name
    if safe_name != filename:
        raise HTTPException(status_code=400, detail="非法文件名")

    path = _JAVA_DIAG_DIR / safe_name
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="诊断结果不存在")

    suffix = path.suffix.lower()
    media_type = {
        ".html": "text/html; charset=utf-8",
        ".txt": "text/plain; charset=utf-8",
        ".svg": "image/svg+xml",
    }.get(suffix, "application/octet-stream")
    encoded = quote(path.name)
    disposition = "inline" if suffix in {".html", ".svg"} else "attachment"
    return Response(
        content=path.read_bytes(),
        media_type=media_type,
        headers={"Content-Disposition": f"{disposition}; filename*=UTF-8''{encoded}"},
    )


# ── 巡检（基于 Prometheus，使用 CMDB 中的 IP 匹配实例）──────────────────────

def _inspection_issue_counts(results: list[dict]) -> list[tuple[str, int]]:
    counter: dict[str, int] = {}
    for result in results:
        for check in result.get("checks", []):
            if check.get("status") == "normal":
                continue
            item = check.get("item", "未知项")
            counter[item] = counter.get(item, 0) + 1
    return sorted(counter.items(), key=lambda item: item[1], reverse=True)


def _build_inspection_fallback_summary(
    results: list[dict], summary: dict, error: Exception | None = None
) -> str:
    total    = summary.get("total", len(results))
    normal   = summary.get("normal", 0)
    warning  = summary.get("warning", 0)
    critical = summary.get("critical", 0)
    issue_counts   = _inspection_issue_counts(results)
    top_issue_text = "、".join(
        f"{item}（{count}台）" for item, count in issue_counts[:3]
    ) or "暂未发现重复性异常项"

    abnormal_hosts = [r for r in results if r.get("overall") != "normal"]
    abnormal_hosts.sort(
        key=lambda item: (
            {"critical": 2, "warning": 1, "normal": 0}.get(item.get("overall", "normal"), 0),
            sum(1 for c in item.get("checks", []) if c.get("status") != "normal"),
        ),
        reverse=True,
    )
    top_host_text = "；".join(
        f"{host.get('hostname') or host.get('instance')}({host.get('ip', '-')})"
        for host in abnormal_hosts[:3]
    ) or "暂无异常主机"

    parts = [
        f"本次巡检共覆盖 {total} 台主机，其中正常 {normal} 台、警告 {warning} 台、严重 {critical} 台。",
        f"当前最集中的异常项为：{top_issue_text}。",
        f"需优先关注的主机：{top_host_text}。",
        "建议先处理严重主机，再按高频异常项排查资源瓶颈、容量风险和连接状态波动。",
        "未来24小时若不处理，异常项可能继续扩散到更多主机，并导致性能抖动、告警增加或服务稳定性下降。",
    ]
    if error:
        parts.append(f"AI 服务暂不可用，当前显示的是规则摘要（{error}）。")
    return "\n\n".join(parts)


def _cmdb_instances_for_group(group_id: str) -> tuple[list[str], str]:
    """根据分组 ID 从 CMDB 获取 Prometheus 实例列表（ip:9100），返回 (instances, group_name)。"""
    hosts = load_hosts_list()
    groups = load_groups()
    g = next((g for g in groups if g["id"] == group_id), None)
    group_name = g["name"] if g else group_id
    instances = [f"{h['ip']}:9100" for h in hosts if h.get("group") == group_id and h.get("ip")]
    return instances, group_name


def _metric_float(metrics: dict, key: str) -> float | None:
    raw = metrics.get(key)
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _metric_int(metrics: dict, key: str) -> int | None:
    value = _metric_float(metrics, key)
    return int(value) if value is not None else None


def _inspection_severity(status: str) -> int:
    return {"critical": 2, "warning": 1, "normal": 0}.get(status, 0)


def _inspection_overall(checks: list[dict]) -> str:
    worst = max((_inspection_severity(c.get("status", "")) for c in checks), default=0)
    return "critical" if worst >= 2 else "warning" if worst == 1 else "normal"


def _bps_to_mib_per_sec(value) -> float | None:
    try:
        return round(float(value) / (1024 * 1024), 2)
    except (TypeError, ValueError):
        return None


def _sync_info_to_inspection_result(host: dict, info: dict) -> dict | None:
    """把 SSH/Python 兜底采集结果转成巡检报告结构。"""
    metrics: dict = {}

    mapping = {
        "cpu_usage_pct": "cpu_usage",
        "memory_usage_pct": "mem_usage",
        "memory_gb": "mem_total_gb",
        "load1": "load1",
        "load5": "load5",
        "load15": "load15",
        "uptime_seconds": "uptime_seconds",
        "tcp_connections": "tcp_estab",
        "tcp_time_wait": "tcp_tw",
    }
    for src, dst in mapping.items():
        value = info.get(src)
        if value is not None:
            metrics[dst] = value

    byte_rate_mapping = {
        "network_rx_bps": "net_recv_mbps",
        "network_tx_bps": "net_send_mbps",
        "disk_io_read_bps": "disk_read_mbps",
        "disk_io_write_bps": "disk_write_mbps",
    }
    for src, dst in byte_rate_mapping.items():
        value = _bps_to_mib_per_sec(info.get(src))
        if value is not None:
            metrics[dst] = value

    partitions: list[dict] = []
    for p in info.get("disk_usage") or []:
        mount = p.get("mount") or p.get("mountpoint") or ""
        if not mount:
            continue
        used_pct = p.get("used_pct")
        if used_pct is None:
            continue
        partitions.append({
            "mountpoint": mount,
            "mount": mount,
            "fstype": p.get("fstype", ""),
            "total_gb": p.get("size_gb") or p.get("total_gb") or 0,
            "used_gb": p.get("used_gb") or 0,
            "avail_gb": p.get("avail_gb") or 0,
            "usage_pct": used_pct,
            "used_pct": used_pct,
        })

    if not metrics and not partitions:
        return None

    cpu_cores = info.get("cpu_cores") or host.get("cpu_cores") or 4
    try:
        checks = prom._build_checks(metrics, int(cpu_cores or 4), partitions)
    except Exception:
        checks = []
    checks.append({
        "item": "指标来源",
        "value": "SSH/Python 兜底采集",
        "status": "normal",
        "threshold": "Prometheus 无数据时自动兜底",
    })

    ip = host.get("ip", "")
    return {
        "instance": f"{ip}:ssh",
        "ip": ip,
        "hostname": info.get("hostname") or host.get("hostname") or ip,
        "os": info.get("os_version") or host.get("os_version", ""),
        "os_version": info.get("os_version") or host.get("os_version", ""),
        "cpu_cores": info.get("cpu_cores") or host.get("cpu_cores"),
        "memory_gb": info.get("memory_gb") or host.get("memory_gb"),
        "disk_gb": info.get("disk_gb") or host.get("disk_gb"),
        "job": "ssh-python-fallback",
        "state": "fallback",
        "group": host.get("group", ""),
        "env": host.get("env", ""),
        "overall": _inspection_overall(checks),
        "checks": checks,
        "metrics": metrics,
        "partitions": partitions,
        "metrics_source": "ssh_python",
    }


async def _collect_inspection_fallbacks(hosts: list[dict]) -> dict[str, dict]:
    """Prometheus 缺失时，通过后端 Python 并发 SSH 到主机采集指标兜底。"""
    if not hosts:
        return {}

    sem = asyncio.Semaphore(_HOST_INSPECT_FALLBACK_CONCURRENCY)

    async def _one(host: dict) -> tuple[str, dict]:
        ip = host.get("ip", "")
        async with sem:
            try:
                info = await _ssh_sync(host)
                result = _sync_info_to_inspection_result(host, info)
                if result:
                    return ip, {"ok": True, "result": result}
                return ip, {"ok": False, "error": "SSH 连接成功，但未采集到服务器指标"}
            except Exception as exc:
                return ip, {"ok": False, "error": _ssh_error_msg(exc, host)}

    pairs = await asyncio.gather(*(_one(h) for h in hosts), return_exceptions=True)
    collected: dict[str, dict] = {}
    for item in pairs:
        if isinstance(item, Exception):
            continue
        ip, payload = item
        if ip:
            collected[ip] = payload
    return collected


@router.get("/api/hosts/inspect")
async def inspect_all_hosts(group_id: Optional[str] = Query(None)):
    """巡检主机 — 以 CMDB 为基准，有 Prometheus 指标的正常巡检，无指标的标记为离线。"""
    async def generate():
        try:
            group_name: str = ""
            all_hosts  = load_hosts_list()

            # 确定本次巡检的目标主机列表
            if group_id:
                groups_list = load_groups()
                g = next((g for g in groups_list if g["id"] == group_id), None)
                group_name = g["name"] if g else group_id
                target_hosts = [h for h in all_hosts if h.get("group") == group_id and h.get("ip")]
                if not target_hosts:
                    yield f"data: {json.dumps({'type': 'error', 'message': f'分组「{group_name}」下没有主机'}, ensure_ascii=False)}\n\n"
                    yield "data: [DONE]\n\n"
                    return
            else:
                target_hosts = [h for h in all_hosts if h.get("ip")]
                if not target_hosts:
                    yield f"data: {json.dumps({'type': 'error', 'message': 'CMDB 中没有主机，请先录入主机'}, ensure_ascii=False)}\n\n"
                    yield "data: [DONE]\n\n"
                    return

            # 从 Prometheus 获取全量指标（不按 instance 过滤，因为 node-exporter 的 instance 是主机名而非 ip:port）
            try:
                prom_results = await prom.inspect_hosts(instances=None)
            except Exception:
                prom_results = []

            # 建立 ip → prom_result 映射（通过 Prometheus 返回的 ip 字段匹配）
            prom_map: dict[str, dict] = {}
            for r in prom_results:
                ip = r.get("ip", "")
                if ip:
                    prom_map[ip] = r

            missing_prom_hosts = [h for h in target_hosts if h.get("ip") not in prom_map]
            fallback_map = await _collect_inspection_fallbacks(missing_prom_hosts)

            # 合并：CMDB 主机为基准
            results = []
            for h in target_hosts:
                ip = h["ip"]
                if ip in prom_map:
                    # 有 Prometheus 数据：使用巡检结果，补全 CMDB 信息
                    r = dict(prom_map[ip])
                    r["hostname"] = h.get("hostname") or r.get("hostname") or ip
                    r["os"]       = r.get("os") or h.get("os_version", "")
                    r["cpu_cores"] = h.get("cpu_cores")
                    r["memory_gb"] = h.get("memory_gb")
                    r["disk_gb"]   = h.get("disk_gb")
                    r["group"]    = h.get("group", "")
                    r["env"]      = h.get("env", "")
                    results.append(r)
                elif fallback_map.get(ip, {}).get("ok") and fallback_map[ip].get("result"):
                    results.append(fallback_map[ip]["result"])
                else:
                    fallback_error = fallback_map.get(ip, {}).get("error", "")
                    # 无 Prometheus 数据：以"无数据"状态显示，确保主机出现在报告中
                    results.append({
                        "instance": f"{ip}:9100",
                        "ip":       ip,
                        "hostname": h.get("hostname") or ip,
                        "os":       h.get("os_version", ""),
                        "cpu_cores": h.get("cpu_cores"),
                        "memory_gb": h.get("memory_gb"),
                        "disk_gb":   h.get("disk_gb"),
                        "job":      "",
                        "state":    "offline",
                        "group":    h.get("group", ""),
                        "env":      h.get("env", ""),
                        "overall":  "warning",
                        "checks": [{
                            "item":      "Prometheus 数据",
                            "value":     "无指标",
                            "status":    "warning",
                            "threshold": "需要安装 node_exporter 或检查 Prometheus 配置",
                        }, {
                            "item":      "SSH/Python 兜底",
                            "value":     fallback_error or "未获取到指标",
                            "status":    "warning",
                            "threshold": "请检查 SSH 凭证、网络连通性和目标主机 Python/shell 环境",
                        }],
                        "metrics":    {},
                        "partitions": [],
                        "metrics_source": "missing",
                    })

            # 把 Prometheus 指标回写到 hosts.json，让 CMDB 列表 / 健康总览 / 卡片墙
            # 同步看到最新的 cpu_usage_pct / memory_usage_pct / load5 / disk_usage
            now = _NOW()
            hosts_by_ip = {h.get("ip"): h for h in all_hosts if h.get("ip")}
            mutated = False
            metrics_updated_count = 0
            for r in results:
                host = hosts_by_ip.get(r.get("ip"))
                if not host:
                    continue
                host_changed = False
                m = r.get("metrics") or {}

                if r.get("metrics_source") == "ssh_python":
                    for src, dst in (
                        ("hostname", "hostname"),
                        ("os_version", "os_version"),
                        ("cpu_cores", "cpu_cores"),
                        ("memory_gb", "memory_gb"),
                        ("disk_gb", "disk_gb"),
                    ):
                        value = r.get(src)
                        if value not in (None, ""):
                            host[dst] = value
                            host_changed = True

                cpu_usage = _metric_float(m, "cpu_usage")
                if cpu_usage is not None:
                    host["cpu_usage_pct"] = round(cpu_usage, 1)
                    host_changed = True
                mem_usage = _metric_float(m, "mem_usage")
                if mem_usage is not None:
                    host["memory_usage_pct"] = round(mem_usage, 1)
                    host_changed = True
                load5 = _metric_float(m, "load5")
                if load5 is not None:
                    host["load5"] = round(load5, 2)
                    host_changed = True
                load1 = _metric_float(m, "load1")
                if load1 is not None:
                    host["load1"] = round(load1, 2)
                    host_changed = True
                load15 = _metric_float(m, "load15")
                if load15 is not None:
                    host["load15"] = round(load15, 2)
                    host_changed = True
                uptime_seconds = _metric_int(m, "uptime_seconds")
                if uptime_seconds is not None:
                    host["uptime_seconds"] = uptime_seconds
                    host["uptime_text"] = _format_uptime(uptime_seconds)
                    host_changed = True
                tcp_estab = _metric_int(m, "tcp_estab")
                if tcp_estab is not None:
                    host["tcp_connections"] = tcp_estab
                    host_changed = True
                tcp_tw = _metric_int(m, "tcp_tw")
                if tcp_tw is not None:
                    host["tcp_time_wait"] = tcp_tw
                    host_changed = True
                net_recv = _metric_float(m, "net_recv_mbps")
                if net_recv is not None:
                    host["network_rx_bps"] = int(net_recv * 1024 * 1024 / 8)
                    host_changed = True
                net_send = _metric_float(m, "net_send_mbps")
                if net_send is not None:
                    host["network_tx_bps"] = int(net_send * 1024 * 1024 / 8)
                    host_changed = True
                disk_read = _metric_float(m, "disk_read_mbps")
                if disk_read is not None:
                    host["disk_io_read_bps"] = int(disk_read * 1024 * 1024 / 8)
                    host_changed = True
                disk_write = _metric_float(m, "disk_write_mbps")
                if disk_write is not None:
                    host["disk_io_write_bps"] = int(disk_write * 1024 * 1024 / 8)
                    host_changed = True
                if r.get("partitions"):
                    host["disk_usage"] = [{
                        "mount": p.get("mountpoint") or p.get("mount") or "",
                        "size_gb": round(float(p.get("total_gb") or p.get("size_gb") or 0), 1),
                        "used_gb": round(float(p.get("used_gb") or 0), 1),
                        "avail_gb": round(float(p.get("avail_gb") or 0), 1),
                        "used_pct": round(float(p.get("used_pct") or p.get("usage_pct") or 0), 1),
                    } for p in r["partitions"] if p.get("mountpoint") or p.get("mount")]
                    host_changed = True
                if host_changed:
                    host["metrics_updated_at"] = now
                    host["last_sync_at"] = now
                    metrics_updated_count += 1
                    mutated = True
            if mutated:
                save_hosts_list(all_hosts)

            # 严重 → 警告 → 正常排序
            sev = {"critical": 2, "warning": 1, "normal": 0}
            results.sort(key=lambda x: -sev.get(x["overall"], 0))

            summary = {
                "total":      len(results),
                "normal":     sum(1 for r in results if r["overall"] == "normal"),
                "warning":    sum(1 for r in results if r["overall"] == "warning"),
                "critical":   sum(1 for r in results if r["overall"] == "critical"),
                "group_id":   group_id or "",
                "group_name": group_name,
                "metrics_updated_at": now if metrics_updated_count else "",
                "metrics_updated_count": metrics_updated_count,
                "metrics_missing_count": sum(1 for r in results if not (r.get("metrics") or r.get("partitions"))),
                "metrics_fallback_count": sum(1 for r in results if r.get("metrics_source") == "ssh_python"),
            }
            yield f"data: {json.dumps({'type': 'inspect_data', 'data': results, 'summary': summary}, ensure_ascii=False)}\n\n"
        except Exception as exc:
            logger.exception("巡检 SSE 流异常")
            yield f"data: {json.dumps({'type': 'error', 'message': str(exc)}, ensure_ascii=False)}\n\n"
        finally:
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class InspectAIRequest(BaseModel):
    results: list
    summary: dict


@router.post("/api/hosts/inspect/ai")
async def inspect_ai_summary(req: InspectAIRequest):
    """对已有巡检结果做 AI 流式分析（按需调用）。"""
    async def generate():
        provider_name = analyzer.provider_name
        yield f"data: {json.dumps({'type': 'ai_meta', 'provider': provider_name, 'fallback': False}, ensure_ascii=False)}\n\n"
        try:
            async for chunk in analyzer.generate_inspection_summary(req.results, req.summary):
                yield f"data: {json.dumps({'type': 'ai_chunk', 'text': chunk}, ensure_ascii=False)}\n\n"
        except Exception as exc:
            logger.exception("巡检 AI 流式总结失败，降级为规则摘要")
            fallback = _build_inspection_fallback_summary(req.results, req.summary, exc)
            yield f"data: {json.dumps({'type': 'ai_meta', 'provider': provider_name, 'fallback': True}, ensure_ascii=False)}\n\n"
            yield f"data: {json.dumps({'type': 'ai_chunk', 'text': fallback}, ensure_ascii=False)}\n\n"
        finally:
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class NotifyGroupsRequest(BaseModel):
    results: list
    summary: dict = {}
    ai_text: str = ""
    group_id: str = ""


@router.post("/api/hosts/inspect/notify-groups")
async def notify_groups_inspect(req: NotifyGroupsRequest):
    """将巡检结果按分组推送到飞书。"""
    try:
        from scheduler import _send_group_inspect_notifications
        results = await _send_group_inspect_notifications(
            req.results,
            force=True,
            group_id=(req.group_id or "").strip(),
        )
        sent    = [item for item in results if item.get("push", {}).get("ok")]
        failed  = [item for item in results if item.get("push") and not item.get("push", {}).get("ok")]
        skipped = [item for item in results if item.get("skipped")]
        target_group_name = next(
            (item.get("group_name") or item.get("group_id") for item in results
             if item.get("group_name") or item.get("group_id")),
            (req.group_id or "").strip(),
        )
        if sent:
            message = f"已推送分组「{target_group_name}」到飞书" if req.group_id else f"已推送 {len(sent)} 个分组到飞书"
            if failed:
                message += f"，{len(failed)} 个分组失败"
        else:
            message = (
                f"分组「{target_group_name}」没有成功推送，请检查巡检数据和飞书 Webhook 配置"
                if req.group_id else "没有成功推送的分组，请检查分组主机和飞书 Webhook 配置"
            )
        return {
            "ok": bool(sent),
            "message": message,
            "summary": {"sent": len(sent), "failed": len(failed), "skipped": len(skipped)},
            "results": results,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class InspectExcelRequest(BaseModel):
    results: list
    summary: dict
    ai_text: str = ""


@router.post("/api/hosts/inspect/excel")
async def export_inspect_excel(req: InspectExcelRequest):
    """根据传入的巡检结果生成 Excel 文件并下载。"""
    try:
        results = req.results
        summary = req.summary
        ai_text = req.ai_text
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        wb = Workbook()

        STATUS_COLOR = {"normal": "FF67C23A", "warning": "FFE6A23C", "critical": "FFF56C6C"}
        STATUS_TEXT  = {"normal": "正常", "warning": "警告", "critical": "严重"}

        def hfill(color="FF1F3A5F"):
            return PatternFill("solid", fgColor=color)

        def tborder():
            s = Side(style="thin", color="FFCCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def header_row(ws, headers, row=1):
            for col, h in enumerate(headers, 1):
                c           = ws.cell(row=row, column=col, value=h)
                c.font      = Font(bold=True, color="FFFFFFFF", size=10)
                c.fill      = hfill()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border    = tborder()

        def data_cell(ws, row, col, value, status=None):
            c           = ws.cell(row=row, column=col, value=value)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = tborder()
            if status and status in STATUS_COLOR:
                c.fill = PatternFill("solid", fgColor=STATUS_COLOR[status])
                c.font = Font(color="FFFFFFFF", size=9)
            else:
                c.font = Font(size=9)

        ws1 = wb.active
        ws1.title = "巡检概况"
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 40

        def s1(r, k, v, bold=False):
            ck = ws1.cell(row=r, column=1, value=k)
            cv = ws1.cell(row=r, column=2, value=v)
            ck.font      = Font(bold=True, color="FFFFFFFF", size=10)
            ck.fill      = hfill("FF243850")
            cv.font      = Font(bold=bold, size=10)
            ck.alignment = cv.alignment = Alignment(vertical="center", wrap_text=True)
            ck.border    = cv.border    = tborder()
            ws1.row_dimensions[r].height = 18

        s1(1, "巡检时间",     now_str, bold=True)
        s1(2, "巡检主机总数", summary.get("total", 0))
        s1(3, "正常主机",     summary.get("normal", 0))
        s1(4, "警告主机",     summary.get("warning", 0))
        s1(5, "严重主机",     summary.get("critical", 0))
        s1(6, "已获取指标主机", summary.get("metrics_updated_count", 0))
        s1(7, "SSH/Python兜底主机", summary.get("metrics_fallback_count", 0))
        s1(8, "未获取指标主机", summary.get("metrics_missing_count", 0))

        issue_cnt: dict[str, int] = {}
        for r in results:
            for c in r.get("checks", []):
                if c.get("status") != "normal":
                    issue_cnt[c.get("item", "未知")] = issue_cnt.get(c.get("item", "未知"), 0) + 1
        top_issues = sorted(issue_cnt.items(), key=lambda x: x[1], reverse=True)[:8]

        ws1.cell(row=10, column=1, value="高频异常项").font = Font(bold=True, size=10)
        header_row(ws1, ["异常检查项", "影响主机数"], row=11)
        for i, (item, cnt) in enumerate(top_issues, 12):
            ws1.cell(row=i, column=1, value=item).border    = tborder()
            ws1.cell(row=i, column=2, value=cnt).border     = tborder()
            ws1.cell(row=i, column=1).font      = Font(size=9)
            ws1.cell(row=i, column=2).font      = Font(size=9)
            ws1.cell(row=i, column=1).alignment = Alignment(vertical="center")
            ws1.cell(row=i, column=2).alignment = Alignment(horizontal="center", vertical="center")

        ai_start = 12 + len(top_issues) + 2
        ws1.cell(row=ai_start, column=1, value="AI 分析总结").font = Font(bold=True, size=10)
        if ai_text:
            c           = ws1.cell(row=ai_start + 1, column=1, value=ai_text)
            c.font      = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border    = tborder()
            ws1.merge_cells(start_row=ai_start + 1, start_column=1, end_row=ai_start + 1, end_column=2)
            ws1.row_dimensions[ai_start + 1].height = max(60, len(ai_text) // 3)

        ws2 = wb.create_sheet("全部主机明细")
        headers2 = [
            "状态", "主机名", "IP", "系统",
            "CPU使用率(%)", "CPU核心", "内存使用率(%)", "内存总量(GB)",
            "负载1m", "负载5m", "负载15m", "运行时长",
            "磁盘挂载", "磁盘使用率(%)", "磁盘容量(已用/总量GB)",
            "网络入(MB/s)", "网络出(MB/s)", "磁盘读(MB/s)", "磁盘写(MB/s)",
            "TCP连接", "TIME_WAIT", "异常项",
        ]
        header_row(ws2, headers2)
        col_widths(ws2, [10, 18, 16, 22, 14, 10, 14, 14, 10, 10, 10, 14, 14, 14, 18, 14, 14, 14, 14, 12, 12, 36])
        ws2.row_dimensions[1].height = 22

        def fmt(v, decimals=1):
            return round(v, decimals) if v is not None else "-"

        def number_or_none(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        def highest_partition(partitions: list[dict]) -> dict | None:
            best = None
            best_usage = -1.0
            for pt in partitions or []:
                usage = number_or_none(pt.get("usage_pct", pt.get("used_pct")))
                if usage is None:
                    continue
                if usage > best_usage:
                    best = pt
                    best_usage = usage
            return best

        def uptime_text(seconds) -> str:
            seconds_value = number_or_none(seconds)
            if seconds_value is None or seconds_value < 0:
                return "-"
            days = int(seconds_value // 86400)
            hours = int((seconds_value % 86400) // 3600)
            minutes = int((seconds_value % 3600) // 60)
            return f"{days}天{hours}小时{minutes}分钟" if days else f"{hours}小时{minutes}分钟"

        def disk_capacity_text(pt: dict | None) -> str:
            if not pt:
                return "-"
            used = number_or_none(pt.get("used_gb"))
            total = number_or_none(pt.get("total_gb", pt.get("size_gb")))
            if used is None and total is None:
                return "-"
            if used is None:
                return f"- / {round(total, 1)}"
            if total is None:
                return f"{round(used, 1)} / -"
            return f"{round(used, 1)} / {round(total, 1)}"

        def issue_text(checks: list[dict]) -> str:
            abnormal = [c for c in checks or [] if c.get("status") != "normal"]
            if not abnormal:
                return "全部正常"
            return "；".join(
                f"{c.get('item', '-')}: {c.get('value', '-')}"
                for c in abnormal
            )

        for ri, h in enumerate(results, 2):
            overall      = h.get("overall", "normal")
            status_label = STATUS_TEXT.get(overall, overall)

            m           = h.get("metrics") or {}
            partitions = h.get("partitions") or []
            disk = highest_partition(partitions)
            disk_usage = disk.get("usage_pct", disk.get("used_pct")) if disk else None

            row_vals = [
                status_label, h.get("hostname", "-"), h.get("ip", "-"), h.get("os", "-"),
                fmt(m.get("cpu_usage")), h.get("cpu_cores") or "-", fmt(m.get("mem_usage")), fmt(m.get("mem_total_gb") or h.get("memory_gb")),
                fmt(m.get("load1")), fmt(m.get("load5")), fmt(m.get("load15")), uptime_text(m.get("uptime_seconds")),
                (disk.get("mountpoint") or disk.get("mount") or "-") if disk else "-",
                fmt(disk_usage),
                disk_capacity_text(disk),
                fmt(m.get("net_recv_mbps")), fmt(m.get("net_send_mbps")), fmt(m.get("disk_read_mbps")), fmt(m.get("disk_write_mbps")),
                m.get("tcp_estab") if m.get("tcp_estab") is not None else "-",
                m.get("tcp_tw") if m.get("tcp_tw") is not None else "-",
                issue_text(h.get("checks") or []),
            ]
            for ci, val in enumerate(row_vals, 1):
                data_cell(ws2, ri, ci, val, overall if ci == 1 else None)
            ws2.row_dimensions[ri].height = 16

        ws2.freeze_panes = "A2"

        ws3 = wb.create_sheet("异常项明细")
        headers3 = ["主机名", "IP地址", "检查项", "当前值", "状态", "阈值说明"]
        header_row(ws3, headers3)
        col_widths(ws3, [18, 16, 20, 20, 10, 30])
        ws3.row_dimensions[1].height = 22

        ri3 = 2
        for h in results:
            for ck in [c for c in h.get("checks", []) if c.get("status") != "normal"]:
                status = ck.get("status", "warning")
                row3 = [
                    h.get("hostname", "-"), h.get("ip", "-"),
                    ck.get("item", "-"), ck.get("value", "-"),
                    STATUS_TEXT.get(status, status), ck.get("threshold", "-"),
                ]
                for ci, val in enumerate(row3, 1):
                    data_cell(ws3, ri3, ci, val, status if ci == 5 else None)
                ws3.row_dimensions[ri3].height = 16
                ri3 += 1

        ws3.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename     = f"主机巡检_{now_str.replace(':', '-')}.xlsx"
        encoded_name = quote(filename, encoding="utf-8")
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
        )
    except Exception as e:
        logger.exception("巡检 Excel 导出失败")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/hosts/{host_id}/inspect")
async def inspect_single_host(host_id: str):
    """巡检单台主机（按 CMDB ID 查找 IP，再查 Prometheus）。"""
    hosts = load_hosts_list()
    host = next((h for h in hosts if h.get("id") == host_id), None)
    if not host:
        raise HTTPException(status_code=404, detail="主机不存在")
    ip = host.get("ip", "")
    if not ip:
        raise HTTPException(status_code=400, detail="主机没有配置 IP")
    try:
        instance = f"{ip}:9100"
        results = await prom.inspect_hosts(instances=[instance])
        if not results:
            raise HTTPException(status_code=404, detail="Prometheus 中未找到该主机指标（请确认 node_exporter 已运行）")
        return results[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))
