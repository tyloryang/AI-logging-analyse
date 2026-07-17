"""运维日报 / 巡检日报路由。

路由前缀：/api/report/*
"""
import asyncio
import json
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import Optional

from json_snapshot_store import read_json_file, write_json_file
from state import (
    loki, prom, analyzer,
    REPORTS_DIR,
    FEISHU_WEBHOOK, FEISHU_KEYWORD,
    DINGTALK_WEBHOOK, DINGTALK_KEYWORD,
    APP_URL,
    load_slowlog_targets, save_slowlog_targets,
    load_groups, load_cmdb,
)
from notifier import send_feishu, send_dingtalk, send_feishu_group_inspect
from report_builder import collect_daily_data, collect_inspect_data, build_inspect_meta
from report_store import save_report_meta, list_report_meta

logger = logging.getLogger(__name__)
router = APIRouter()


def _slowlog_ai_timeout_seconds() -> int:
    try:
        return max(30, int(os.getenv("SLOWLOG_AI_TIMEOUT_SECONDS", "120")))
    except (TypeError, ValueError):
        return 120


async def _stream_ai_with_timeout(prompt: str, *, max_tokens: int, timeout_seconds: int):
    iterator = analyzer.provider.stream(prompt, max_tokens=max_tokens).__aiter__()
    deadline = asyncio.get_running_loop().time() + timeout_seconds
    try:
        while True:
            remaining = deadline - asyncio.get_running_loop().time()
            if remaining <= 0:
                raise asyncio.TimeoutError
            try:
                chunk = await asyncio.wait_for(iterator.__anext__(), timeout=remaining)
            except StopAsyncIteration:
                break
            yield chunk
    finally:
        close = getattr(iterator, "aclose", None)
        if close:
            try:
                await close()
            except Exception:
                pass


def _read_report_json(path):
    data = read_json_file(path, default=None)
    return data if isinstance(data, dict) else None


def _write_report_json(path, data: dict) -> None:
    write_json_file(path, data, ensure_parent=True)


# ── 运维日报 ──────────────────────────────────────────────────────────────────

class ReportMeta(BaseModel):
    id: str
    title: str
    created_at: str
    health_score: int
    total_logs: int
    total_errors: int
    service_count: int
    active_alerts: int


@router.get("/api/report/generate")
async def generate_report():
    """触发生成运维日报，流式返回 AI 分析内容"""
    try:
        data = await collect_daily_data()
        meta = data["meta"]

        report_path    = REPORTS_DIR / f"{meta['id']}.json"
        ai_content_parts: list[str] = []

        async def generate():
            yield f"data: __META__{json.dumps(meta, ensure_ascii=False)}\n\n"
            try:
                async for chunk in analyzer.generate_daily_report(
                    error_counts=data["error_counts"],
                    total_logs=data["total_logs"],
                    service_count=data["service_count"],
                    node_status=data["node_status"],
                    active_alerts=data["active_alerts"],
                    sample_errors=data["error_logs"],
                    service_error_summaries=data["service_error_summaries"],
                    error_keywords=data["error_keywords"],
                    interface_status=data["interface_status"],
                ):
                    ai_content_parts.append(chunk)
                    yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
            except Exception as exc:
                logger.exception("日报 AI 生成异常")
                ai_content_parts.append(f"\n[AI生成出错] {exc}")
                yield f"data: {json.dumps('[AI生成出错] ' + str(exc), ensure_ascii=False)}\n\n"
            meta["ai_analysis"] = "".join(ai_content_parts)
            _write_report_json(report_path, meta)
            asyncio.create_task(save_report_meta(meta))
            # 后台写入 Milvus（不阻塞 SSE 流）
            try:
                from agent.report_memory import get_report_memory
                asyncio.create_task(get_report_memory().save(meta))
            except Exception:
                pass
            yield "data: [DONE]\n\n"

        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/report/list")
async def list_reports():
    """历史报告列表（从 DB 查询元数据，不再逐文件读取）"""
    rows = await list_report_meta(limit=50)
    # DB 为空时降级到文件扫描（首次启动 sync 尚未完成的容错）
    if not rows:
        for p in sorted(REPORTS_DIR.glob("*.json"), reverse=True)[:50]:
            try:
                data = _read_report_json(p)
                if not data:
                    continue
                rows.append({k: data[k] for k in [
                    "id", "type", "title", "created_at", "health_score",
                    "total_logs", "total_errors", "service_count", "active_alerts",
                    "total_queries", "alert_queries", "hosts_count",
                ] if k in data})
            except Exception:
                pass
    return {"data": rows}


# ── 巡检日报 ──────────────────────────────────────────────────────────────────

UNGROUPED_GROUP_ID = "__ungrouped__"
UNGROUPED_GROUP_NAME = "未分组"


def _build_inspect_summary(results: list[dict], group_name: str = "") -> dict:
    total = len(results)
    summary = {
        "total": total,
        "normal": sum(1 for r in results if r.get("overall") == "normal"),
        "warning": sum(1 for r in results if r.get("overall") == "warning"),
        "critical": sum(1 for r in results if r.get("overall") == "critical"),
        "cmdb_total": total,
        "prometheus_extra_count": 0,
        "scope": "cmdb",
    }
    summary["scope_note"] = f"统计口径：按 CMDB 主机 {total} 台统计，正常/警告/严重数量均以 CMDB 主机为基准。"
    if group_name:
        summary["group_name"] = group_name
    return summary


def _build_inspect_top_issues(results: list[dict]) -> list[dict]:
    issue_cnt: dict[str, int] = {}
    for r in results:
        for c in r.get("checks", []):
            if c.get("status") != "normal":
                item = c.get("item", "未知")
                issue_cnt[item] = issue_cnt.get(item, 0) + 1
    return [{"item": k, "count": v} for k, v in sorted(issue_cnt.items(), key=lambda x: x[1], reverse=True)[:10]]


def _sort_abnormal_hosts(results: list[dict]) -> list[dict]:
    abnormal_hosts = [r for r in results if r.get("overall") != "normal"]
    abnormal_hosts.sort(
        key=lambda r: {"critical": 2, "warning": 1}.get(r.get("overall", "normal"), 0),
        reverse=True,
    )
    return abnormal_hosts


def _append_missing_cmdb_hosts(results: list[dict], group_id: str = "") -> list[dict]:
    cmdb = load_cmdb()
    expected = [
        {"ip": ip, **meta}
        for ip, meta in cmdb.items()
        if not group_id or meta.get("group") == group_id
    ]
    found_ips = {
        (r.get("ip") or r.get("instance", "").split(":")[0]).strip()
        for r in results
        if (r.get("ip") or r.get("instance"))
    }
    merged = list(results)
    for host in expected:
        ip = (host.get("ip") or "").strip()
        if not ip or ip in found_ips:
            continue
        merged.append({
            "instance": ip,
            "ip": ip,
            "hostname": host.get("hostname") or ip,
            "os": host.get("os") or host.get("platform", ""),
            "job": "",
            "state": "missing",
            "overall": "critical",
            "checks": [{
                "item": "Prometheus 监控",
                "value": "未发现该主机监控数据",
                "status": "critical",
                "threshold": "应存在可抓取 target",
            }],
            "metrics": {},
            "partitions": [],
        })
    merged.sort(
        key=lambda r: {"critical": 2, "warning": 1}.get(r.get("overall", "normal"), 0),
        reverse=True,
    )
    return merged


def _inspect_health_score(summary: dict) -> int:
    total = summary.get("total", 0)
    if total <= 0:
        return 100
    score = 100
    score -= min(40, int(summary.get("critical", 0) / total * 200))
    score -= min(30, int(summary.get("warning", 0) / total * 100))
    return max(0, score)


def _group_inspect_results(results: list[dict]) -> list[dict]:
    groups = load_groups()
    group_order = {g["id"]: idx for idx, g in enumerate(groups)}
    group_names = {g["id"]: g.get("name", g["id"]) for g in groups}
    cmdb = load_cmdb()

    grouped: dict[str, dict] = {}
    for result in results:
        ip = (result.get("ip") or result.get("instance", "").split(":")[0]).strip()
        host_meta = cmdb.get(ip, {})
        raw_group_id = (host_meta.get("group") or "").strip()
        key = raw_group_id or UNGROUPED_GROUP_ID
        group_name = group_names.get(raw_group_id, raw_group_id) if raw_group_id else UNGROUPED_GROUP_NAME
        bucket = grouped.setdefault(key, {
            "group_id": "" if key == UNGROUPED_GROUP_ID else key,
            "group_name": group_name,
            "sort_index": group_order.get(raw_group_id, len(group_order) + (1 if raw_group_id else 2)),
            "results": [],
        })
        bucket["results"].append(result)

    return sorted(grouped.values(), key=lambda item: (item["sort_index"], item["group_name"]))


def _group_ai_fallback(group_name: str, summary: dict) -> str:
    return (
        f"分组「{group_name}」共 {summary.get('total', 0)} 台主机，"
        f"正常 {summary.get('normal', 0)} 台、告警 {summary.get('warning', 0)} 台、"
        f"严重 {summary.get('critical', 0)} 台。"
        "建议优先处理严重主机和高频异常项。"
    )


async def _build_group_analyses(grouped_results: list[dict]) -> list[dict]:
    analyses = []
    for item in grouped_results:
        summary = _build_inspect_summary(item["results"], item["group_name"])
        ai_parts: list[str] = []
        try:
            async for chunk in analyzer.generate_inspection_summary(item["results"], summary):
                ai_parts.append(chunk)
            ai_text = "".join(ai_parts).strip()
        except Exception as exc:
            logger.warning("分组 %s AI 分析生成失败: %s", item["group_name"], exc)
            ai_text = _group_ai_fallback(item["group_name"], summary)

        analyses.append({
            "group_id": item["group_id"],
            "group_name": item["group_name"],
            "host_summary": summary,
            "top_issues": _build_inspect_top_issues(item["results"]),
            "abnormal_hosts": _sort_abnormal_hosts(item["results"])[:20],
            "health_score": _inspect_health_score(summary),
            "ai_analysis": ai_text,
        })
    return analyses


def _compose_group_ai_text(group_analyses: list[dict]) -> str:
    sections: list[str] = []
    for item in group_analyses:
        summary = item.get("host_summary", {})
        header = (
            f"【分组：{item.get('group_name') or UNGROUPED_GROUP_NAME}】\n"
            f"主机 {summary.get('total', 0)} 台，正常 {summary.get('normal', 0)} 台，"
            f"告警 {summary.get('warning', 0)} 台，严重 {summary.get('critical', 0)} 台"
        )
        body = (item.get("ai_analysis") or "").strip() or "暂无 AI 分析结果"
        sections.append(f"{header}\n{body}")
    return "\n\n".join(sections).strip()


def _build_group_report_from_inspect(report: dict, group_analysis: dict, group_results: list[dict]) -> dict:
    created_at = report.get("created_at", "")
    date_text = created_at[:10] if created_at else ""
    group_name = group_analysis.get("group_name") or UNGROUPED_GROUP_NAME
    summary = group_analysis.get("host_summary") or _build_inspect_summary(group_results, group_name)
    return {
        "id": report.get("id", ""),
        "type": "inspect",
        "group_id": group_analysis.get("group_id", ""),
        "group_name": group_name,
        "title": f"主机巡检日报【{group_name}】 {date_text}".strip(),
        "created_at": created_at,
        "health_score": group_analysis.get("health_score", _inspect_health_score(summary)),
        "host_summary": summary,
        "top_issues": group_analysis.get("top_issues") or _build_inspect_top_issues(group_results),
        "abnormal_hosts": group_analysis.get("abnormal_hosts") or _sort_abnormal_hosts(group_results)[:20],
        "all_hosts": group_results,
        "prometheus_extra_hosts": [],
        "prometheus_extra_count": 0,
        "summary_scope_note": summary.get("scope_note", ""),
        "ai_analysis": (group_analysis.get("ai_analysis") or "").strip() or report.get("ai_analysis", ""),
        "group_analyses": [group_analysis],
    }


def _split_inspect_report_by_group(report: dict) -> dict[str, dict]:
    if report.get("group_id"):
        return {report.get("group_id", ""): report}

    all_hosts = report.get("all_hosts", [])
    if not all_hosts:
        return {}

    grouped_results = _group_inspect_results(all_hosts)
    analyses = {
        (item.get("group_id") or UNGROUPED_GROUP_ID): item
        for item in (report.get("group_analyses") or [])
    }

    split_reports: dict[str, dict] = {}
    for item in grouped_results:
        key = item["group_id"] or UNGROUPED_GROUP_ID
        analysis = analyses.get(key, {
            "group_id": item["group_id"],
            "group_name": item["group_name"],
            "host_summary": _build_inspect_summary(item["results"], item["group_name"]),
            "top_issues": _build_inspect_top_issues(item["results"]),
            "abnormal_hosts": _sort_abnormal_hosts(item["results"])[:20],
            "health_score": _inspect_health_score(_build_inspect_summary(item["results"], item["group_name"])),
            "ai_analysis": "",
        })
        split_reports[item["group_id"]] = _build_group_report_from_inspect(report, analysis, item["results"])
    return split_reports


@router.get("/api/report/inspect/generate")
async def generate_inspect_report(group_id: Optional[str] = Query(None)):
    """生成主机巡检日报，流式返回 AI 分析；可通过 group_id 限定分组"""
    try:
        group_name = ""
        if group_id:
            groups = load_groups()
            g = next((g for g in groups if g["id"] == group_id), None)
            group_name = g["name"] if g else group_id
        data = await collect_inspect_data(group_id=group_id or "", group_name=group_name)
        results = data["results"]
        summary = data["summary"]
        top_issues = data["top_issues"]
        abnormal_hosts = data["abnormal_hosts"]
        all_hosts_brief = data["all_hosts"]
        health_score = data["health_score"]
        meta = build_inspect_meta(data, group_id or "", group_name)

        report_path = REPORTS_DIR / f"{meta['id']}.json"
        async def generate():
            yield f"data: __META__{json.dumps(meta, ensure_ascii=False)}\n\n"
            if group_id:
                ai_parts: list[str] = []
                try:
                    async for chunk in analyzer.generate_host_inspect_report(results, summary):
                        ai_parts.append(chunk)
                        yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
                except Exception as exc:
                    logger.exception("巡检日报 AI 生成异常")
                    ai_parts.append(f"\n[AI生成出错] {exc}")
                    yield f"data: {json.dumps('[AI生成出错] ' + str(exc), ensure_ascii=False)}\n\n"
                ai_text = "".join(ai_parts).strip()
                meta["group_analyses"] = [{
                    "group_id": group_id or "",
                    "group_name": group_name or UNGROUPED_GROUP_NAME,
                    "host_summary": summary,
                    "top_issues": top_issues,
                    "abnormal_hosts": abnormal_hosts[:20],
                    "health_score": health_score,
                    "ai_analysis": ai_text,
                }]
                meta["ai_analysis"] = ai_text
            else:
                grouped_results = _group_inspect_results(all_hosts_brief)
                if not grouped_results:
                    meta["group_analyses"] = []
                    meta["ai_analysis"] = "未找到可用于分组分析的主机数据。"
                    yield f"data: {json.dumps(meta['ai_analysis'], ensure_ascii=False)}\n\n"
                else:
                    group_analyses = await _build_group_analyses(grouped_results)
                    combined_text = _compose_group_ai_text(group_analyses)
                    meta["group_analyses"] = group_analyses
                    meta["ai_analysis"] = combined_text
                    for chunk in [combined_text[i:i + 800] for i in range(0, len(combined_text), 800)]:
                        yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
            _write_report_json(report_path, meta)
            asyncio.create_task(save_report_meta(meta))
            # 后台写入 Milvus
            try:
                from agent.report_memory import get_report_memory
                asyncio.create_task(get_report_memory().save(meta))
            except Exception:
                pass
            yield "data: [DONE]\n\n"

        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _find_latest_group_inspect_report(group_id: str) -> Optional[dict]:
    """在 REPORTS_DIR 中找到指定分组最新的、已含 AI 分析的巡检报告，没有则返回 None"""
    best = None
    for p in sorted(REPORTS_DIR.glob("inspect_*.json"), reverse=True):
        try:
            data = _read_report_json(p)
            if not data:
                continue
            if data.get("type") == "inspect" and data.get("group_id") == group_id and data.get("ai_analysis"):
                best = data
                break   # 按文件名倒序，第一个即最新
        except Exception:
            continue
    return best


@router.post("/api/report/inspect/generate-groups")
async def generate_inspect_report_all_groups():
    """为每个配置了飞书 webhook 的分组生成巡检报告并推送。
    优先复用已有 AI 分析的最新报告；若不存在则重新采集并生成 AI 后推送。"""
    groups = load_groups()
    cmdb   = load_cmdb()

    results = []
    for group in groups:
        webhook = group.get("feishu_webhook", "")
        if not webhook:
            results.append({"group_id": group["id"], "group_name": group.get("name", ""), "skipped": True, "reason": "未配置飞书 webhook"})
            continue

        group_id   = group["id"]
        group_name = group.get("name", group_id)
        keyword    = group.get("feishu_keyword", "")

        instances = [inst for inst, meta in cmdb.items() if meta.get("group") == group_id]
        if not instances:
            results.append({"group_id": group_id, "group_name": group_name, "skipped": True, "reason": "该分组无主机"})
            continue

        try:
            # ① 优先使用已有报告（有 AI 分析）
            existing = _find_latest_group_inspect_report(group_id)
            if existing:
                host_results = existing.get("all_hosts", [])
                ai_text      = existing.get("ai_analysis", "")
                source       = "cached"
            else:
                # ② 重新采集 + 生成 AI + 保存报告
                inspect_data = await collect_inspect_data(
                    instances=instances,
                    group_id=group_id,
                    group_name=group_name,
                )
                full_results = inspect_data["results"]
                summary = inspect_data["summary"]
                ai_parts: list[str] = []
                try:
                    async for chunk in analyzer.generate_host_inspect_report(full_results, summary):
                        ai_parts.append(chunk)
                except Exception as ai_exc:
                    logger.warning("分组 %s AI生成失败: %s", group_name, ai_exc)
                ai_text = "".join(ai_parts)
                meta_g = build_inspect_meta(inspect_data, group_id, group_name)
                meta_g["ai_analysis"] = ai_text
                meta_g["group_analyses"] = [{
                    "group_id": group_id,
                    "group_name": group_name,
                    "host_summary": summary,
                    "top_issues": inspect_data["top_issues"],
                    "abnormal_hosts": inspect_data["abnormal_hosts"],
                    "health_score": inspect_data["health_score"],
                    "ai_analysis": ai_text,
                }]
                try:
                    _write_report_json(REPORTS_DIR / f"{meta_g['id']}.json", meta_g)
                    asyncio.create_task(save_report_meta(meta_g))
                except Exception as save_exc:
                    logger.warning("分组 %s 报告保存失败: %s", group_name, save_exc)

                # 推送时使用完整巡检结果（含 checks 字段）
                host_results = full_results
                source = "generated"

            push_result = await send_feishu_group_inspect(
                group_name, host_results, webhook, keyword=keyword, ai_text=ai_text
            )
            results.append({
                "group_id":   group_id,
                "group_name": group_name,
                "hosts":      len(host_results),
                "source":     source,
                "push":       push_result,
            })
        except Exception as exc:
            logger.exception("分组 %s 巡检推送失败", group_name)
            results.append({"group_id": group_id, "group_name": group_name, "error": str(exc)})

    return {"results": results}


@router.get("/api/report/inspect/{report_id}/excel")
async def download_inspect_report_excel(report_id: str):
    """下载主机巡检日报 Excel 文件"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    data = _read_report_json(p)
    if not data:
        raise HTTPException(status_code=404, detail="报告内容不存在")
    if data.get("type") != "inspect":
        raise HTTPException(status_code=400, detail="该报告不是巡检类型")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()

        STATUS_COLOR = {"normal": "FF67C23A", "warning": "FFE6A23C", "critical": "FFF56C6C"}
        STATUS_TEXT  = {"normal": "正常", "warning": "警告", "critical": "严重"}

        def header_fill(color="FF1F3A5F"):
            return PatternFill("solid", fgColor=color)

        def thin_border():
            s = Side(style="thin", color="FFCCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def write_header_row(ws, headers, row=1):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font      = Font(bold=True, color="FFFFFFFF", size=10)
                c.fill      = header_fill()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border    = thin_border()

        def write_data_cell(ws, row, col, value, status=None):
            c           = ws.cell(row=row, column=col, value=value)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = thin_border()
            if status and status in STATUS_COLOR:
                c.fill = PatternFill("solid", fgColor=STATUS_COLOR[status])
                c.font = Font(color="FFFFFFFF", size=9)
            else:
                c.font = Font(size=9)
            return c

        summary_d  = data.get("host_summary", {})
        all_hosts  = data.get("all_hosts", [])
        top_issues = data.get("top_issues", [])
        ai_text    = data.get("ai_analysis", "")
        title      = data.get("title", "主机巡检日报")
        created    = data.get("created_at", "")[:19].replace("T", " ")

        # Sheet 1: 巡检概况
        ws1 = wb.active
        ws1.title = "巡检概况"
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 40

        def s1_row(r, k, v, bold=False):
            ck = ws1.cell(row=r, column=1, value=k)
            cv = ws1.cell(row=r, column=2, value=v)
            ck.font      = Font(bold=True, color="FFFFFFFF", size=10)
            ck.fill      = header_fill("FF243850")
            cv.font      = Font(bold=bold, size=10)
            ck.alignment = cv.alignment = Alignment(vertical="center", wrap_text=True)
            ck.border    = cv.border    = thin_border()
            ws1.row_dimensions[r].height = 18

        s1_row(1, "报告名称",     title, bold=True)
        s1_row(2, "生成时间",     created)
        s1_row(3, "健康评分",     f"{data.get('health_score', '-')}/100")
        s1_row(4, "巡检主机总数", summary_d.get("total", 0))
        s1_row(5, "正常主机",     summary_d.get("normal", 0))
        s1_row(6, "警告主机",     summary_d.get("warning", 0))
        s1_row(7, "严重主机",     summary_d.get("critical", 0))
        s1_row(8, "已获取指标主机", summary_d.get("metrics_updated_count", 0))
        s1_row(9, "SSH/Python兜底主机", summary_d.get("metrics_fallback_count", 0))
        s1_row(10, "未获取指标主机", summary_d.get("metrics_missing_count", 0))

        ws1.cell(row=12, column=1, value="高频异常项").font = Font(bold=True, size=10)
        write_header_row(ws1, ["异常检查项", "影响主机数"], row=13)
        for i, issue in enumerate(top_issues, 14):
            ws1.cell(row=i, column=1, value=issue.get("item", "")).border    = thin_border()
            ws1.cell(row=i, column=2, value=issue.get("count", 0)).border    = thin_border()
            ws1.cell(row=i, column=1).font      = Font(size=9)
            ws1.cell(row=i, column=2).font      = Font(size=9)
            ws1.cell(row=i, column=1).alignment = Alignment(vertical="center")
            ws1.cell(row=i, column=2).alignment = Alignment(horizontal="center", vertical="center")

        ai_start = 14 + len(top_issues) + 2
        ws1.cell(row=ai_start, column=1, value="AI 分析总结").font = Font(bold=True, size=10)
        if ai_text:
            c           = ws1.cell(row=ai_start + 1, column=1, value=ai_text)
            c.font      = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border    = thin_border()
            ws1.merge_cells(
                start_row=ai_start + 1, start_column=1,
                end_row=ai_start + 1,   end_column=2,
            )
            ws1.row_dimensions[ai_start + 1].height = max(60, len(ai_text) // 3)

        # Sheet 2: 全部主机明细
        ws2 = wb.create_sheet("全部主机明细")
        headers2 = [
            "分组", "状态", "服务器IP", "服务器", "主机名", "负责人", "机房",
            "CPU使用率(%)", "CPU核心", "内存使用率(%)", "内存总量(GB)",
            "负载1m", "负载5m", "负载15m", "运行时长",
            "磁盘具体占用",
            "网络入(MB/s)", "网络出(MB/s)", "磁盘读(MB/s)", "磁盘写(MB/s)",
            "TCP连接", "TIME_WAIT", "进程占用Top10", "异常项",
        ]
        write_header_row(ws2, headers2)
        set_col_widths(ws2, [16, 10, 16, 18, 18, 14, 16, 14, 10, 14, 14, 10, 10, 10, 14, 44, 14, 14, 14, 14, 12, 12, 52, 36])
        ws2.row_dimensions[1].height = 22

        def fmt(v, decimals=1):
            return round(v, decimals) if v is not None else "-"

        def number_or_none(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        def highest_partition(partitions: list[dict]) -> dict | None:
            best = None
            best_usage = -1.0
            for pt in partitions or []:
                usage = number_or_none(pt.get("usage_pct", pt.get("used_pct")))
                if usage is None:
                    continue
                if usage > best_usage:
                    best = pt
                    best_usage = usage
            return best

        def uptime_text(seconds) -> str:
            seconds_value = number_or_none(seconds)
            if seconds_value is None or seconds_value < 0:
                return "-"
            days = int(seconds_value // 86400)
            hours = int((seconds_value % 86400) // 3600)
            minutes = int((seconds_value % 3600) // 60)
            return f"{days}天{hours}小时{minutes}分钟" if days else f"{hours}小时{minutes}分钟"

        def disk_capacity_text(pt: dict | None) -> str:
            if not pt:
                return "-"
            used = number_or_none(pt.get("used_gb"))
            total = number_or_none(pt.get("total_gb", pt.get("size_gb")))
            if used is None and total is None:
                return "-"
            if used is None:
                return f"- / {round(total, 1)}"
            if total is None:
                return f"{round(used, 1)} / -"
            return f"{round(used, 1)} / {round(total, 1)}"

        def issue_text(checks: list[dict]) -> str:
            abnormal = [c for c in checks or [] if c.get("status") != "normal"]
            if not abnormal:
                return "全部正常"
            return "；".join(
                f"{c.get('item', '-')}: {c.get('value', '-')}"
                for c in abnormal
            )

        def partitions_text(partitions: list[dict]) -> str:
            return "\n".join(
                f"{pt.get('mountpoint') or pt.get('mount') or '-'} "
                f"{pt.get('usage_pct', pt.get('used_pct', '-'))}% "
                f"({pt.get('used_gb', '-')}/{pt.get('total_gb', pt.get('size_gb', '-'))}GB)"
                for pt in partitions or []
            ) or "未采集"

        def processes_text(processes: list[dict], error: str = "") -> str:
            return "\n".join(
                f"{idx}. {proc.get('service') or proc.get('comm') or '-'} "
                f"CPU {proc.get('cpu', '-')}% / MEM {proc.get('mem', '-')}%"
                for idx, proc in enumerate(processes or [], 1)
            ) or error or "未采集"

        for ri, h in enumerate(all_hosts, 2):
            overall      = h.get("overall", "normal")
            status_label = STATUS_TEXT.get(overall, overall)

            partitions = h.get("partitions") or []
            disk = highest_partition(partitions)
            disk_usage = disk.get("usage_pct", disk.get("used_pct")) if disk else None

            row_vals = [
                h.get("group_name") or "未分组", status_label, h.get("ip", "-"),
                h.get("role") or h.get("os") or "-", h.get("hostname", "-"),
                h.get("owner") or "未配置", h.get("datacenter") or "未配置",
                fmt(h.get("cpu_pct")), h.get("cpu_cores") or "-", fmt(h.get("mem_pct")), fmt(h.get("mem_total")),
                fmt(h.get("load1")), fmt(h.get("load5")), fmt(h.get("load15")), uptime_text(h.get("uptime_s")),
                partitions_text(partitions),
                fmt(h.get("net_recv")), fmt(h.get("net_send")), fmt(h.get("disk_read")), fmt(h.get("disk_write")),
                h.get("tcp_estab") if h.get("tcp_estab") is not None else "-",
                h.get("tcp_tw") if h.get("tcp_tw") is not None else "-",
                processes_text(h.get("process_top10") or [], h.get("process_error") or ""),
                issue_text(h.get("checks") or []),
            ]
            for ci, val in enumerate(row_vals, 1):
                st = overall if ci == 2 else None
                write_data_cell(ws2, ri, ci, val, st)
            ws2.row_dimensions[ri].height = max(30, min(150, 15 * max(len(partitions), len(h.get("process_top10") or []), 1)))

        ws2.freeze_panes = "A2"

        # Sheet 3: 异常项明细
        ws3 = wb.create_sheet("异常项明细")
        headers3 = ["主机名", "IP地址", "检查项", "当前值", "状态", "阈值说明"]
        write_header_row(ws3, headers3)
        set_col_widths(ws3, [18, 16, 20, 20, 10, 30])
        ws3.row_dimensions[1].height = 22

        ri3 = 2
        for h in all_hosts:
            for ck in [c for c in h.get("checks", []) if c.get("status") != "normal"]:
                status = ck.get("status", "warning")
                row3 = [
                    h.get("hostname", "-"), h.get("ip", "-"),
                    ck.get("item", "-"), ck.get("value", "-"),
                    STATUS_TEXT.get(status, status), ck.get("threshold", "-"),
                ]
                for ci, val in enumerate(row3, 1):
                    write_data_cell(ws3, ri3, ci, val, status if ci == 5 else None)
                ws3.row_dimensions[ri3].height = 16
                ri3 += 1

        ws3.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename     = f"巡检日报_{report_id}.xlsx"
        encoded_name = quote(filename, encoding="utf-8")
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
        )
    except Exception as e:
        logger.exception("Excel 导出失败")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/report/{report_id}")
async def get_report(report_id: str):
    """获取指定报告详情"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    data = _read_report_json(p)
    if not data:
        raise HTTPException(status_code=404, detail="报告内容不存在")
    return data


@router.get("/api/report/{report_id}/export.html")
async def export_report_html(report_id: str):
    """将报告渲染为可打印 HTML，浏览器 Ctrl+P → 另存为 PDF 即可。"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    data = _read_report_json(p)
    if not data:
        raise HTTPException(status_code=404, detail="报告内容不存在")
    html = _render_report_html(data)
    encoded = quote(f"{data.get('title', report_id)}.html")
    return Response(
        content=html,
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f'inline; filename*=UTF-8\'\'{encoded}'},
    )


@router.get("/api/report/{report_id}/export.pdf")
async def export_report_pdf(report_id: str):
    """报告直接导出 PDF（reportlab 生成，支持运维日报/主机巡检/慢日志三类）。"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    data = _read_report_json(p)
    if not data:
        raise HTTPException(status_code=404, detail="报告内容不存在")

    from services.report_pdf import build_report_pdf
    try:
        # reportlab 是纯 CPU 同步渲染，放线程池避免阻塞事件循环
        pdf_bytes = await asyncio.to_thread(build_report_pdf, data)
    except Exception as exc:
        logger.exception("[report] PDF 导出失败 %s", report_id)
        raise HTTPException(status_code=500, detail=f"PDF 生成失败: {exc}")

    encoded = quote(f"{data.get('title', report_id)}.pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{encoded}'},
    )


_SAFE_REPORT_ID = re.compile(r"^[A-Za-z0-9_-]{1,64}$")


def _public_inspect_report_path(report_id: str) -> Path:
    if not _SAFE_REPORT_ID.fullmatch(report_id or ""):
        raise HTTPException(status_code=404, detail="巡检报告不存在")
    reports_root = REPORTS_DIR.resolve()
    report_path = (reports_root / f"{report_id}.json").resolve()
    if report_path.parent != reports_root:
        raise HTTPException(status_code=404, detail="巡检报告不存在")
    return report_path


@router.get("/api/public/report/inspect/{report_id}.pdf")
async def public_inspect_report_pdf(report_id: str):
    report_path = _public_inspect_report_path(report_id)
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="巡检报告不存在")
    data = _read_report_json(report_path)
    if not data or data.get("type") != "inspect":
        raise HTTPException(status_code=404, detail="巡检报告不存在")

    from services.report_pdf import build_report_pdf
    try:
        pdf_bytes = await asyncio.to_thread(build_report_pdf, data)
    except Exception as exc:
        logger.exception("[report] 公开巡检 PDF 导出失败 %s", report_id)
        raise HTTPException(status_code=500, detail=f"PDF 生成失败: {exc}") from exc

    encoded = quote(f"{data.get('title', report_id)}.pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{encoded}"},
    )


def _render_report_html(data: dict) -> str:
    title       = data.get("title", "运维报告")
    created_at  = data.get("created_at", "")[:19].replace("T", " ")
    score       = data.get("health_score", "-")
    try:
        score_value = int(score)
    except (TypeError, ValueError):
        score_value = 0
    score_color = "#22c55e" if score_value >= 80 else ("#f59e0b" if score_value >= 60 else "#ef4444")
    ai_text     = data.get("ai_analysis", "（暂无 AI 分析）")
    report_type = data.get("type", "daily")
    report_type_label = {
        "daily": "运维日报",
        "inspect": "主机巡检日报",
        "slowlog": "MySQL 慢日志报告",
    }.get(report_type, "分析报告")

    # 转义 HTML 特殊字符
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    # AI 分析：Markdown 粗转 HTML（加粗、换行、标题）
    def md2html(text):
        import re
        lines, out = text.split("\n"), []
        for ln in lines:
            ln = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', ln)
            ln = re.sub(r'^#{1,3}\s+(.+)', r'<h4>\1</h4>', ln)
            ln = re.sub(r'^[-*]\s+', '• ', ln)
            out.append(f'<p>{ln}</p>' if not ln.startswith('<h4>') else ln)
        return "\n".join(out)

    # 节点状态 / 巡检状态
    node = data.get("node_status", {})
    summary = data.get("host_summary") or data.get("summary", {})
    stats_block = ""
    if node or summary:
        items = []
        if node:
            items += [f"节点正常: {node.get('normal',0)}", f"节点异常: {node.get('abnormal',0)}"]
        if summary:
            items += [f"主机总数: {summary.get('total',0)}",
                      f"正常: {summary.get('normal',0)}",
                      f"警告: {summary.get('warning',0)}",
                      f"严重: {summary.get('critical',0)}"]
        stats_block = "<div class='stats-row'>" + "".join(
            f"<div class='stat-card'>{esc(i)}</div>" for i in items
        ) + "</div>"

    detail_block = ""
    if report_type == "daily":
        service_rows = "".join(
            "<tr>"
            f"<td>{esc(item.get('service', '-'))}</td>"
            f"<td>{esc('、'.join(item.get('keywords') or []) or '样本不足')}</td>"
            f"<td>{esc(item.get('summary', '-'))}</td>"
            "</tr>"
            for item in data.get("service_error_summaries") or []
        )
        keyword_rows = "".join(
            "<tr>"
            f"<td>{esc(item.get('keyword', '-'))}</td>"
            f"<td>{esc(item.get('count', 0))}</td>"
            f"<td>{esc('、'.join(item.get('services') or []) or '-')}</td>"
            "</tr>"
            for item in data.get("error_keywords") or []
        )
        interface = data.get("interface_status") or {}
        interface_rows = "".join(
            "<tr>"
            f"<td>{esc(row.get('status', '-'))}</td>"
            f"<td>{esc(row.get('application', '-'))}</td>"
            f"<td>{esc(row.get('method', '-'))}</td>"
            f"<td>{esc(row.get('uri', '-'))}</td>"
            f"<td>{esc(row.get('error_ratio', 0))}%</td>"
            f"<td>{esc(row.get('p95_ms', 0))}ms</td>"
            "</tr>"
            for row in interface.get("rows") or []
        )
        detail_block = f"""
        <h2>第二部分 · 微服务错误一句话概括</h2>
        <table><thead><tr><th>微服务</th><th>错误关键字</th><th>一句话结论</th></tr></thead>
        <tbody>{service_rows or '<tr><td colspan="3">未发现微服务错误</td></tr>'}</tbody></table>
        <h2>第三部分 · 高频错误关键字</h2>
        <table><thead><tr><th>关键字</th><th>样本频次</th><th>涉及服务</th></tr></thead>
        <tbody>{keyword_rows or '<tr><td colspan="3">暂无可统计关键字</td></tr>'}</tbody></table>
        <h2>第四部分 · 接口监控状况评估</h2>
        <table><thead><tr><th>状态</th><th>应用</th><th>方法</th><th>接口</th><th>5xx率</th><th>P95</th></tr></thead>
        <tbody>{interface_rows or '<tr><td colspan="6">未采集到 http_server 接口指标</td></tr>'}</tbody></table>
        """
    elif report_type == "inspect":
        sections = data.get("group_sections") or [{
            "group_name": data.get("group_name") or "全部主机",
            "hosts": data.get("all_hosts") or [],
        }]
        section_html = []
        for group in sections:
            rows = []
            for host in group.get("hosts") or []:
                disks = "；".join(
                    f"{p.get('mountpoint') or p.get('mount') or '-'} "
                    f"{p.get('usage_pct', p.get('used_pct', '-'))}% "
                    f"({p.get('used_gb', '-')}/{p.get('total_gb', p.get('size_gb', '-'))}GB)"
                    for p in host.get("partitions") or []
                ) or "未采集"
                processes = "；".join(
                    f"{i}. {proc.get('service') or proc.get('comm') or '-'} "
                    f"CPU {proc.get('cpu', '-')}%/MEM {proc.get('mem', '-')}%"
                    for i, proc in enumerate(host.get("process_top10") or [], 1)
                ) or host.get("process_error") or "未采集"
                rows.append(
                    "<tr>"
                    f"<td>{esc(host.get('overall', '-'))}</td><td>{esc(host.get('ip', '-'))}</td>"
                    f"<td>{esc(host.get('role') or host.get('os') or '-')}</td>"
                    f"<td>{esc(host.get('hostname', '-'))}</td>"
                    f"<td>{esc(host.get('cpu_pct', '-'))}%</td><td>{esc(host.get('mem_pct', '-'))}%</td>"
                    f"<td>{esc(disks)}</td><td>{esc(processes)}</td>"
                    f"<td>{esc(host.get('owner') or '未配置')}</td>"
                    f"<td>{esc(host.get('datacenter') or '未配置')}</td>"
                    "</tr>"
                )
            section_html.append(
                f"<h3>{esc(group.get('group_name', '未分组'))}（{len(group.get('hosts') or [])} 台，告警优先）</h3>"
                "<div class='table-scroll'><table><thead><tr>"
                "<th>状态</th><th>服务器IP</th><th>服务器</th><th>主机名</th><th>CPU</th><th>内存</th>"
                "<th>磁盘具体占用</th><th>进程占用Top10</th><th>负责人</th><th>机房</th>"
                f"</tr></thead><tbody>{''.join(rows)}</tbody></table></div>"
            )
        detail_block = "<h2>第二部分 · 按组分类主机巡检明细</h2>" + "".join(section_html)
    elif report_type == "slowlog":
        slow_rows = "".join(
            "<tr>"
            f"<td>{esc(i)}</td><td>{esc(row.get('host_ip', '-'))}</td>"
            f"<td>{esc(row.get('query_time', 0))}s</td>"
            f"<td>{esc(row.get('rows_examined', 0))}</td><td>{esc(row.get('sql_brief', '-'))}</td>"
            "</tr>"
            for i, row in enumerate(data.get("top_slow") or [], 1)
        )
        detail_block = f"""
        <h2>慢日志分析结果</h2>
        <div class="stats-row"><div class="stat-card">分析时段: {esc(data.get('date_from', '-'))} ~ {esc(data.get('date_to', '-'))}</div>
        <div class="stat-card">慢查询: {esc(data.get('total_queries', 0))}</div>
        <div class="stat-card">高耗时告警: {esc(data.get('alert_queries', 0))}</div>
        <div class="stat-card">最大耗时: {esc(data.get('max_query_time', 0))}s</div></div>
        <h3>Top 10 最慢查询</h3>
        <table><thead><tr><th>#</th><th>主机IP</th><th>耗时</th><th>扫描行</th><th>SQL摘要</th></tr></thead>
        <tbody>{slow_rows or '<tr><td colspan="5">未发现慢查询</td></tr>'}</tbody></table>
        """

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>{esc(title)}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "PingFang SC","Microsoft YaHei",Arial,sans-serif; font-size: 13px;
          color: #1a1a2e; background: #fff; padding: 32px 40px; max-width: 900px; margin: auto; }}
  h1 {{ font-size: 22px; color: #1e3a8a; border-bottom: 2px solid #3b82f6; padding-bottom: 10px; margin-bottom: 6px; }}
  h2 {{ font-size: 16px; color: #1e40af; margin: 24px 0 10px; border-left: 4px solid #3b82f6; padding-left: 10px; }}
  h3 {{ font-size: 14px; color: #374151; margin: 18px 0 8px; }}
  h4 {{ font-size: 13px; font-weight: 700; margin: 10px 0 4px; color: #1f2937; }}
  .meta {{ color: #6b7280; font-size: 12px; margin-bottom: 18px; }}
  .score-badge {{ display: inline-block; padding: 4px 14px; border-radius: 999px;
                  background: {score_color}22; color: {score_color};
                  font-size: 14px; font-weight: 700; border: 1px solid {score_color}55; }}
  .kv-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 14px 0; }}
  .kv-card {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px;
              padding: 12px 14px; text-align: center; }}
  .kv-card .val {{ font-size: 20px; font-weight: 700; color: #1e3a8a; }}
  .kv-card .lbl {{ font-size: 11px; color: #6b7280; margin-top: 2px; }}
  .stats-row {{ display: flex; gap: 10px; flex-wrap: wrap; margin: 12px 0; }}
  .stat-card {{ background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px;
                padding: 8px 16px; font-size: 12px; color: #1d4ed8; }}
  .ai-section {{ background: #fafafa; border: 1px solid #e5e7eb; border-radius: 10px;
                 padding: 20px; margin: 10px 0; line-height: 1.8; }}
  .ai-section p {{ margin: 4px 0; color: #374151; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 12px; }}
  th {{ background: #eff6ff; color: #1d4ed8; padding: 8px 12px; text-align: left;
        border-bottom: 2px solid #bfdbfe; }}
  td {{ padding: 7px 12px; border-bottom: 1px solid #f3f4f6; }}
  tr:nth-child(even) td {{ background: #f9fafb; }}
  .table-scroll {{ width: 100%; overflow-x: auto; }}
  .table-scroll table {{ min-width: 1300px; }}
  .footer {{ margin-top: 32px; padding-top: 12px; border-top: 1px solid #e5e7eb;
             color: #9ca3af; font-size: 11px; text-align: center; }}
  @media print {{
    body {{ padding: 20px; }}
    .no-print {{ display: none !important; }}
    a {{ text-decoration: none; color: inherit; }}
    @page {{ margin: 15mm 15mm; size: A4; }}
  }}
</style>
</head>
<body>
<div class="no-print" style="background:#eff6ff;border:1px solid #93c5fd;border-radius:8px;
     padding:10px 16px;margin-bottom:20px;font-size:12px;color:#1d4ed8;display:flex;align-items:center;gap:10px;">
  <span>💡 在浏览器按 <strong>Ctrl+P</strong>（Mac: ⌘+P），选择「另存为 PDF」即可导出 PDF 文件。</span>
  <button onclick="window.print()" style="margin-left:auto;padding:5px 14px;background:#3b82f6;color:#fff;
    border:none;border-radius:6px;cursor:pointer;font-size:12px;">打印 / 存为 PDF</button>
</div>

<h1>{esc(title)}</h1>
<div class="meta">生成时间：{esc(created_at)} &nbsp;·&nbsp; 报告类型：{report_type_label}</div>

<div class="kv-grid">
  <div class="kv-card"><div class="val"><span class="score-badge">{esc(score)}</span></div><div class="lbl">健康评分</div></div>
  <div class="kv-card"><div class="val">{esc(data.get("service_count", data.get("hosts_count", summary.get("total", 0))))}</div><div class="lbl">覆盖对象</div></div>
  <div class="kv-card"><div class="val">{esc(data.get("active_alerts", data.get("alert_queries", summary.get("warning", 0))))}</div><div class="lbl">需关注</div></div>
  <div class="kv-card"><div class="val">{esc((data.get("interface_status") or {}).get("status", data.get("max_query_time", summary.get("critical", "-"))))}</div><div class="lbl">关键状态</div></div>
</div>

{stats_block}

<h2>第一部分 · AI 结论</h2>
<div class="ai-section">{md2html(esc(ai_text))}</div>

{detail_block}

<div class="footer">由 AIOps 智能运维平台生成 · {esc(created_at)}</div>
</body>
</html>"""


# ── 通知推送 ──────────────────────────────────────────────────────────────────

class NotifyRequest(BaseModel):
    channels: list[str]  # ["feishu", "dingtalk"]


@router.post("/api/report/{report_id}/notify")
async def notify_report(report_id: str, body: NotifyRequest):
    """将指定报告推送到飞书 / 钉钉"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    report = _read_report_json(p)
    if not report:
        raise HTTPException(status_code=404, detail="报告内容不存在")
    report_url = f"{APP_URL}/report/{report_id}" if APP_URL else ""

    results = {}
    for ch in body.channels:
        if ch == "feishu":
            if not FEISHU_WEBHOOK:
                results["feishu"] = {"ok": False, "msg": "未配置 FEISHU_WEBHOOK"}
            else:
                results["feishu"] = await send_feishu(
                    report, FEISHU_WEBHOOK, keyword=FEISHU_KEYWORD, report_url=report_url
                )
        elif ch == "dingtalk":
            if not DINGTALK_WEBHOOK:
                results["dingtalk"] = {"ok": False, "msg": "未配置 DINGTALK_WEBHOOK"}
            else:
                results["dingtalk"] = await send_dingtalk(
                    report, DINGTALK_WEBHOOK, keyword=DINGTALK_KEYWORD
                )
        else:
            results[ch] = {"ok": False, "msg": f"不支持的渠道: {ch}"}

    return {"results": results}


@router.post("/api/report/{report_id}/notify-groups")
async def notify_report_groups(report_id: str, group_id: Optional[str] = Query(None)):
    """将指定报告推送到分组。
    若传入 group_id 则只推送该分组；否则推送所有已配置 webhook 的分组。"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    report = _read_report_json(p)
    if not report:
        raise HTTPException(status_code=404, detail="报告内容不存在")
    report_url = f"{APP_URL}/report/{report_id}" if APP_URL else ""

    all_groups = load_groups()
    group_reports = _split_inspect_report_by_group(report) if report.get("type") == "inspect" else {}
    # 若指定分组则只取该分组，否则取全部
    target_groups = [g for g in all_groups if g["id"] == group_id] if group_id else all_groups

    results = []
    for group in target_groups:
        gid        = group["id"]
        group_name = group.get("name", gid)
        pushed: dict[str, dict] = {}
        payload_report = group_reports.get(gid, report) if group_reports else report

        if report.get("type") == "inspect" and group_reports and gid not in group_reports:
            results.append({"group_id": gid, "group_name": group_name, "skipped": True, "reason": "该分组在当前报告中无主机"})
            continue

        feishu_wh = group.get("feishu_webhook", "")
        if feishu_wh:
            keyword = group.get("feishu_keyword", FEISHU_KEYWORD)
            pushed["feishu"] = await send_feishu(payload_report, feishu_wh, keyword=keyword, report_url=report_url)

        dingtalk_wh = group.get("dingtalk_webhook", "")
        if dingtalk_wh:
            keyword = group.get("dingtalk_keyword", DINGTALK_KEYWORD)
            pushed["dingtalk"] = await send_dingtalk(payload_report, dingtalk_wh, keyword=keyword)

        if not pushed:
            results.append({"group_id": gid, "group_name": group_name, "skipped": True, "reason": "未配置 webhook"})
        else:
            results.append({"group_id": gid, "group_name": group_name, "push": pushed})

    return {"results": results}


# ── 慢日志分析报告 ─────────────────────────────────────────────────────────────

class SlowlogTargetsConfig(BaseModel):
    enabled:       bool  = False
    date_days:     int   = 1
    threshold_sec: float = 1.0
    alert_sec:     float = 10.0
    targets:       list  = []


@router.get("/api/report/slowlog/targets")
async def get_slowlog_targets():
    """获取慢日志定时报告目标配置"""
    return load_slowlog_targets()


@router.put("/api/report/slowlog/targets")
async def put_slowlog_targets(body: SlowlogTargetsConfig):
    """保存慢日志定时报告目标配置"""
    save_slowlog_targets(body.model_dump())
    return {"ok": True}


@router.get("/api/report/slowlog/generate")
async def generate_slowlog_report():
    """根据已保存的目标配置生成慢日志分析报告，SSE 流式返回"""
    from report_builder import collect_slowlog_data, build_slowlog_ai_prompt

    config = load_slowlog_targets()
    if not config.get("targets"):
        raise HTTPException(400, detail="未配置慢日志分析目标，请先在「慢日志报告 · 配置目标」中添加主机")

    async def generate():
        data = await collect_slowlog_data(config)
        if data is None:
            yield f"data: {json.dumps('未找到满足条件的慢查询记录', ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"
            return

        meta = {k: v for k, v in data.items() if not k.startswith("_")}
        report_path = REPORTS_DIR / f"{meta['id']}.json"
        initial_report = {
            **meta,
            "ai_analysis": "AI 分析生成中，基础慢日志统计已保存。",
        }
        _write_report_json(report_path, initial_report)
        await save_report_meta(initial_report)
        yield f"data: __META__{json.dumps(meta, ensure_ascii=False)}\n\n"

        if not data.get("_all_entries"):
            meta["ai_analysis"] = "未找到满足条件的慢查询记录，请检查目标主机、日志路径及时间范围配置。"
            _write_report_json(report_path, meta)
            await save_report_meta(meta)
            yield "data: [DONE]\n\n"
            return

        prompt = build_slowlog_ai_prompt(data)
        ai_parts: list[str] = []
        try:
            timeout_seconds = _slowlog_ai_timeout_seconds()
            async for chunk in _stream_ai_with_timeout(
                prompt,
                max_tokens=1800,
                timeout_seconds=timeout_seconds,
            ):
                ai_parts.append(chunk)
                yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
        except asyncio.TimeoutError:
            msg = (
                f"\n[AI生成超时] 已保存基础慢日志报告；AI 分析超过 "
                f"{_slowlog_ai_timeout_seconds()} 秒未完成，请稍后重试或检查模型服务。"
            )
            logger.warning("[slowlog_report] AI 生成超时")
            ai_parts.append(msg)
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
        except Exception as exc:
            logger.exception("[slowlog_report] AI 生成异常")
            ai_parts.append(f"\n[AI生成出错] {exc}")
            yield f"data: {json.dumps('[AI生成出错] ' + str(exc), ensure_ascii=False)}\n\n"

        meta["ai_analysis"] = "".join(ai_parts)
        _write_report_json(report_path, meta)
        await save_report_meta(meta)
        try:
            from agent.report_memory import get_report_memory
            asyncio.create_task(get_report_memory().save(meta))
        except Exception:
            pass
        yield "data: [DONE]\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )
