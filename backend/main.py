"""AI Ops 日志分析系统 - FastAPI 后端"""
import json
import logging
import os
import time
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from urllib.parse import quote

import asyncio

import httpx
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from dotenv import load_dotenv
from fastapi import FastAPI, Query, HTTPException, WebSocket
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel

from cryptography.fernet import Fernet

from loki_client import LokiClient
from ai_analyzer import AIAnalyzer
from log_clusterer import LogClusterer
from notifier import send_feishu, send_dingtalk
from prom_client import PrometheusClient
from ssh_bridge import ssh_websocket_handler

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

LOKI_URL = os.getenv("LOKI_URL", "http://localhost:3100")
LOKI_USERNAME = os.getenv("LOKI_USERNAME", "")
LOKI_PASSWORD = os.getenv("LOKI_PASSWORD", "")
PROMETHEUS_URL = os.getenv("PROMETHEUS_URL", "http://localhost:9090")
PROMETHEUS_USERNAME = os.getenv("PROMETHEUS_USERNAME", "")
PROMETHEUS_PASSWORD = os.getenv("PROMETHEUS_PASSWORD", "")
REPORTS_DIR = Path(os.getenv("REPORTS_DIR", "./reports"))
REPORTS_DIR.mkdir(exist_ok=True)
CMDB_FILE = Path(os.getenv("CMDB_FILE", "./cmdb_hosts.json"))
CREDENTIALS_FILE = Path(os.getenv("CREDENTIALS_FILE", "./ssh_credentials.json"))
# SSH 密码加密密钥（自动生成并持久化到 .ssh_key 文件）
_SSH_KEY_FILE = Path(os.getenv("SSH_KEY_FILE", "./.ssh_key"))
if _SSH_KEY_FILE.exists():
    _FERNET_KEY = _SSH_KEY_FILE.read_bytes().strip()
else:
    _FERNET_KEY = Fernet.generate_key()
    _SSH_KEY_FILE.write_bytes(_FERNET_KEY)
    logger.info("[启动] 已生成 SSH 密码加密密钥: %s", _SSH_KEY_FILE)
_fernet = Fernet(_FERNET_KEY)

FEISHU_WEBHOOK    = os.getenv("FEISHU_WEBHOOK", "")
FEISHU_KEYWORD    = os.getenv("FEISHU_KEYWORD", "")
DINGTALK_WEBHOOK  = os.getenv("DINGTALK_WEBHOOK", "")
DINGTALK_KEYWORD  = os.getenv("DINGTALK_KEYWORD", "")
APP_URL           = os.getenv("APP_URL", "").rstrip("/")

# ── 定时推送配置 ──────────────────────────────
# SCHEDULE_CRON：cron 表达式，默认每天 09:00（服务器本地时间）
# 格式：分 时 日 月 周，例如 "0 9 * * *"
SCHEDULE_CRON     = os.getenv("SCHEDULE_CRON", "0 9 * * *")
# SCHEDULE_CHANNELS：推送渠道，逗号分隔，例如 "feishu,dingtalk"
SCHEDULE_CHANNELS = [
    ch.strip() for ch in os.getenv("SCHEDULE_CHANNELS", "").split(",") if ch.strip()
]

loki = LokiClient(LOKI_URL, LOKI_USERNAME, LOKI_PASSWORD)
logger.info("[启动] PROMETHEUS_URL=%s, auth=%s", PROMETHEUS_URL, "yes" if PROMETHEUS_USERNAME else "no")
prom = PrometheusClient(PROMETHEUS_URL, PROMETHEUS_USERNAME, PROMETHEUS_PASSWORD)
analyzer = AIAnalyzer()
clusterer = LogClusterer()


def _load_cmdb() -> dict:
    """加载 CMDB 数据（手动补充的字段：owner/env/role/notes）"""
    if CMDB_FILE.exists():
        return json.loads(CMDB_FILE.read_text(encoding="utf-8"))
    return {}


def _encrypt_password(plain: str) -> str:
    """加密 SSH 密码"""
    return _fernet.encrypt(plain.encode()).decode()


def _decrypt_password(cipher: str) -> str:
    """解密 SSH 密码"""
    return _fernet.decrypt(cipher.encode()).decode()


def _save_cmdb(data: dict):
    CMDB_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_credentials() -> list[dict]:
    """加载凭证库"""
    if CREDENTIALS_FILE.exists():
        return json.loads(CREDENTIALS_FILE.read_text(encoding="utf-8"))
    return []


def _save_credentials(data: list[dict]):
    CREDENTIALS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ────────── 定时任务 ──────────

async def _build_and_save_report() -> dict:
    """生成并保存运维日报（非流式），返回完整 report dict。"""
    error_counts = await loki.count_errors_by_service(hours=24)
    error_logs   = await loki.query_error_logs(hours=24, limit=1000)
    services     = await loki.get_services()

    node_status       = {"normal": 27, "abnormal": 0}
    active_alerts     = min(len(error_counts), 3)
    total_error_count = sum(error_counts.values())
    total_logs        = max(total_error_count * 8, total_error_count)

    health_score = await analyzer.calculate_health_score(
        total_logs, total_error_count, active_alerts, node_status["abnormal"]
    )

    now       = datetime.now(timezone.utc)
    report_id = now.strftime("%Y%m%d%H%M%S")

    report = {
        "id":           report_id,
        "title":        f"运维日报 {now.strftime('%Y-%m-%d')}",
        "created_at":   now.isoformat(),
        "health_score": health_score,
        "total_logs":   total_logs,
        "total_errors": total_error_count,
        "service_count": len(services),
        "active_alerts": active_alerts,
        "node_status":  node_status,
        "top10_errors": [
            {"service": k, "count": v}
            for k, v in list(error_counts.items())[:10]
        ],
    }

    ai_parts = []
    async for chunk in analyzer.generate_daily_report(
        error_counts=error_counts,
        total_logs=total_logs,
        service_count=len(services),
        node_status=node_status,
        active_alerts=active_alerts,
        sample_errors=error_logs,
    ):
        ai_parts.append(chunk)
    report["ai_analysis"] = "".join(ai_parts)

    report_path = REPORTS_DIR / f"{report_id}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


async def _scheduled_report_job():
    """定时任务入口：生成日报并推送到已配置的渠道。"""
    if not SCHEDULE_CHANNELS:
        logger.info("[scheduler] SCHEDULE_CHANNELS 未配置，跳过推送")
        return
    logger.info("[scheduler] 开始生成定时日报 ...")
    try:
        report = await _build_and_save_report()
        report_url = f"{APP_URL}/report/{report['id']}" if APP_URL else ""
        for ch in SCHEDULE_CHANNELS:
            if ch == "feishu":
                if not FEISHU_WEBHOOK:
                    logger.warning("[scheduler] FEISHU_WEBHOOK 未配置，跳过飞书推送")
                    continue
                result = await send_feishu(report, FEISHU_WEBHOOK, keyword=FEISHU_KEYWORD, report_url=report_url)
                logger.info("[scheduler] 飞书推送结果: %s", result)
            elif ch == "dingtalk":
                if not DINGTALK_WEBHOOK:
                    logger.warning("[scheduler] DINGTALK_WEBHOOK 未配置，跳过钉钉推送")
                    continue
                result = await send_dingtalk(report, DINGTALK_WEBHOOK, keyword=DINGTALK_KEYWORD)
                logger.info("[scheduler] 钉钉推送结果: %s", result)
            else:
                logger.warning("[scheduler] 不支持的推送渠道: %s", ch)
        logger.info("[scheduler] 定时日报完成，report_id=%s", report["id"])
    except Exception:
        logger.exception("[scheduler] 定时日报任务异常")


@asynccontextmanager
async def lifespan(app: FastAPI):
    scheduler = AsyncIOScheduler()
    try:
        minute, hour, day, month, day_of_week = SCHEDULE_CRON.split()
    except ValueError:
        minute, hour, day, month, day_of_week = "0", "9", "*", "*", "*"
        logger.warning("[scheduler] SCHEDULE_CRON 格式错误，使用默认值 '0 9 * * *'")

    scheduler.add_job(
        _scheduled_report_job,
        CronTrigger(minute=minute, hour=hour, day=day, month=month, day_of_week=day_of_week),
        id="daily_report",
        replace_existing=True,
    )
    scheduler.start()
    logger.info(
        "[scheduler] 定时推送已启动，cron='%s'，渠道=%s",
        SCHEDULE_CRON, SCHEDULE_CHANNELS or "（未配置，仅生成报告不推送）",
    )
    yield
    scheduler.shutdown(wait=False)
    logger.info("[scheduler] 定时推送已停止")


app = FastAPI(title="AI Ops Log Analysis", version="1.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ────────── 服务列表 ──────────

@app.get("/api/services")
async def get_services():
    """获取所有服务及其错误数"""
    try:
        services = await loki.get_services()
        return {"data": services, "total": len(services)}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Loki 连接失败: {e}")


# ────────── 日志查询 ──────────

def _parse_time_ns(dt_str: Optional[str]) -> Optional[int]:
    """将 ISO datetime 字符串解析为纳秒时间戳，解析失败返回 None。"""
    if not dt_str:
        return None
    try:
        # 兼容带 / 不带时区的格式
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"):
            try:
                dt = datetime.strptime(dt_str, fmt).replace(tzinfo=timezone.utc)
                return int(dt.timestamp() * 1e9)
            except ValueError:
                continue
        # 尝试 fromisoformat（Python 3.11+ 支持更多格式）
        dt = datetime.fromisoformat(dt_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return int(dt.timestamp() * 1e9)
    except Exception:
        return None


@app.get("/api/logs")
async def get_logs(
    service: Optional[str] = Query(None, description="服务名称"),
    hours: int = Query(24, description="查询时长（小时）"),
    limit: int = Query(2000, le=10000, description="返回条数"),
    level: Optional[str] = Query(None, description="日志级别过滤: error/warn/info"),
    keyword: Optional[str] = Query(None, description="关键字过滤（不区分大小写）"),
    start_time: Optional[str] = Query(None, description="自定义开始时间 ISO 格式，如 2024-01-01T00:00"),
    end_time: Optional[str] = Query(None, description="自定义结束时间 ISO 格式，如 2024-01-01T23:59"),
):
    """查询日志"""
    try:
        logs = await loki.query_logs(
            service=service,
            hours=hours,
            limit=limit,
            level=level or None,
            keyword=keyword or None,
            start_ns=_parse_time_ns(start_time),
            end_ns=_parse_time_ns(end_time),
        )
        return {
            "data": logs,
            "total": len(logs),
            "service": service,
            "hours": hours,
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))


@app.get("/api/logs/errors")
async def get_error_logs(
    hours: int = Query(24),
    limit: int = Query(5000, le=20000),
):
    """查询全量错误日志"""
    try:
        logs = await loki.query_error_logs(hours=hours, limit=limit)
        return {"data": logs, "total": len(logs)}
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))


@app.get("/api/metrics/errors")
async def get_error_metrics(hours: int = Query(24)):
    """各服务错误数统计"""
    try:
        counts = await loki.count_errors_by_service(hours=hours)
        return {
            "data": [{"service": k, "count": v} for k, v in counts.items()],
            "total_errors": sum(counts.values()),
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))


# ────────── 日志模板聚合 ──────────

@app.get("/api/logs/templates")
async def get_log_templates(
    service: Optional[str] = Query(None, description="服务名称"),
    hours: int = Query(24, description="查询时长（小时）"),
    limit: int = Query(10000, le=50000, description="参与聚类的日志上限"),
    top_n: int = Query(100, le=500, description="返回模板数上限"),
    level: Optional[str] = Query(None, description="日志级别过滤: error/warn，不传则聚类全量日志"),
    keyword: Optional[str] = Query(None, description="关键字过滤"),
    start_time: Optional[str] = Query(None, description="自定义开始时间 ISO 格式"),
    end_time: Optional[str] = Query(None, description="自定义结束时间 ISO 格式"),
):
    """Drain3 日志模板聚类：将重复日志归纳为带 <*> 占位符的模板"""
    try:
        logs = await loki.query_logs(
            service=service, hours=hours, limit=limit, level=level,
            keyword=keyword or None,
            start_ns=_parse_time_ns(start_time),
            end_ns=_parse_time_ns(end_time),
        )
        templates = clusterer.cluster(logs, top_n=top_n)
        return {
            "data": templates,
            "total_logs": len(logs),
            "total_templates": len(templates),
            "service": service,
            "hours": hours,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ────────── AI 分析 ──────────

@app.get("/api/analyze/stream")
async def analyze_logs_stream(
    service: Optional[str] = Query(None),
    hours: int = Query(24),
):
    """流式 AI 分析日志（SSE）"""
    try:
        logs = await loki.query_error_logs(service=service, hours=hours)
        if not logs:
            async def empty():
                yield "data: 该时间范围内未发现错误日志。\n\n"
                yield "data: [DONE]\n\n"
            return StreamingResponse(empty(), media_type="text/event-stream")

        async def generate():
            try:
                async for chunk in analyzer.analyze_logs_stream(logs, service or ""):
                    yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
                yield "data: [DONE]\n\n"
            except Exception as exc:
                logger.exception("AI 分析流式输出异常")
                yield f"data: {json.dumps('[AI分析出错] ' + str(exc), ensure_ascii=False)}\n\n"
                yield "data: [DONE]\n\n"

        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ────────── 日报 ──────────

class ReportMeta(BaseModel):
    id: str
    title: str
    created_at: str
    health_score: int
    total_logs: int
    total_errors: int
    service_count: int
    active_alerts: int


@app.get("/api/report/generate")
async def generate_report():
    """触发生成运维日报，流式返回 AI 分析内容"""
    try:
        # 并发获取数据
        error_counts = await loki.count_errors_by_service(hours=24)
        error_logs = await loki.query_error_logs(hours=24, limit=1000)

        # 节点状态（模拟，实际可接 Prometheus）
        node_status = {"normal": 27, "abnormal": 0}
        active_alerts = min(len(error_counts), 3)

        services = await loki.get_services()
        total_error_count = sum(error_counts.values())

        # 估算总日志量（错误日志通常占总量 5-15%）
        total_logs = max(total_error_count * 8, total_error_count)

        health_score = await analyzer.calculate_health_score(
            total_logs, total_error_count, active_alerts, node_status["abnormal"]
        )

        now = datetime.now(timezone.utc)
        report_id = now.strftime("%Y%m%d%H%M%S")

        # 保存报告元数据（不含 AI 分析，AI 分析实时流式返回）
        meta = {
            "id": report_id,
            "title": f"运维日报 {now.strftime('%Y-%m-%d')}",
            "created_at": now.isoformat(),
            "health_score": health_score,
            "total_logs": total_logs,
            "total_errors": total_error_count,
            "service_count": len(services),
            "active_alerts": active_alerts,
            "node_status": node_status,
            "top10_errors": [
                {"service": k, "count": v}
                for k, v in list(error_counts.items())[:10]
            ],
        }

        report_path = REPORTS_DIR / f"{report_id}.json"
        ai_content_parts = []

        async def generate():
            # 先发送报告元数据
            yield f"data: __META__{json.dumps(meta, ensure_ascii=False)}\n\n"

            try:
                # 流式 AI 分析
                async for chunk in analyzer.generate_daily_report(
                    error_counts=error_counts,
                    total_logs=total_logs,
                    service_count=len(services),
                    node_status=node_status,
                    active_alerts=active_alerts,
                    sample_errors=error_logs,
                ):
                    ai_content_parts.append(chunk)
                    yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
            except Exception as exc:
                logger.exception("日报 AI 生成异常")
                ai_content_parts.append(f"\n[AI生成出错] {exc}")
                yield f"data: {json.dumps('[AI生成出错] ' + str(exc), ensure_ascii=False)}\n\n"

            # 保存完整报告
            meta["ai_analysis"] = "".join(ai_content_parts)
            report_path.write_text(
                json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            yield "data: [DONE]\n\n"

        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/list")
async def list_reports():
    """历史报告列表"""
    reports = []
    for p in sorted(REPORTS_DIR.glob("*.json"), reverse=True)[:30]:
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            reports.append({k: data[k] for k in [
                "id", "title", "created_at", "health_score",
                "total_logs", "total_errors", "service_count", "active_alerts"
            ] if k in data})
        except Exception:
            pass
    return {"data": reports}


@app.get("/api/report/inspect/generate")
async def generate_inspect_report():
    """生成主机巡检日报，流式返回 AI 分析"""
    try:
        results = await prom.inspect_hosts()
        summary = {
            "total":    len(results),
            "normal":   sum(1 for r in results if r["overall"] == "normal"),
            "warning":  sum(1 for r in results if r["overall"] == "warning"),
            "critical": sum(1 for r in results if r["overall"] == "critical"),
        }

        # 高频异常项 top10
        issue_cnt: dict[str, int] = {}
        for r in results:
            for c in r.get("checks", []):
                if c.get("status") != "normal":
                    issue_cnt[c.get("item", "未知")] = issue_cnt.get(c.get("item", "未知"), 0) + 1
        top_issues = sorted(issue_cnt.items(), key=lambda x: x[1], reverse=True)[:10]

        # 异常主机列表
        abnormal_hosts = [r for r in results if r.get("overall") != "normal"]
        abnormal_hosts.sort(key=lambda r: {"critical": 2, "warning": 1}.get(r.get("overall", "normal"), 0), reverse=True)

        health_score = await analyzer.calculate_host_health_score(summary)

        now = datetime.now(timezone.utc)
        report_id = "inspect_" + now.strftime("%Y%m%d%H%M%S")

        # 全量主机精简数据（用于 Excel 导出）
        all_hosts_brief = []
        for r in results:
            m = r.get("metrics") or {}
            all_hosts_brief.append({
                "hostname":   r.get("hostname") or r.get("instance", ""),
                "ip":         r.get("ip", ""),
                "os":         r.get("os", ""),
                "state":      r.get("state", ""),
                "overall":    r.get("overall", "normal"),
                "cpu_pct":    m.get("cpu_usage"),
                "mem_pct":    m.get("mem_usage"),
                "mem_total":  m.get("mem_total_gb"),
                "load5":      m.get("load5"),
                "net_recv":   m.get("net_recv_mbps"),
                "net_send":   m.get("net_send_mbps"),
                "tcp_estab":  m.get("tcp_estab"),
                "uptime_s":   m.get("uptime_seconds"),
                "checks":     r.get("checks", []),
                "partitions": r.get("partitions", []),
            })

        meta = {
            "id":             report_id,
            "type":           "inspect",
            "title":          f"主机巡检日报 {now.strftime('%Y-%m-%d')}",
            "created_at":     now.isoformat(),
            "health_score":   health_score,
            "host_summary":   summary,
            "top_issues":     [{"item": k, "count": v} for k, v in top_issues],
            "abnormal_hosts": abnormal_hosts[:20],
            "all_hosts":      all_hosts_brief,
        }

        report_path = REPORTS_DIR / f"{report_id}.json"
        ai_parts = []

        async def generate():
            yield f"data: __META__{json.dumps(meta, ensure_ascii=False)}\n\n"
            try:
                async for chunk in analyzer.generate_host_inspect_report(results, summary):
                    ai_parts.append(chunk)
                    yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
            except Exception as exc:
                logger.exception("巡检日报 AI 生成异常")
                ai_parts.append(f"\n[AI生成出错] {exc}")
                yield f"data: {json.dumps('[AI生成出错] ' + str(exc), ensure_ascii=False)}\n\n"
            meta["ai_analysis"] = "".join(ai_parts)
            report_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
            yield "data: [DONE]\n\n"

        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/inspect/{report_id}/excel")
async def download_inspect_report_excel(report_id: str):
    """下载主机巡检日报 Excel 文件"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    data = json.loads(p.read_text(encoding="utf-8"))
    if data.get("type") != "inspect":
        raise HTTPException(status_code=400, detail="该报告不是巡检类型")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()

        STATUS_COLOR = {"normal": "FF67C23A", "warning": "FFE6A23C", "critical": "FFF56C6C"}
        STATUS_TEXT  = {"normal": "正常", "warning": "警告", "critical": "严重"}

        def header_fill(color="FF1F3A5F"):
            return PatternFill("solid", fgColor=color)

        def thin_border():
            s = Side(style="thin", color="FFCCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def write_header_row(ws, headers, row=1):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFFFF", size=10)
                c.fill = header_fill()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = thin_border()

        def write_data_cell(ws, row, col, value, status=None):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
            if status and status in STATUS_COLOR:
                c.fill = PatternFill("solid", fgColor=STATUS_COLOR[status])
                c.font = Font(color="FFFFFFFF", size=9)
            else:
                c.font = Font(size=9)
            return c

        summary   = data.get("host_summary", {})
        all_hosts = data.get("all_hosts", [])
        top_issues = data.get("top_issues", [])
        ai_text   = data.get("ai_analysis", "")
        title     = data.get("title", "主机巡检日报")
        created   = data.get("created_at", "")[:19].replace("T", " ")

        # ── Sheet 1: 巡检概况 ───────────────────────────────────
        ws1 = wb.active
        ws1.title = "巡检概况"
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 40

        def s1_row(r, k, v, bold=False):
            ck = ws1.cell(row=r, column=1, value=k)
            cv = ws1.cell(row=r, column=2, value=v)
            ck.font = Font(bold=True, size=10)
            cv.font = Font(bold=bold, size=10)
            ck.fill = header_fill("FF243850")
            ck.font = Font(bold=True, color="FFFFFFFF", size=10)
            ck.alignment = cv.alignment = Alignment(vertical="center", wrap_text=True)
            ck.border = cv.border = thin_border()
            ws1.row_dimensions[r].height = 18

        s1_row(1, "报告名称", title, bold=True)
        s1_row(2, "生成时间", created)
        s1_row(3, "健康评分", f"{data.get('health_score', '-')}/100")
        s1_row(4, "巡检主机总数", summary.get("total", 0))
        s1_row(5, "正常主机", summary.get("normal", 0))
        s1_row(6, "警告主机", summary.get("warning", 0))
        s1_row(7, "严重主机", summary.get("critical", 0))

        # 高频异常项
        ws1.cell(row=9, column=1, value="高频异常项").font = Font(bold=True, size=10)
        write_header_row(ws1, ["异常检查项", "影响主机数"], row=10)
        for i, issue in enumerate(top_issues, 11):
            ws1.cell(row=i, column=1, value=issue.get("item", "")).border = thin_border()
            ws1.cell(row=i, column=2, value=issue.get("count", 0)).border = thin_border()
            ws1.cell(row=i, column=1).font = Font(size=9)
            ws1.cell(row=i, column=2).font = Font(size=9)
            ws1.cell(row=i, column=1).alignment = Alignment(vertical="center")
            ws1.cell(row=i, column=2).alignment = Alignment(horizontal="center", vertical="center")

        # AI 分析
        ai_start = 11 + len(top_issues) + 2
        ws1.cell(row=ai_start, column=1, value="AI 分析总结").font = Font(bold=True, size=10)
        if ai_text:
            c = ws1.cell(row=ai_start + 1, column=1, value=ai_text)
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = thin_border()
            ws1.merge_cells(
                start_row=ai_start + 1, start_column=1,
                end_row=ai_start + 1, end_column=2
            )
            ws1.row_dimensions[ai_start + 1].height = max(60, len(ai_text) // 3)

        # ── Sheet 2: 全部主机明细 ───────────────────────────────
        ws2 = wb.create_sheet("全部主机明细")
        headers2 = [
            "主机名", "IP地址", "操作系统", "状态", "巡检结果",
            "CPU使用率(%)", "内存使用率(%)", "内存总量(GB)",
            "负载(5m)", "网络收(Mbps)", "网络发(Mbps)",
            "TCP连接数", "运行时长(天)", "磁盘分区详情"
        ]
        write_header_row(ws2, headers2)
        set_col_widths(ws2, [18, 16, 22, 10, 10, 14, 14, 14, 10, 14, 14, 12, 12, 40])
        ws2.row_dimensions[1].height = 22

        for ri, h in enumerate(all_hosts, 2):
            overall = h.get("overall", "normal")
            status_label = STATUS_TEXT.get(overall, overall)

            def fmt(v, decimals=1):
                return round(v, decimals) if v is not None else "-"

            uptime_days = "-"
            us = h.get("uptime_s")
            if us is not None:
                uptime_days = round(us / 86400, 1)

            partitions = h.get("partitions") or []
            disk_text = "  ".join(
                f"{pt.get('mountpoint','?')} {pt.get('usage_pct','-')}%"
                for pt in partitions
            ) if partitions else "-"

            row_vals = [
                h.get("hostname", "-"),
                h.get("ip", "-"),
                h.get("os", "-"),
                h.get("state", "-"),
                status_label,
                fmt(h.get("cpu_pct")),
                fmt(h.get("mem_pct")),
                fmt(h.get("mem_total")),
                fmt(h.get("load5")),
                fmt(h.get("net_recv")),
                fmt(h.get("net_send")),
                h.get("tcp_estab") or "-",
                uptime_days,
                disk_text,
            ]
            for ci, val in enumerate(row_vals, 1):
                st = overall if ci == 5 else None
                write_data_cell(ws2, ri, ci, val, st)
            ws2.row_dimensions[ri].height = 16

        ws2.freeze_panes = "A2"

        # ── Sheet 3: 异常项明细 ─────────────────────────────────
        ws3 = wb.create_sheet("异常项明细")
        headers3 = ["主机名", "IP地址", "检查项", "当前值", "状态", "阈值说明"]
        write_header_row(ws3, headers3)
        set_col_widths(ws3, [18, 16, 20, 20, 10, 30])
        ws3.row_dimensions[1].height = 22

        ri3 = 2
        for h in all_hosts:
            checks = h.get("checks", [])
            bad = [c for c in checks if c.get("status") != "normal"]
            for ck in bad:
                status = ck.get("status", "warning")
                row3 = [
                    h.get("hostname", "-"),
                    h.get("ip", "-"),
                    ck.get("item", "-"),
                    ck.get("value", "-"),
                    STATUS_TEXT.get(status, status),
                    ck.get("threshold", "-"),
                ]
                for ci, val in enumerate(row3, 1):
                    st = status if ci == 5 else None
                    write_data_cell(ws3, ri3, ci, val, st)
                ws3.row_dimensions[ri3].height = 16
                ri3 += 1

        ws3.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"巡检日报_{report_id}.xlsx"
        encoded_fname = quote(filename, encoding="utf-8")
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fname}"},
        )
    except Exception as e:
        logger.exception("Excel 导出失败")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/{report_id}")
async def get_report(report_id: str):
    """获取指定报告详情"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    data = json.loads(p.read_text(encoding="utf-8"))
    return data


# ────────── 通知推送 ──────────

class NotifyRequest(BaseModel):
    channels: list[str]  # ["feishu", "dingtalk"]


@app.post("/api/report/{report_id}/notify")
async def notify_report(report_id: str, body: NotifyRequest):
    """将指定报告推送到飞书 / 钉钉"""
    p = REPORTS_DIR / f"{report_id}.json"
    if not p.exists():
        raise HTTPException(status_code=404, detail="报告不存在")
    report = json.loads(p.read_text(encoding="utf-8"))

    report_url = f"{APP_URL}/report/{report_id}" if APP_URL else ""

    results = {}
    for ch in body.channels:
        if ch == "feishu":
            if not FEISHU_WEBHOOK:
                results["feishu"] = {"ok": False, "msg": "未配置 FEISHU_WEBHOOK"}
            else:
                results["feishu"] = await send_feishu(report, FEISHU_WEBHOOK, keyword=FEISHU_KEYWORD, report_url=report_url)
        elif ch == "dingtalk":
            if not DINGTALK_WEBHOOK:
                results["dingtalk"] = {"ok": False, "msg": "未配置 DINGTALK_WEBHOOK"}
            else:
                results["dingtalk"] = await send_dingtalk(report, DINGTALK_WEBHOOK, keyword=DINGTALK_KEYWORD)
        else:
            results[ch] = {"ok": False, "msg": f"不支持的渠道: {ch}"}

    return {"results": results}


# ────────── CMDB 主机管理 ──────────

@app.get("/api/hosts/inspect")
async def inspect_all_hosts():
    """巡检全量主机，SSE 流式返回数据（不含 AI 分析）"""
    async def generate():
        try:
            results = await prom.inspect_hosts()
            summary = {
                "total": len(results),
                "normal": sum(1 for r in results if r["overall"] == "normal"),
                "warning": sum(1 for r in results if r["overall"] == "warning"),
                "critical": sum(1 for r in results if r["overall"] == "critical"),
            }
            yield f"data: {json.dumps({'type': 'inspect_data', 'data': results, 'summary': summary}, ensure_ascii=False)}\n\n"
        except Exception as exc:
            logger.exception("巡检 SSE 流异常")
            yield f"data: {json.dumps({'type': 'error', 'message': str(exc)}, ensure_ascii=False)}\n\n"
        finally:
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class InspectAIRequest(BaseModel):
    results: list
    summary: dict


@app.post("/api/hosts/inspect/ai")
async def inspect_ai_summary(req: InspectAIRequest):
    """对已有巡检结果做 AI 流式分析（按需调用）"""
    async def generate():
        provider_name = analyzer.provider_name
        yield f"data: {json.dumps({'type': 'ai_meta', 'provider': provider_name, 'fallback': False}, ensure_ascii=False)}\n\n"
        try:
            async for chunk in analyzer.generate_inspection_summary(req.results, req.summary):
                yield f"data: {json.dumps({'type': 'ai_chunk', 'text': chunk}, ensure_ascii=False)}\n\n"
        except Exception as exc:
            logger.exception("巡检 AI 流式总结失败，降级为规则摘要")
            fallback = _build_inspection_fallback_summary(req.results, req.summary, exc)
            yield f"data: {json.dumps({'type': 'ai_meta', 'provider': provider_name, 'fallback': True}, ensure_ascii=False)}\n\n"
            yield f"data: {json.dumps({'type': 'ai_chunk', 'text': fallback}, ensure_ascii=False)}\n\n"
        finally:
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class InspectExcelRequest(BaseModel):
    results: list
    summary: dict
    ai_text: str = ""


@app.post("/api/hosts/inspect/excel")
async def export_inspect_excel(req: InspectExcelRequest):
    """根据传入的巡检结果生成 Excel 文件并下载"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        results  = req.results
        summary  = req.summary
        ai_text  = req.ai_text
        now_str  = datetime.now().strftime("%Y-%m-%d %H:%M")

        wb = Workbook()

        STATUS_COLOR = {"normal": "FF67C23A", "warning": "FFE6A23C", "critical": "FFF56C6C"}
        STATUS_TEXT  = {"normal": "正常", "warning": "警告", "critical": "严重"}

        def hfill(color="FF1F3A5F"):
            return PatternFill("solid", fgColor=color)

        def tborder():
            s = Side(style="thin", color="FFCCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def header_row(ws, headers, row=1):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFFFF", size=10)
                c.fill = hfill()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = tborder()

        def data_cell(ws, row, col, value, status=None):
            c = ws.cell(row=row, column=col, value=value)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = tborder()
            if status and status in STATUS_COLOR:
                c.fill = PatternFill("solid", fgColor=STATUS_COLOR[status])
                c.font = Font(color="FFFFFFFF", size=9)
            else:
                c.font = Font(size=9)

        # ── Sheet 1: 巡检概况 ────────────────────────────────────
        ws1 = wb.active
        ws1.title = "巡检概况"
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 40

        def s1(r, k, v, bold=False):
            ck = ws1.cell(row=r, column=1, value=k)
            cv = ws1.cell(row=r, column=2, value=v)
            ck.font = Font(bold=True, color="FFFFFFFF", size=10)
            ck.fill = hfill("FF243850")
            cv.font = Font(bold=bold, size=10)
            ck.alignment = cv.alignment = Alignment(vertical="center", wrap_text=True)
            ck.border = cv.border = tborder()
            ws1.row_dimensions[r].height = 18

        s1(1, "巡检时间", now_str, bold=True)
        s1(2, "巡检主机总数", summary.get("total", 0))
        s1(3, "正常主机", summary.get("normal", 0))
        s1(4, "警告主机", summary.get("warning", 0))
        s1(5, "严重主机", summary.get("critical", 0))

        # 高频异常项
        issue_cnt: dict[str, int] = {}
        for r in results:
            for c in r.get("checks", []):
                if c.get("status") != "normal":
                    issue_cnt[c.get("item", "未知")] = issue_cnt.get(c.get("item", "未知"), 0) + 1
        top_issues = sorted(issue_cnt.items(), key=lambda x: x[1], reverse=True)[:10]

        ws1.cell(row=7, column=1, value="高频异常项").font = Font(bold=True, size=10)
        header_row(ws1, ["异常检查项", "影响主机数"], row=8)
        for i, (item, cnt) in enumerate(top_issues, 9):
            ws1.cell(row=i, column=1, value=item).border = tborder()
            ws1.cell(row=i, column=2, value=cnt).border = tborder()
            ws1.cell(row=i, column=1).font = Font(size=9)
            ws1.cell(row=i, column=2).font = Font(size=9)
            ws1.cell(row=i, column=2).alignment = Alignment(horizontal="center")

        if ai_text:
            ai_row = 9 + len(top_issues) + 1
            ws1.cell(row=ai_row, column=1, value="AI 分析总结").font = Font(bold=True, size=10)
            c = ws1.cell(row=ai_row + 1, column=1, value=ai_text)
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = tborder()
            ws1.merge_cells(start_row=ai_row+1, start_column=1, end_row=ai_row+1, end_column=2)
            ws1.row_dimensions[ai_row + 1].height = max(60, len(ai_text) // 3)

        # ── Sheet 2: 全部主机明细 ─────────────────────────────────
        ws2 = wb.create_sheet("全部主机明细")
        h2 = ["主机名", "IP地址", "操作系统", "状态", "巡检结果",
              "CPU使用率(%)", "内存使用率(%)", "内存总量(GB)",
              "负载(5m)", "网络收(Mbps)", "网络发(Mbps)",
              "TCP连接数", "运行时长(天)", "磁盘分区详情"]
        header_row(ws2, h2)
        col_widths(ws2, [18, 16, 22, 10, 10, 14, 14, 14, 10, 14, 14, 12, 12, 40])
        ws2.row_dimensions[1].height = 22

        def fmt(v, d=1):
            return round(v, d) if v is not None else "-"

        for ri, h in enumerate(results, 2):
            m = h.get("metrics") or {}
            overall = h.get("overall", "normal")
            us = m.get("uptime_seconds")
            uptime_days = round(us / 86400, 1) if us is not None else "-"
            parts = h.get("partitions") or []
            disk_text = "  ".join(f"{p.get('mountpoint','?')} {p.get('usage_pct','-')}%" for p in parts) or "-"

            row_vals = [
                h.get("hostname") or h.get("instance", "-"),
                h.get("ip", "-"),
                h.get("os", "-"),
                h.get("state", "-"),
                STATUS_TEXT.get(overall, overall),
                fmt(m.get("cpu_usage")),
                fmt(m.get("mem_usage")),
                fmt(m.get("mem_total_gb")),
                fmt(m.get("load5")),
                fmt(m.get("net_recv_mbps")),
                fmt(m.get("net_send_mbps")),
                m.get("tcp_estab") or "-",
                uptime_days,
                disk_text,
            ]
            for ci, val in enumerate(row_vals, 1):
                data_cell(ws2, ri, ci, val, overall if ci == 5 else None)
            ws2.row_dimensions[ri].height = 16
        ws2.freeze_panes = "A2"

        # ── Sheet 3: 异常项明细 ──────────────────────────────────
        ws3 = wb.create_sheet("异常项明细")
        header_row(ws3, ["主机名", "IP地址", "检查项", "当前值", "状态", "阈值说明"])
        col_widths(ws3, [18, 16, 20, 20, 10, 30])
        ws3.row_dimensions[1].height = 22

        ri3 = 2
        for h in results:
            for ck in h.get("checks", []):
                if ck.get("status") == "normal":
                    continue
                st = ck.get("status", "warning")
                data_cell(ws3, ri3, 1, h.get("hostname") or h.get("instance", "-"))
                data_cell(ws3, ri3, 2, h.get("ip", "-"))
                data_cell(ws3, ri3, 3, ck.get("item", "-"))
                data_cell(ws3, ri3, 4, ck.get("value", "-"))
                data_cell(ws3, ri3, 5, STATUS_TEXT.get(st, st), st)
                data_cell(ws3, ri3, 6, ck.get("threshold", "-"))
                ws3.row_dimensions[ri3].height = 16
                ri3 += 1
        ws3.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"巡检报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        encoded_fname = quote(fname, encoding="utf-8")
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fname}"},
        )
    except Exception as e:
        logger.exception("Excel 导出失败")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hosts")
async def get_hosts():
    """获取所有主机列表（Prometheus 发现 + CMDB 补充字段 + 实时指标 + 分区）"""
    try:
        hosts, all_metrics, all_partitions = await asyncio.gather(
            prom.discover_hosts(),
            prom.get_all_host_metrics(),
            prom.get_all_partitions(),
        )
        cmdb = _load_cmdb()
        _cmdb_dirty = False

        for host in hosts:
            inst = host["instance"]
            host["metrics"] = all_metrics.get(inst, {})
            host["partitions"] = all_partitions.get(inst, [])
            # 合并 CMDB 手动字段
            extra = cmdb.get(inst, {})
            host["owner"] = extra.get("owner", "")
            host["env"] = extra.get("env", "")
            host["role"] = extra.get("role", "")
            host["notes"] = extra.get("notes", "")
            # CMDB 中存储的 SSH 凭据（密码不下发到前端，只告知是否已配置）
            host["ssh_port"] = extra.get("ssh_port", 22)
            host["ssh_user"] = extra.get("ssh_user", "")
            host["ssh_saved"] = bool(extra.get("ssh_password"))
            host["credential_id"] = extra.get("credential_id", "")
            # 将真实 IP 写入 CMDB 缓存，供 SSH 连接时使用
            if host["ip"] and host["ip"] != extra.get("ip"):
                cmdb.setdefault(inst, {})["ip"] = host["ip"]
                _cmdb_dirty = True

        if _cmdb_dirty:
            _save_cmdb(cmdb)
        return {"data": hosts, "total": len(hosts), "prometheus_url": PROMETHEUS_URL}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Prometheus 连接失败: {e}")


class HostUpdateRequest(BaseModel):
    owner: Optional[str] = None
    env: Optional[str] = None
    role: Optional[str] = None
    notes: Optional[str] = None
    ssh_port: Optional[int] = None
    ssh_user: Optional[str] = None
    ssh_password: Optional[str] = None  # 明文传入，加密存储
    credential_id: Optional[str] = None  # 关联凭证库中的凭证


@app.put("/api/hosts/{instance:path}")
async def update_host(instance: str, body: HostUpdateRequest):
    """更新主机 CMDB 信息"""
    cmdb = _load_cmdb()
    if instance not in cmdb:
        cmdb[instance] = {}
    for field in ("owner", "env", "role", "notes", "ssh_port", "ssh_user"):
        val = getattr(body, field)
        if val is not None:
            cmdb[instance][field] = val
    # SSH 密码加密存储
    if body.ssh_password is not None:
        if body.ssh_password:
            cmdb[instance]["ssh_password"] = _encrypt_password(body.ssh_password)
        else:
            cmdb[instance].pop("ssh_password", None)  # 空密码 = 清除
    # 关联凭证库
    if body.credential_id is not None:
        if body.credential_id:
            cmdb[instance]["credential_id"] = body.credential_id
            # 使用凭证库时清除独立保存的密码
            cmdb[instance].pop("ssh_password", None)
        else:
            cmdb[instance].pop("credential_id", None)  # 空 = 解绑
    _save_cmdb(cmdb)
    return {"ok": True, "instance": instance}


# ────────── SSH 凭证库 ──────────

class CredentialRequest(BaseModel):
    name: str                          # 凭证名称，如 "生产环境 root"
    username: str = "root"
    password: str                      # 明文传入，加密存储
    port: int = 22


@app.get("/api/ssh/credentials")
async def list_credentials():
    """列出所有凭证（不返回密码）"""
    creds = _load_credentials()
    return {"data": [
        {"id": c["id"], "name": c["name"], "username": c["username"], "port": c["port"]}
        for c in creds
    ]}


@app.post("/api/ssh/credentials")
async def create_credential(body: CredentialRequest):
    """创建凭证"""
    creds = _load_credentials()
    cred_id = f"cred_{int(time.time() * 1000)}"
    creds.append({
        "id": cred_id,
        "name": body.name,
        "username": body.username,
        "password": _encrypt_password(body.password),
        "port": body.port,
    })
    _save_credentials(creds)
    return {"ok": True, "id": cred_id}


@app.put("/api/ssh/credentials/{cred_id}")
async def update_credential(cred_id: str, body: CredentialRequest):
    """更新凭证"""
    creds = _load_credentials()
    for c in creds:
        if c["id"] == cred_id:
            c["name"] = body.name
            c["username"] = body.username
            c["port"] = body.port
            if body.password:
                c["password"] = _encrypt_password(body.password)
            _save_credentials(creds)
            return {"ok": True}
    raise HTTPException(status_code=404, detail="凭证不存在")


@app.delete("/api/ssh/credentials/{cred_id}")
async def delete_credential(cred_id: str):
    """删除凭证"""
    creds = _load_credentials()
    creds = [c for c in creds if c["id"] != cred_id]
    _save_credentials(creds)
    # 清除引用该凭证的主机
    cmdb = _load_cmdb()
    changed = False
    for inst, entry in cmdb.items():
        if entry.get("credential_id") == cred_id:
            entry.pop("credential_id", None)
            changed = True
    if changed:
        _save_cmdb(cmdb)
    return {"ok": True}


# ────────── 巡检 ──────────

def _inspection_issue_counts(results: list[dict]) -> list[tuple[str, int]]:
    counter: dict[str, int] = {}
    for result in results:
        for check in result.get("checks", []):
            if check.get("status") == "normal":
                continue
            item = check.get("item", "未知项")
            counter[item] = counter.get(item, 0) + 1
    return sorted(counter.items(), key=lambda item: item[1], reverse=True)


def _build_inspection_fallback_summary(results: list[dict], summary: dict, error: Exception | None = None) -> str:
    total = summary.get("total", len(results))
    normal = summary.get("normal", 0)
    warning = summary.get("warning", 0)
    critical = summary.get("critical", 0)
    issue_counts = _inspection_issue_counts(results)
    top_issue_text = "、".join(
        f"{item}（{count}台）" for item, count in issue_counts[:3]
    ) or "暂未发现重复性异常项"

    abnormal_hosts = [r for r in results if r.get("overall") != "normal"]
    abnormal_hosts.sort(
        key=lambda item: (
            {"critical": 2, "warning": 1, "normal": 0}.get(item.get("overall", "normal"), 0),
            sum(1 for c in item.get("checks", []) if c.get("status") != "normal"),
        ),
        reverse=True,
    )
    top_host_text = "；".join(
        f"{host.get('hostname') or host.get('instance')}({host.get('ip', '-')})"
        for host in abnormal_hosts[:3]
    ) or "暂无异常主机"

    parts = [
        f"本次巡检共覆盖 {total} 台主机，其中正常 {normal} 台、警告 {warning} 台、严重 {critical} 台。",
        f"当前最集中的异常项为：{top_issue_text}。",
        f"需优先关注的主机：{top_host_text}。",
        "建议先处理严重主机，再按高频异常项排查资源瓶颈、容量风险和连接状态波动。",
        "未来24小时若不处理，异常项可能继续扩散到更多主机，并导致性能抖动、告警增加或服务稳定性下降。",
    ]
    if error:
        parts.append(f"AI 服务暂不可用，当前显示的是规则摘要（{error}）。")
    return "\n\n".join(parts)


async def _generate_inspection_summary(results: list[dict], summary: dict) -> tuple[str, str, bool, str]:
    provider_name = analyzer.provider_name
    try:
        chunks = []
        async for chunk in analyzer.generate_inspection_summary(results, summary):
            chunks.append(chunk)
        content = "".join(chunks).strip()
        if not content:
            raise RuntimeError("AI 返回内容为空")
        return content, provider_name, False, ""
    except Exception as exc:
        logger.exception("巡检 AI 总结生成失败")
        fallback = _build_inspection_fallback_summary(results, summary, exc)
        return fallback, provider_name, True, str(exc)


async def _build_inspection_response(results: list[dict]) -> dict:
    summary = {
        "total": len(results),
        "normal": sum(1 for r in results if r["overall"] == "normal"),
        "warning": sum(1 for r in results if r["overall"] == "warning"),
        "critical": sum(1 for r in results if r["overall"] == "critical"),
    }
    ai_summary, ai_provider, ai_fallback, ai_error = await _generate_inspection_summary(results, summary)
    return {
        "data": results,
        "summary": summary,
        "ai_summary": ai_summary,
        "ai_provider": ai_provider,
        "ai_fallback": ai_fallback,
        "ai_error": ai_error,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/hosts/{instance:path}/inspect")
async def inspect_single_host(instance: str):
    """巡检单台主机"""
    try:
        results = await prom.inspect_hosts(instances=[instance])
        if not results:
            raise HTTPException(status_code=404, detail="主机未找到")
        return results[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=str(e))


# ────────── SSH WebSocket ──────────

@app.websocket("/api/ws/ssh")
async def ws_ssh(ws: WebSocket):
    """WebSocket SSH 终端代理（凭据通过首条 WebSocket 消息传递，不暴露在 URL 中）"""

    def _resolve_credential(instance: str, credential_id: str = ""):
        """从 CMDB 或凭证库中解密已保存的 SSH 凭证"""
        cmdb = _load_cmdb()
        entry = cmdb.get(instance, {}) if instance else {}
        host = entry.get("ip", instance.split(":")[0]) if instance else ""

        # 直接指定凭证 ID（SSH toolbar 直选）或 CMDB 绑定的凭证 ID
        cred_id = credential_id or entry.get("credential_id", "")
        if cred_id:
            creds = _load_credentials()
            cred = next((c for c in creds if c["id"] == cred_id), None)
            if cred:
                try:
                    password = _decrypt_password(cred["password"])
                except Exception:
                    return None
                return {
                    "host": host,
                    "port": cred.get("port", 22),
                    "username": cred.get("username", "root"),
                    "password": password,
                }

        # 回退到主机独立保存的密码
        cipher = entry.get("ssh_password", "")
        if not cipher:
            return None
        try:
            password = _decrypt_password(cipher)
        except Exception:
            return None
        return {
            "host": host,
            "port": entry.get("ssh_port", 22),
            "username": entry.get("ssh_user", "root"),
            "password": password,
        }

    await ssh_websocket_handler(ws, resolve_credential=_resolve_credential)


# ────────── 健康检查 ──────────

@app.get("/api/health")
async def health():
    async def _check_loki():
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(5.0)) as c:
                kw = {"url": f"{LOKI_URL}/loki/api/v1/labels"}
                if LOKI_USERNAME:
                    kw["auth"] = (LOKI_USERNAME, LOKI_PASSWORD)
                resp = await c.get(**kw)
                return resp.status_code == 200
        except Exception as e:
            logger.warning("[health] Loki 连接失败: %s", e)
            return False

    async def _check_prom():
        try:
            await prom.query("up", timeout=5)
            return True
        except Exception as e:
            logger.warning("[health] Prometheus 连接失败: %s", e)
            return False

    # 并行检测，总耗时 ≤ 5 秒
    loki_ok, prom_ok = await asyncio.gather(_check_loki(), _check_prom())

    try:
        ai_name = analyzer.provider_name
        ai_ok = True
    except Exception as e:
        ai_name = str(e)
        ai_ok = False

    return {
        "status": "ok",
        "version": "2.0",
        "loki_connected": loki_ok,
        "loki_url": LOKI_URL,
        "prometheus_connected": prom_ok,
        "prometheus_url": PROMETHEUS_URL,
        "ai_provider": ai_name,
        "ai_ready": ai_ok,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host=os.getenv("APP_HOST", "0.0.0.0"),
        port=int(os.getenv("APP_PORT", "8000")),
        reload=os.getenv("DEV_RELOAD", "").lower() in ("1", "true", "yes"),
    )
